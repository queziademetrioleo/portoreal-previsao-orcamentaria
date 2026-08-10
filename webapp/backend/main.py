#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PREVISAO ORCAMENTARIA — Backend Web (FastAPI)

Fluxo human-in-the-loop:
  1. POST /api/sessao            — upload dos 4 relatorios -> analise (regras + IA)
  2. GET  /api/sessao/{id}       — estado da sessao (itens p/ revisao)
  3. POST /api/sessao/{id}/gerar — decisoes humanas -> xlsx final (estrutura identica ao manual)

Dev:  uvicorn main:app --reload --port 8000
Prod: ver Dockerfile na raiz do projeto.
"""
import os
import re
import sys
import json
import uuid
import asyncio
import copy
import threading
import tempfile
import datetime
import logging
import openpyxl

from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

import db
import previsao as core
from gerador_previsao import gerar_previsao_adaptativa
from relatorio_pdf import gerar_relatorio_pdf

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

app = FastAPI(title='Previsao Orcamentaria', version='2.0')

_cors_origins = os.environ.get('CORS_ORIGINS', '*').split(',')
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in _cors_origins if o.strip()] or ['*'],
    allow_methods=['*'],
    allow_headers=['*'],
)

ARQUIVOS_ESPERADOS = {
    'balanual': 'balanual.xls',
    'desbai': 'desbai06.xls',
    'dessin': 'dessin02.xls',
    'inad': 'inad01.xls',
    'rec': 'rec02.xls',
}


# ---------------------------------------------------------------------------
# Helpers de sessao  (persistencia via MySQL)
# ---------------------------------------------------------------------------

def _carregar_estado(sid):
    """Carrega o estado JSON da sessao do MySQL."""
    row = db.carregar_sessao(sid)
    if not row or not row.get('estado_json'):
        raise HTTPException(404, 'Sessao nao encontrada')
    return json.loads(row['estado_json'])


def _json_dumps(obj):
    """Serializa objeto para JSON, convertendo chaves tupla em string."""
    def _converter(o):
        if isinstance(o, dict):
            return {str(k) if isinstance(k, tuple) else k: _converter(v) for k, v in o.items()}
        if isinstance(o, list):
            return [_converter(v) for v in o]
        if isinstance(o, tuple):
            return str(o)
        return o
    return json.dumps(_converter(obj), ensure_ascii=False, default=str)


def _json_loads(s):
    """Carrega JSON, convertendo chaves que parecem tuplas de volta para tupla."""
    import ast
    def _converter(o):
        if isinstance(o, dict):
            result = {}
            for k, v in o.items():
                if isinstance(k, str) and k.startswith('(') and k.endswith(')'):
                    try:
                        k = ast.literal_eval(k)
                    except (ValueError, SyntaxError):
                        pass
                result[k] = _converter(v)
            return result
        if isinstance(o, list):
            return [_converter(v) for v in o]
        return o
    return _converter(json.loads(s))


def _obter_R(sid):
    """Retorna o cache de analise (R) do MySQL; fallback para core.analisar()."""
    row = db.carregar_sessao(sid)
    if row is None:
        raise HTTPException(404, 'Sessao nao encontrada')
    if row.get('cache_analise'):
        try:
            return _json_loads(row['cache_analise'])
        except (json.JSONDecodeError, TypeError):
            pass
    # Fallback: reconstituir arquivos do MySQL e re-analisar
    with tempfile.TemporaryDirectory() as tmpdir:
        for chave, fname in ARQUIVOS_ESPERADOS.items():
            content = db.obter_arquivo(sid, chave)
            if content:
                with open(os.path.join(tmpdir, fname), 'wb') as f:
                    f.write(content)
        R = core.analisar(tmpdir)
        db.salvar_cache_analise(sid, _json_dumps(R))
        return R


def _decisao_payload(decisoes, item_id):
    raw = decisoes.get(str(item_id))
    if isinstance(raw, dict):
        return raw.get('decisao'), raw.get('valor'), raw.get('nota')
    return raw, None, None


def _aplicar_decisao_editavel(item, decisao, valor, nota, permitidas):
    if decisao in permitidas:
        item['decisao'] = decisao
    if valor is not None and valor != '':
        try:
            item['valor_editado'] = round(float(valor), 2)
        except (TypeError, ValueError):
            pass
    if nota is not None:
        item['nota'] = str(nota).strip()[:1000]


def _valor_revisado(item):
    try:
        return float(item.get('valor_editado', item.get('valor', 0)) or 0)
    except (TypeError, ValueError):
        return float(item.get('valor', 0) or 0)


# Traduz o motivo tecnico (vindo de classify()/outliers/IA) em motivos curtos
# e objetivos, na ordem em que fazem sentido para leitura. Cada regra devolve
# no maximo 1 frase; a primeira que bater e usada.
_MOTIVOS_PADROES = [
    (re.compile(r'obras?|benfeitoria', re.I), 'é uma obra ou reforma'),
    (re.compile(r'rescis|indeniza', re.I), 'é rescisão ou indenização de funcionário'),
    (re.compile(r'reparo|conserto', re.I), 'foi um conserto pontual, fora da rotina'),
    (re.compile(r'outlier|\bmad\b', re.I), 'o valor ficou bem acima do normal para essa conta'),
    (re.compile(r'valor alto', re.I), 'o valor ficou bem mais alto que o de costume'),
    (re.compile(r'capital', re.I), 'a descrição indica gasto de obra/capital'),
    (re.compile(r'sem regra explicita', re.I), 'não segue um padrão claro nas contas do condomínio'),
    (re.compile(r'periodic|ambigu', re.I), 'pode ou não se repetir — fica em revisão'),
    (re.compile(r'recorrente', re.I), 'é um gasto do dia a dia do condomínio'),
]


def _motivos_legiveis(motivo, n_meses):
    """Extrai motivos curtos e objetivos a partir do texto tecnico do motor.
    Sempre retorna pelo menos 1 item (fallback por frequencia ou generico)."""
    motivos = []
    texto = str(motivo or '')
    if texto.startswith('IA:'):
        texto_ia = texto[3:].strip()
        if texto_ia:
            motivos.append(texto_ia.rstrip('.'))
    else:
        for padrao, frase in _MOTIVOS_PADROES:
            if padrao.search(texto):
                motivos.append(frase)
                break
    if n_meses is not None and n_meses <= 2 and not any('rotina' in m or 'repetir' in m for m in motivos):
        motivos.append(f'só apareceu em {n_meses} de 12 meses')
    if not motivos:
        motivos.append('identificado pela análise como fora do padrão de gasto recorrente')
    return motivos


def _explicacao_deterministica_despesa(item):
    decisao = item.get('decisao')
    motivos = _motivos_legiveis(item.get('motivo'), item.get('n_meses'))
    if decisao == 'aprovada':
        resumo = f"Removido por motivo{'s' if len(motivos) > 1 else ''} de: {', '.join(motivos)}."
    elif decisao == 'reprovada':
        resumo = f"Mantido por motivo{'s' if len(motivos) > 1 else ''} de: {', '.join(motivos)}."
    else:
        resumo = 'Item aguardando decisão humana.'
    evidencias = []
    origem = item.get('origem')
    if origem:
        evidencias.append(f'Origem da classificação: {origem}.')
    if item.get('n_meses') is not None:
        evidencias.append(f'Frequência observada: {item.get("n_meses")}/12 meses.')
    if item.get('motivo'):
        evidencias.append(f'Motivo técnico: {item.get("motivo")}.')
    if item.get('nota'):
        evidencias.append(f'Nota humana: {item.get("nota")}.')
    return {'resumo': resumo, 'evidencias': evidencias}


def _explicacao_deterministica_inad(item):
    critica = bool(item.get('critica'))
    meses = item.get('meses_atraso', 0)
    if item.get('decisao') == 'abater':
        resumo = (
            f'Abatido da receita por motivo de: {meses} mês(es) consecutivos em atraso, '
            'o que reduz o quanto o condomínio deve efetivamente receber.'
        )
    else:
        resumo = (
            f'Ignorado por motivo de: atraso de {meses} mês(es), abaixo do limite '
            'considerado crítico (3 meses seguidos).'
        )
    evidencias = [
        'Inadimplência crítica (3+ meses seguidos).' if critica else 'Inadimplência recente.',
        f'{meses} mês(es) em atraso.',
    ]
    if item.get('nota'):
        evidencias.append(f'Nota humana: {item.get("nota")}.')
    return {'resumo': resumo, 'evidencias': evidencias}


def _enriquecer_explicacoes_ia(estado):
    """Gera explicacoes curtas e humanas para o relatório final.

    A IA é opcional: se não houver chave/resposta válida, mantemos explicações
    determinísticas para não bloquear a geração.
    """
    despesas = [
        item for item in (estado.get('extraordinarias', []) + estado.get('revisar', []))
        if item.get('decisao') != 'pendente'
    ]
    inad = estado.get('inadimplencia', [])
    for item in despesas:
        item['explicacao'] = _explicacao_deterministica_despesa(item)
    for item in inad:
        item['explicacao'] = _explicacao_deterministica_inad(item)

    if not core._ia_disponivel() or not despesas:
        return

    payload = []
    for item in despesas[:80]:
        payload.append({
            'id': item['id'],
            'decisao': 'removido' if item.get('decisao') == 'aprovada' else 'mantido',
            'grupo': item.get('grupo'),
            'classe': item.get('classe'),
            'descricao': item.get('descricao'),
            'valor': _valor_revisado(item),
            'frequencia_12m': item.get('n_meses'),
            'motivo_tecnico': item.get('motivo'),
            'nota_humana': item.get('nota'),
            'origem': item.get('origem'),
        })
    sistema = (
        'Você é um analista financeiro de condomínios. Explique decisões de revisão '
        'orçamentária para um síndico idoso, com linguagem simples, respeitosa e direta. '
        'Não invente fatos. Use apenas os dados recebidos.'
    )
    user = (
        'Para cada item, retorne JSON no formato '
        '{"itens":[{"id":1,"resumo":"frase curta","evidencias":["ponto 1","ponto 2"]}]}. '
        'Explique por que foi removido ou mantido, destacando pontualidade, frequência, '
        'parcelas, regra/IA e nota humana quando houver.\n\n'
        f'Itens: {json.dumps(payload, ensure_ascii=False)}'
    )
    try:
        resp = core._claude_chat(sistema, user, max_tokens=5000, temperature=0)
        if not resp:
            return
        data = core._extrai_json(resp)
        por_id = {int(i.get('id')): i for i in data.get('itens', []) if i.get('id') is not None}
        for item in despesas:
            exp = por_id.get(int(item['id']))
            if not exp:
                continue
            resumo = str(exp.get('resumo') or '').strip()
            evidencias = exp.get('evidencias') if isinstance(exp.get('evidencias'), list) else []
            evidencias = [str(e).strip() for e in evidencias if str(e).strip()][:5]
            if resumo:
                item['explicacao'] = {'resumo': resumo[:500], 'evidencias': evidencias}
    except Exception as exc:
        logger.warning('Falha ao gerar explicacoes IA: %s', exc)


def _fluxo_mensal_balanco(bal):
    meses = bal.get('meses') or []
    n = len(meses)
    out = []
    for i in range(n):
        receita = sum(
            float((ln.get('monthly') or [0] * n)[i] or 0)
            for ln in bal.get('receitas', [])
            if i < len(ln.get('monthly') or [])
        )
        despesa = sum(
            float((ln.get('monthly') or [0] * n)[i] or 0)
            for ln in bal.get('despesas', [])
            if i < len(ln.get('monthly') or [])
        )
        out.append({
            'mes': str(meses[i]),
            'receita': round(receita, 2),
            'despesa': round(despesa, 2),
            'saldo': round(receita - despesa, 2),
        })
    return out


def _aplicar_decisoes(estado, dec):
    """Aplica as decisoes humanas no estado (in-place)."""
    if getattr(dec, 'inflacao_pct', None) is not None:
        estado['resumo']['inflacao'] = float(dec.inflacao_pct)
    if getattr(dec, 'ultimo_reajuste', None) is not None:
        estado['resumo']['ultimo_reajuste'] = dec.ultimo_reajuste or None
    for item in estado['extraordinarias']:
        d, valor, nota = _decisao_payload(dec.extraordinarias, item['id'])
        _aplicar_decisao_editavel(item, d, valor, nota, ('aprovada', 'reprovada'))
    for item in estado['revisar']:
        d, valor, nota = _decisao_payload(dec.revisar, item['id'])
        _aplicar_decisao_editavel(item, d, valor, nota, ('aprovada', 'reprovada', 'pendente'))
    for item in estado['inadimplencia']:
        d, valor, nota = _decisao_payload(dec.inadimplencia, item['id'])
        _aplicar_decisao_editavel(item, d, valor, nota, ('abater', 'ignorar'))


def _recalcular_com_decisoes(sid, estado):
    """Recalcula subtotal/total/impacto a partir das decisoes ja no 'estado'.
    Retorna (R2, impacto_receita). Nao gera xlsx."""
    ids_remover = {i['id'] for i in estado['extraordinarias'] if i['decisao'] == 'aprovada'}
    ids_remover |= {i['id'] for i in estado['revisar'] if i['decisao'] == 'aprovada'}

    R = copy.deepcopy(_obter_R(sid))
    R['inflacao_pct'] = float(estado['resumo'].get('inflacao') or core.INFLACAO)
    valores_editados = {
        i['id']: _valor_revisado(i)
        for i in (estado['extraordinarias'] + estado['revisar'])
        if 'valor_editado' in i
    }
    for idx, it in enumerate(R['des']['itens']):
        if idx in valores_editados:
            it['valor_pago'] = valores_editados[idx]
            it['valor_lcto'] = valores_editados[idx]
        it['cat'] = 'Extraordinaria' if idx in ids_remover else (
            'Recorrente' if it['cat'] in ('Extraordinaria', 'Revisar') else it['cat'])
    R2 = core.recalcular(R)

    unidades = {}
    for item in estado['inadimplencia']:
        if item['decisao'] == 'abater':
            unidades.setdefault(item['unidade'], []).append(_valor_revisado(item))
    impacto = sum(sum(v) / len(v) for v in unidades.values())
    return R2, impacto


def _extrair_previsao_final(path):
    """Extrai a área visível da aba PREVISÃO (2), fonte oficial do resultado."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        nn = core._norm(name).replace(' ', '')
        if 'previs' in nn and '(2)' in name:
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:
            nn = core._norm(name).replace(' ', '')
            if 'previs' in nn:
                ws = wb[name]
                break
    if ws is None:
        return []

    rows = []
    for r in range(9, 56):
        label = ws.cell(r, 3).value
        anual = ws.cell(r, 4).value
        rateio = ws.cell(r, 5).value
        mensal = ws.cell(r, 6).value
        if label is None and anual is None and rateio is None and mensal is None:
            continue
        rows.append({
            'row': r,
            'label': str(label).strip() if label is not None else '',
            'anual': round(float(anual), 2) if isinstance(anual, (int, float)) else anual,
            'rateio': round(float(rateio), 2) if isinstance(rateio, (int, float)) else rateio,
            'mensal': round(float(mensal), 2) if isinstance(mensal, (int, float)) else mensal,
        })
    return rows


def _montar_estado(sid, nome, ano, R):
    """Converte o resultado de core.analisar() no payload de revisao humana."""
    des = R['des']

    extraordinarias, revisar = [], []
    for idx, it in enumerate(des['itens']):
        item = {
            'id': idx,
            'grupo': it['grupo'],
            'classe': it['classe'],
            'data': str(it['data'] or ''),
            'descricao': it['descricao'],
            'valor': round(it['valor_pago'], 2),
            'motivo': it['motivo'],
            'n_meses': it.get('n_meses'),
        }
        if it['cat'] == 'Extraordinaria':
            ia = R.get('sugestoes_ia', {}).get(idx) or R.get('sugestoes_ia', {}).get(str(idx))
            item['origem'] = 'IA' if (it['motivo'] or '').startswith('IA:') else 'Regra'
            item['decisao'] = 'aprovada'        # default: remover da base
            extraordinarias.append(item)
        elif it['cat'] == 'Revisar':
            item['origem'] = 'IA' if (it['motivo'] or '').startswith('IA:') else 'Regra'
            item['decisao'] = 'pendente'        # humano decide
            revisar.append(item)

    inad_itens = []
    if R['inad']:
        # Calcular a ultima parcela (mes_ref mais recente) por unidade
        ultima_parcela_por_unidade = {}
        for it in R['inad']['itens']:
            u = it['unidade']
            mes = it.get('mes_ref', '')
            atual = ultima_parcela_por_unidade.get(u, '')
            if mes > atual:
                ultima_parcela_por_unidade[u] = mes

        for i, it in enumerate(R['inad']['itens']):
            critica = it.get('critica', (it['meses_atraso'] or 0) >= 3)
            inad_itens.append({
                'id': i,
                'unidade': it['unidade'],
                'classe': it['classe'],
                'mes_ref': it['mes_ref'],
                'vencimento': str(it['vencimento'] or ''),
                'valor': round(it['valor'], 2),
                'meses_atraso': it['meses_atraso'],
                'critica': critica,
                'decisao': 'abater' if critica else 'ignorar',
                'ultima_parcela': ultima_parcela_por_unidade.get(it['unidade'], it.get('mes_ref', '')),
            })

    linhas = [{
        'grupo': l['grupo'], 'classe': l['classe'],
        'base': round(l['base'], 2), 'deducao': round(l['deducao'], 2),
        'final': round(l['final'], 2), 'regra': l['regra'],
        'n_meses': l['n_meses'],
    } for l in R['linhas']]

    bal = R['bal']
    return {
        'sessao_id': sid,
        'nome_condominio': nome,
        'ano_previsao': ano,
        'criado_em': datetime.datetime.now().isoformat(timespec='seconds'),
        'modelo_ia': core._ia_modelo(),
        'ia_ativa': core._ia_disponivel(),
        'resumo': {
            'base_total': round(R['base_total'], 2),
            'desconsideracoes': round(R['desconsideracoes'], 2),
            'prov_laudo': round(R['prov_laudo'], 2),
            'prov_incendio': round(R['prov_incendio'], 2),
            'subtotal': round(R['subtotal'], 2),
            'inflacao': core.INFLACAO,
            'total_previsto': round(R['total_previsto'], 2),
            'receita_anual': round(R.get('receita_anual') or 0, 2),
            'receita_mensal': round((R.get('receita_anual') or 0) / 12, 2),
            'rec_mes_ref': (R.get('rec') or {}).get('mes_ref'),
            'n_meses_balanco': bal.get('n_meses'),
            'ultimo_reajuste': None,  # 'AAAA-MM', preenchido na revisao p/ o relatorio PDF
            'cenarios': R.get('cenarios'),
            'periodo': [str(R['des']['periodo'][0]), str(R['des']['periodo'][1])],
        },
        'extraordinarias': extraordinarias,
        'revisar': revisar,
        'inadimplencia': inad_itens,
        'inad_meta': ({
            'total': round(R['inad']['total'], 2),
            'critica': round(R['inad']['critica'], 2),
            'data_base': str(R['inad']['data_base']),
        } if R['inad'] else None),
        'linhas_contas': linhas,
        'previsao_final': [],
        'fluxo_mensal': _fluxo_mensal_balanco(bal),
        'status': 'em_revisao',
    }


# ---------------------------------------------------------------------------
# Modelo de decisoes
# ---------------------------------------------------------------------------
class Decisoes(BaseModel):
    extraordinarias: dict = Field(default_factory=dict)
    revisar: dict = Field(default_factory=dict)
    inadimplencia: dict = Field(default_factory=dict)
    # Fracao (ex.: 0.10). None = manter o valor atual da sessao.
    inflacao_pct: float | None = Field(default=None, ge=0.0, le=1.0)
    # 'AAAA-MM' (ex.: '2022-03') — mes/ano do ultimo reajuste da taxa
    # condominial, usado no item 6 do relatorio PDF. None = nao informado.
    ultimo_reajuste: str | None = Field(default=None, max_length=7)


# ---------------------------------------------------------------------------
# Startup — limpeza de sessoes antigas
# ---------------------------------------------------------------------------
@app.on_event("startup")
async def startup():
    removidas = db.limpar_sessoes_antigas(90)
    if removidas:
        logger.info(f'{removidas} sessoes antigas removidas (TTL=90d)')


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------
@app.get("/api/health")
async def health():
    db_ok = db.verificar_conexao()
    return {"status": "ok" if db_ok else "degraded", "db": "connected" if db_ok else "disconnected"}


# ---------------------------------------------------------------------------
# 1) UPLOAD + ANALISE
# ---------------------------------------------------------------------------
MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # 20 MB por arquivo

@app.post('/api/sessao')
async def criar_sessao(
    nome_condominio: str = Form(..., max_length=200),
    ano_previsao: int = Form(..., ge=2000, le=2100),
    balanual: UploadFile = File(...),
    desbai: UploadFile = File(...),
    rec: UploadFile = File(...),
    dessin: UploadFile = File(None),
    inad: UploadFile = File(None),
):
    # Nome agora é opcional — o sistema detecta automaticamente do arquivo REC
    # (feedback CEO 07/2026). Se vazio, usamos placeholder que será substituído
    # na análise SSE assim que o REC for processado.
    nome_condominio = (nome_condominio or '').strip() or '(Detectando...)'

    sid = uuid.uuid4().hex[:12]

    # Criar registro PRIMEIRO (INSERT), depois salvar arquivos (UPDATE)
    db.criar_sessao(sid, nome_condominio.strip(), ano_previsao)

    uploads = {
        'balanual': balanual,
        'desbai': desbai,
        'rec': rec,
        'dessin': dessin,
        'inad': inad,
    }
    file_bytes = {}
    for chave, up in uploads.items():
        if up is None:
            continue
        # Validar extensao — xlrd so suporta .xls (BIFF)
        if up.filename and not up.filename.lower().endswith('.xls'):
            db.deletar_sessao(sid)
            raise HTTPException(400, f'Arquivo {up.filename}: apenas .xls e suportado. Converta .xlsx para .xls antes de enviar.')
        # Validar tamanho
        if up.size and up.size > MAX_UPLOAD_BYTES:
            db.deletar_sessao(sid)
            raise HTTPException(413, f'Arquivo {up.filename} excede o limite de 20 MB.')
        conteudo = await up.read()
        file_bytes[chave] = conteudo
        db.salvar_arquivo(sid, chave, conteudo)

    if 'balanual' not in file_bytes or 'desbai' not in file_bytes or 'rec' not in file_bytes:
        db.deletar_sessao(sid)
        raise HTTPException(400, 'Arquivos obrigatorios ausentes (balanual, desbai06 e REC)')
    logger.info(f'Sessao {sid} criada: {nome_condominio.strip()} ({ano_previsao})')

    return {'sessao_id': sid, 'nome_condominio': nome_condominio.strip(),
            'ano_previsao': ano_previsao, 'status': 'pendente'}


@app.get('/api/sessao/{sid}/analisar')
async def analisar_sse(sid: str):
    """SSE endpoint: executa analise e transmite progresso em tempo real."""
    from fastapi.responses import StreamingResponse

    row = db.carregar_sessao(sid)
    if not row:
        raise HTTPException(404, 'Sessao nao encontrada')

    eventos = []
    lock = threading.Lock()
    cancelado = threading.Event()  # sinaliza cancelamento para a thread de analise

    def on_progress(p):
        if cancelado.is_set():
            return
        with lock:
            eventos.append(p)

    async def gerar():
        loop = asyncio.get_event_loop()

        try:
            # Reconstitui arquivos do MySQL para diretorio temporario
            with tempfile.TemporaryDirectory() as tmpdir:
                for chave, fname in ARQUIVOS_ESPERADOS.items():
                    content = db.obter_arquivo(sid, chave)
                    if content:
                        with open(os.path.join(tmpdir, fname), 'wb') as f:
                            f.write(content)

                # Inicia analise em thread
                future = loop.run_in_executor(None, core.analisar, tmpdir, on_progress)

                # Stream progress enquanto analise roda
                last_idx = 0
                while not future.done():
                    await asyncio.sleep(0.3)
                    with lock:
                        while last_idx < len(eventos):
                            ev = eventos[last_idx]
                            last_idx += 1
                            yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
                    yield ": keepalive\n\n"  # mantem conexao viva durante fases longas (IA, parsing)

                # Pega ultimos eventos
                with lock:
                    while last_idx < len(eventos):
                        ev = eventos[last_idx]
                        last_idx += 1
                        yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"

                # Resultado final
                R = future.result()
                db.salvar_cache_analise(sid, _json_dumps(R))

                nome = row['nome_condominio']
                ano = row['ano_previsao']

                # Nome automatico (feedback CEO 07/2026): extrai do REC se
                # disponivel e atualiza a sessao no banco.
                rec_nome = (R.get('rec') or {}).get('nome_condominio')
                if rec_nome and rec_nome.strip():
                    nome = rec_nome.strip()
                    db.atualizar_nome_condominio(sid, nome)

                estado = _montar_estado(sid, nome, ano, R)
                _salvar_estado_sync(sid, estado)

                yield f"data: {json.dumps({'done': True, 'sessao_id': sid}, ensure_ascii=False)}\n\n"

        except asyncio.CancelledError:
            cancelado.set()
            logger.warning('SSE cancelado pelo cliente — sessao %s', sid)
            # Tenta cancelar a future se ainda nao terminou
            if not future.done():
                future.cancel()
        except Exception as exc:
            logger.error('Erro na analise SSE da sessao %s: %s', sid, exc)
            yield f"data: {json.dumps({'error': str(exc)}, ensure_ascii=False)}\n\n"

    return StreamingResponse(gerar(), media_type='text/event-stream',
                             headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


def _salvar_estado_sync(sid, estado):
    db.salvar_estado(sid, json.dumps(estado, ensure_ascii=False, default=str))


# ---------------------------------------------------------------------------
# 2) CONSULTA
# ---------------------------------------------------------------------------
@app.get('/api/sessao/{sid}')
def obter_sessao(sid: str):
    return _carregar_estado(sid)


@app.delete("/api/sessao/{sid}")
def deletar_sessao(sid: str):
    """Remove uma sessao e todos os seus dados."""
    try:
        db.deletar_sessao(sid)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, f"Erro ao deletar sessao: {e}")


@app.get('/api/sessoes')
def listar_sessoes():
    sessoes = db.listar_sessoes()
    return [{
        'sessao_id': s['id'],
        'nome': s['nome_condominio'],
        'ano': s['ano_previsao'],
        'criado_em': str(s['criado_em']),
        'status': s.get('status', 'em_revisao'),
    } for s in sessoes]


# ---------------------------------------------------------------------------
# 3) DECISOES + GERACAO DO DOCUMENTO FINAL
# ---------------------------------------------------------------------------
@app.post('/api/sessao/{sid}/preview')
def preview(sid: str, dec: Decisoes):
    """Recalcula subtotal/total/impacto com as decisoes atuais SEM gerar o xlsx.
    Permite que a interface atualize os numeros ao vivo a cada clique."""
    estado = _carregar_estado(sid)
    _aplicar_decisoes(estado, dec)
    R2, impacto = _recalcular_com_decisoes(sid, estado)
    return {'ok': True,
            'subtotal': round(R2['subtotal'], 2),
            'total_previsto': round(R2['total_previsto'], 2),
            'impacto_receita_mensal': round(impacto, 2),
            'inflacao': R2.get('inflacao_pct'),
            'cenarios': R2.get('cenarios')}


@app.post('/api/sessao/{sid}/gerar')
def gerar(sid: str, dec: Decisoes):
    estado = _carregar_estado(sid)

    _aplicar_decisoes(estado, dec)
    pendentes = [
        item for item in (estado['extraordinarias'] + estado['revisar'])
        if item.get('decisao') == 'pendente'
    ]
    if pendentes:
        raise HTTPException(
            400,
            f'Existem {len(pendentes)} itens pendentes de revisão. '
            'Decida se cada item deve ser removido ou mantido antes de gerar.'
        )
    _enriquecer_explicacoes_ia(estado)
    R2, impacto_receita = _recalcular_com_decisoes(sid, estado)

    # Gerar xlsx em arquivo temporario e salvar bytes no MySQL
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        out_path = tmp.name
    try:
        modelo = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'templates', 'modelo_previsao.xlsx')
        gerar_previsao_adaptativa(
            destino=out_path,
            R=R2,
            nome_condominio=estado['nome_condominio'],
            ano=estado['ano_previsao'],
            inflacao=float(estado['resumo'].get('inflacao') or core.INFLACAO),
            impacto_receita_mensal=impacto_receita,
            inad_detalhe=estado['inadimplencia'],
            inad_meta=estado['inad_meta'],
            referencia=modelo if os.path.exists(modelo) else None,
        )
        previsao_final = _extrair_previsao_final(out_path)
        with open(out_path, 'rb') as f:
            xlsx_bytes = f.read()
    finally:
        os.unlink(out_path)

    db.salvar_arquivo(sid, 'xlsx', xlsx_bytes)

    estado['status'] = 'gerado'
    estado['resumo']['subtotal'] = round(R2['subtotal'], 2)
    estado['resumo']['total_previsto'] = round(R2['total_previsto'], 2)
    estado['resumo']['impacto_receita_mensal'] = round(impacto_receita, 2)
    estado['resumo']['cenarios'] = R2.get('cenarios')
    estado['previsao_final'] = previsao_final
    db.salvar_estado(sid, json.dumps(estado, ensure_ascii=False, default=str))

    return {'ok': True, 'sessao_id': sid,
            'subtotal': round(R2['subtotal'], 2),
            'total_previsto': round(R2['total_previsto'], 2),
            'impacto_receita_mensal': round(impacto_receita, 2),
            'download': f'/api/sessao/{sid}/download'}


@app.post('/api/sessao/{sid}/salvar-decisoes')
def salvar_decisoes(sid: str, decisoes: Decisoes):
    """Salva decisoes parciais do usuario sem gerar o documento final.
    Permite que o usuario retome a revisao depois."""
    estado = _carregar_estado(sid)
    _aplicar_decisoes(estado, decisoes)
    db.salvar_estado(sid, json.dumps(estado, ensure_ascii=False, default=str))
    return {'ok': True, 'sessao_id': sid}


@app.get('/api/sessao/{sid}/relatorio-pdf')
def relatorio_pdf(sid: str):
    """Relatorio final em PDF, para entrega ao condominio (logo + receitas/
    despesas/quadro/insights/graficos/conclusao + Considerações Importantes).
    Requer que o documento xlsx ja tenha sido gerado (mesma fonte de dados)."""
    estado = _carregar_estado(sid)
    if estado.get('status') != 'gerado':
        raise HTTPException(400, 'Gere o documento (xlsx) antes do relatório em PDF.')
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'assets', 'logo.png')
    pdf_bytes = gerar_relatorio_pdf(estado, logo_path=logo_path if os.path.exists(logo_path) else None)
    filename = f"Relatorio {estado['ano_previsao']} - {estado['nome_condominio']}.pdf"
    return Response(
        content=pdf_bytes,
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.get('/api/sessao/{sid}/download')
def download(sid: str):
    estado = _carregar_estado(sid)
    xlsx_bytes = db.obter_arquivo(sid, 'xlsx')
    if not xlsx_bytes:
        raise HTTPException(404, 'Documento ainda nao gerado')
    filename = f"Previsão {estado['ano_previsao']} - {estado['nome_condominio']}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Frontend estatico (build do React) — prod
# ---------------------------------------------------------------------------
_STATIC = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
if os.path.isdir(_STATIC):
    app.mount('/', StaticFiles(directory=_STATIC, html=True), name='frontend')
