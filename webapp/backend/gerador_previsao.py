#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GERADOR DE PREVISAO ORCAMENTARIA
Usa o modelo_previsao.xlsx como template VISUAL (layout, cores, colunas)
e preenche com os dados extraidos de cada condominio.

Template fixo = visual (cabecalhos, formatacao, abas)
Conteudo dinamico = contas preenchidas conforme os dados reais
"""

import os, re, datetime, unicodedata, warnings, shutil
from copy import copy
warnings.filterwarnings('ignore')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

THIN = Side(style='thin', color='B0B7C3')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
SUB_FILL = PatternFill('solid', fgColor='D6E4F0')
MONEY = '#,##0.00'

MESES_PT = ['janeiro', 'fevereiro', 'marco', 'abril', 'maio', 'junho',
            'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']


def _norm(s):
    s = str(s or '').lower().strip()
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')


# Similaridade minima (Jaccard de tokens) para aceitar um match aproximado.
# Substring por comprimento e cego: ou e frouxo (o generico "contrato" vira
# substring de "Contrato de Manutencao do Jardim" -> contas fantasma) ou e
# estrito demais (descarta "Manut. Extintores e/ou Teste" x "Manut. Extintores").
# Jaccard de tokens distingue os dois casos.
_SIM_MIN_MATCH = 0.5

# Conectores ignorados na tokenizacao (nao carregam significado de conta).
_STOPWORDS = {'de', 'do', 'da', 'dos', 'das', 'e', 'ou', 'com', 'para', 'p',
              'no', 'na', 'a', 'o', 'as', 'os', 'em', 'ao', 'eou'}

# Palavras-CATEGORIA genericas: aparecem em muitas linhas e NAO distinguem a
# conta (ex.: "Contrato de Manutencao do Jardim" x "... da E.T.A." compartilham
# {contrato, manutencao}). O match exige pelo menos um token ESPECIFICO em comum
# (jardim/eta/elevador/...), nunca so as categoricas — senao o generico
# "Contrato de Manutencao" cairia em "...do Jardim" (a 1a linha que aparece).
_GENERICAS = {'contrato', 'manutencao', 'manut', 'despesas', 'taxa', 'geral'}
_TOKENS_CURTOS = {'tv', 'oi', 'oii', 'eta', 'gas', 'fgts', 'inss', 'pis',
                  'csll', 'irrf', 'iptu', 'net'}


def _tokens(nome):
    nome = _norm(nome)
    brutos = re.split(r'[^a-z0-9]+', nome)
    tokens = []
    i = 0
    while i < len(brutos):
        tok = brutos[i]
        if len(tok) == 1 and tok.isalnum():
            j = i
            sigla = ''
            while j < len(brutos) and len(brutos[j]) == 1 and brutos[j].isalnum():
                sigla += brutos[j]
                j += 1
            if len(sigla) >= 2:
                tokens.append(sigla)
            i = j
            continue
        tokens.append(tok)
        i += 1
    return {t for t in tokens
            if ((len(t) >= 3) or t in _TOKENS_CURTOS) and t not in _STOPWORDS}


def _data_extenso(dt=None):
    dt = dt or datetime.date.today()
    return f'{dt.day} de {MESES_PT[dt.month - 1]} de {dt.year}'


def _adicionar_consideracoes(ws, ano):
    """Insere ou substitui a nota final exigida abaixo da tabela da PREVISAO (2)."""
    start = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, 9):
            if 'consideracoes importantes' in _norm(ws.cell(r, c).value):
                start = r
                break
        if start:
            break

    if start is None:
        last_row = 1
        for r in range(1, ws.max_row + 1):
            if any(ws.cell(r, c).value not in (None, '') for c in range(1, 9)):
                last_row = r
        start = last_row + 3

    # Remove conteudo/mesclas antigas da área de considerações para não duplicar
    # datas do template manual.
    end = start + 34
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= end and rng.max_row >= start:
            ws.unmerge_cells(str(rng))
    for r in range(start, min(end, ws.max_row + 40) + 1):
        for c in range(1, 9):
            ws.cell(r, c).value = None

    ws.cell(start, 3).value = 'CONSIDERAÇÕES IMPORTANTES'
    ws.cell(start, 3).font = Font(bold=True, underline='single', size=12)
    ws.cell(start, 3).alignment = Alignment(horizontal='center')
    try:
        ws.merge_cells(start_row=start, start_column=3, end_row=start, end_column=6)
    except ValueError:
        pass

    nota_row = start + 2
    ws.cell(nota_row, 1).value = (
        '1) Para o cálculo desta previsão, levamos em consideração a média '
        'aritmética dos últimos 12 meses'
    )
    ws.cell(nota_row, 1).font = Font(bold=True, size=9)
    try:
        ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=8)
    except ValueError:
        pass

    assinatura_row = nota_row + 28
    ws.cell(assinatura_row, 2).value = 'Cabo Frio,'
    ws.cell(assinatura_row, 3).value = _data_extenso()
    ws.cell(assinatura_row, 5).value = 'PORTO REAL IMÓVEIS'
    for c in (2, 3, 5):
        ws.cell(assinatura_row, c).font = Font(size=10)
        ws.cell(assinatura_row, c).alignment = Alignment(horizontal='center')


def _manter_apenas_previsao2(wb):
    """O arquivo entregue ao usuário deve conter somente PREVISÃO (2)."""
    alvo = None
    for nome in wb.sheetnames:
        nn = _norm(nome).replace(' ', '')
        if 'previsao' in nn and '(2)' in nome:
            alvo = nome
            break
    if alvo is None:
        return
    ws = wb[alvo]
    ws.title = 'PREVISÃO (2)'
    for nome in list(wb.sheetnames):
        if nome != ws.title:
            wb.remove(wb[nome])
    wb.active = 0


def _linhas_contratuais(linhas):
    """Linhas contratuais/pro-labore em ordem util para exibicao no template."""
    out = []
    for idx, linha in enumerate(linhas):
        ng = _norm(linha.get('grupo'))
        nc = _norm(linha.get('classe'))
        if not (('contrato' in ng) or ('pro-labore' in ng) or ('prolabore' in ng)
                or ('sindico' in ng and 'reembolso' in ng)):
            continue
        if abs(linha.get('base', 0) or 0) <= 0.005 and abs(linha.get('final', 0) or 0) <= 0.005:
            continue
        out.append({
            'idx': idx,
            'grupo': linha.get('grupo') or '',
            'classe': linha.get('classe') or '',
            'base': float(linha.get('base') or 0),
            'final': float(linha.get('final') or 0),
            'norm': nc,
        })

    def _ordem(item):
        nc = item['norm']
        tk = _tokens(nc)
        if 'elevador' in nc:
            return (0, item['idx'])
        if 'jardim' in nc:
            return (1, item['idx'])
        if ('tv' in tk) or ('cabo' in tk):
            return (2, item['idx'])
        if ('eta' in tk) or ('piscina' in tk):
            return (3, item['idx'])
        if ('hidraul' in nc) or ('eletric' in nc):
            return (4, item['idx'])
        if any(k in nc for k in ('interf', 'camera', 'portao', 'antena')):
            return (5, item['idx'])
        if 'vigia' in nc:
            return (6, item['idx'])
        if 'contab' in nc:
            return (7, item['idx'])
        if ('internet' in nc) or (tk & {'net', 'oi', 'vivo', 'claro', 'fibra'}):
            return (8, item['idx'])
        if 'administr' in nc:
            return (9, item['idx'])
        if ('sindico' in nc) or ('pro-labore' in nc) or ('prolabore' in nc) or ('ajuda de custo' in nc):
            return (10, item['idx'])
        return (11, item['idx'])

    return sorted(out, key=_ordem)


def _achar_valor(nn, valores, usados):
    """Encontra o melhor valor para a conta normalizada 'nn'.

    1) match exato; 2) similaridade de tokens (Jaccard) >= _SIM_MIN_MATCH E pelo
    menos um token ESPECIFICO (nao-categorico) em comum. Cada chave de 'valores'
    so e consumida uma vez (registrada em 'usados'). Retorna (chave, valor) ou
    (None, None).
    """
    if nn in valores and nn not in usados:
        return nn, valores[nn]
    tn = _tokens(nn)
    espec_n = tn - _GENERICAS
    if not tn or not espec_n:
        return None, None
    melhor_chave = None
    melhor_sim = 0.0
    for vn, vv in valores.items():
        if vn in usados:
            continue
        tv = _tokens(vn)
        if not tv:
            continue
        # Exige token especifico em comum (alem de eventuais categoricas).
        if not (espec_n & (tv - _GENERICAS)):
            continue
        sim = len(tn & tv) / len(tn | tv)
        if sim > melhor_sim:
            melhor_sim = sim
            melhor_chave = vn
    if melhor_chave is not None and melhor_sim >= _SIM_MIN_MATCH:
        return melhor_chave, valores[melhor_chave]
    return None, None


def _receita_mensal(ln):
    """Valor mensal de uma receita = media dos meses ATIVOS (total / n_meses).

    Receitas pontuais (1 mes) nao sao recorrentes e nao entram na previsao
    mensal — retorna None para que a conta seja ignorada.
    """
    nm = ln.get('n_meses') or 0
    if nm <= 1:
        return None
    return round(ln['total'] / nm, 2)


def _copy_row_style(ws, src_row, dst_row, cols):
    """Replica estilo/altura de uma linha visual do template."""
    if src_row < 1 or dst_row < 1:
        return
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in cols:
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)


def _insert_rows_for_dynamic_section(ws, before_row, amount, style_row, cols):
    """Abre espaco antes de uma linha de total sem depender do tamanho fixo do modelo."""
    if amount <= 0:
        return
    ws.insert_rows(before_row, amount)
    for row in range(before_row, before_row + amount):
        _copy_row_style(ws, style_row, row, cols)


def gerar_previsao_adaptativa(destino, R, nome_condominio, ano,
                               num_fracoes=None, inflacao=0.10,
                               impacto_receita_mensal=0.0,
                               inad_detalhe=None, inad_meta=None,
                               referencia=None):
    """
    Gera Previsao.xlsx:
    1. Se tem referencia, clona o layout e substitui valores
    2. Senao, gera do zero com layout padrao
    """
    if referencia and os.path.exists(referencia):
        return _gerar_via_template(referencia, destino, R, nome_condominio, ano,
                                   num_fracoes, inflacao, inad_detalhe, inad_meta)
    else:
        return _gerar_do_zero(destino, R, nome_condominio, ano,
                              num_fracoes, inflacao, inad_detalhe, inad_meta)


# ===========================================================================
# Geração via template (clona layout, substitui valores)
# ===========================================================================
def _gerar_via_template(template_path, destino, R, nome_condominio, ano,
                        num_fracoes, inflacao, inad_detalhe, inad_meta):
    """Clona o template e preenche com os dados do condominio."""
    shutil.copy2(template_path, destino)
    wb = openpyxl.load_workbook(destino)
    bal = R['bal']
    linhas = R['linhas']
    linhas_contratuais = _linhas_contratuais(linhas)
    num_frac = num_fracoes if num_fracoes is not None else 12
    data_ext = _data_extenso()

    # Mapas separados: por CLASSE e por GRUPO. Manter separados evita que uma
    # chave de grupo (ex.: "despesas diversas") seja substring de uma linha de
    # classe (ex.: "Estorno de Despesas Diversas") e contamine o valor dela.
    valores_classe = {}
    valores_grupo = {}
    classes_por_grupo = {}   # grupo_norm -> {classe_norm: final}
    for l in linhas:
        valores_classe[_norm(l['classe'])] = l['final']
        ng = _norm(l['grupo'] or '')
        valores_grupo[ng] = valores_grupo.get(ng, 0) + l['final']
        classes_por_grupo.setdefault(ng, {})[_norm(l['classe'])] = l['final']
    # Para a PREVISAO (que mistura linhas de classe e de grupo) usamos o
    # combinado, com a classe tendo prioridade.
    valores = {**valores_grupo, **valores_classe}

    def _classes_do_grupo(grupo_tmpl):
        """Classes parseadas dos grupos que casam com o grupo do template.
        Restringir ao mesmo grupo evita que uma conta homonima de outro grupo
        (ex.: '13o Taxa de Administracao' existe em Diversas E Administrativas)
        seja roteada para a linha errada."""
        gt = _tokens(grupo_tmpl)
        if not gt:
            return valores_classe
        out = {}
        for g, classes in classes_por_grupo.items():
            tg = _tokens(g)
            if tg and len(gt & tg) / len(gt | tg) >= 0.5:
                out.update(classes)
        # Sem fallback para "todas as classes": se nenhum grupo parseado casa com
        # o grupo do template, as linhas ficam vazias (evita contaminacao entre
        # grupos). Os nomes de grupo do Condominio21 sao padronizados e casam.
        return out

    # ---------- CONTAS ----------
    if ' C O N T A S ' in wb.sheetnames:
        ws_c = wb[' C O N T A S ']

        # Atualizar Confronto Inicial
        for r in range(1, 10):
            n3 = _norm(ws_c.cell(r, 3).value)
            if 'valor transportado' in n3:
                ws_c.cell(r, 4).value = round(R['base_total'], 2)
            elif 'desconsidera' in n3:
                ws_c.cell(r, 4).value = round(R['desconsideracoes'], 2)
            elif 'subtotal inicial' in n3:
                ws_c.cell(r, 4).value = round(R['base_total'] - R['desconsideracoes'], 2)
            elif 'subtotal atual' in n3:
                ws_c.cell(r, 4).value = round(R['subtotal'], 2)

        # Limpar dados RESIDUAIS (lixo de preenchimentos anteriores) que o template
        # carrega: ajustes manuais em E:H (ex.: "Outros Materiais" E=1000) e valores
        # literais em D/I de linhas de classe que ficaram de um condominio anterior
        # (ex.: "Contrato" D=5234.52). Sem isso, linhas nao preenchidas nesta
        # geracao mantem numeros antigos e inflam o subtotal. Preserva FORMULAS do
        # template (VLOOKUP, =SUM, =D-SUM(E:H), =+D205).
        for r in range(11, ws_c.max_row + 1):
            for col in (5, 6, 7, 8):
                if isinstance(ws_c.cell(r, col).value, (int, float)):
                    ws_c.cell(r, col).value = None
            code_b = str(ws_c.cell(r, 2).value or '').strip()
            if re.match(r'^\d{2}\.\d{2}$', code_b):  # linha de classe
                for col in (4, 9):
                    if isinstance(ws_c.cell(r, col).value, (int, float)):
                        ws_c.cell(r, col).value = None

        # Receitas — procura cada conta de receita e atualiza D
        for ln in bal.get('receitas', []):
            val = _receita_mensal(ln)
            if val is None:
                continue
            nc = _norm(ln['classe'])
            for r in range(1, ws_c.max_row + 1):
                nn = _norm(ws_c.cell(r, 3).value)
                if nn == nc:
                    ws_c.cell(r, 4).value = val
                    ws_c.cell(r, 9).value = val
                    break

        # Despesas — percorre linhas do template e atualiza com valores calculados.
        # 'usados' garante que cada valor parseado preencha no maximo uma linha.
        # 'grupo_atual' guarda o ultimo cabecalho de grupo (XX) para restringir o
        # match as classes do mesmo grupo.
        usados_c = set()
        grupo_atual = ''
        for r in range(1, ws_c.max_row + 1):
            code_a = str(ws_c.cell(r, 1).value or '').strip()  # col A
            code_b = str(ws_c.cell(r, 2).value or '').strip()  # col B
            nome = str(ws_c.cell(r, 3).value or '').strip()
            if not nome:
                continue
            # Cabecalho de grupo (codigo XX). O template e inconsistente: grupos
            # 02-10 trazem o codigo na col A, mas 14-17 (Admin, Montagem, Fundo
            # Reserva, Pro-labore) trazem na col B. Aceitar ambas, senao o
            # 'grupo_atual' trava e contas desses grupos nao sao preenchidas.
            grp = code_a if re.match(r'^\d{2}$', code_a) else (
                  code_b if re.match(r'^\d{2}$', code_b) else None)
            if grp:
                if not grp.startswith('01'):
                    grupo_atual = _norm(nome)
                continue
            # So linhas de CLASSE de despesa (codigo XX.YY na col B). Cabecalhos de
            # grupo mantem suas formulas =SUM(...). 01.xx sao receitas.
            if not re.match(r'^\d{2}\.\d{2}$', code_b) or code_b.startswith('01'):
                continue
            nn = _norm(nome)
            candidatos = _classes_do_grupo(grupo_atual)
            chave, val = _achar_valor(nn, candidatos, usados_c)
            if val is not None and abs(val) > 0.005:
                ws_c.cell(r, 4).value = round(val, 2)
                ws_c.cell(r, 9).value = round(val, 2)
                usados_c.add(chave)

        # Provisoes R4 (Laudo de Autovistoria) e R5 (Sistema de Incendio).
        # Sao gravadas como ajuste NEGATIVO na coluna E: como a coluna I do
        # template e =D-SUM(E:H), um E negativo soma a provisao ao valor final,
        # que entao flui para a PREVISAO via formula. Sem isso o "Subtotal Atual"
        # da CONTAS nao bate com a soma da PREVISAO.
        def _acha_linha(*termos):
            for rr in range(1, ws_c.max_row + 1):
                n = _norm(ws_c.cell(rr, 3).value)
                if n and all(t in n for t in termos):
                    return rr
            return None

        r_laudo = _acha_linha('laudo', 'autovistoria') or _acha_linha('autovistoria')
        if r_laudo and R.get('prov_laudo', 0) > 0.005:
            if not ws_c.cell(r_laudo, 4).value:
                ws_c.cell(r_laudo, 4).value = 0
            ws_c.cell(r_laudo, 5).value = -round(R['prov_laudo'], 2)

        r_inc = (_acha_linha('sistema', 'combate', 'incendio')
                 or _acha_linha('registro', 'convencao'))
        if r_inc and R.get('prov_incendio', 0) > 0.005:
            if not ws_c.cell(r_inc, 4).value:
                ws_c.cell(r_inc, 4).value = 0
            ws_c.cell(r_inc, 5).value = -round(R['prov_incendio'], 2)

        # Secoes de contratos e pro-labore variam bastante entre condominios.
        # Em vez de manter rotulos fixos do modelo, reescrevemos essas linhas com
        # as classes reais do balanco para evitar numeros em contas erradas.
        contratos = [ln for ln in linhas_contratuais if 'contrato' in _norm(ln['grupo'])]
        prolabore = [ln for ln in linhas_contratuais if 'contrato' not in _norm(ln['grupo'])]

        for rr in range(193, 204):
            ws_c.cell(rr, 3).value = None
            ws_c.cell(rr, 4).value = None
            ws_c.cell(rr, 9).value = None
            for cc in (5, 6, 7, 8):
                ws_c.cell(rr, cc).value = None
        for rr, ln in zip(range(193, 204), contratos):
            ws_c.cell(rr, 3).value = ln['classe']
            ws_c.cell(rr, 4).value = round(ln['base'], 2)
            ws_c.cell(rr, 9).value = round(ln['final'], 2)

        for rr in (256, 257):
            ws_c.cell(rr, 3).value = None
            ws_c.cell(rr, 4).value = None
            ws_c.cell(rr, 9).value = None
            for cc in (5, 6, 7, 8):
                ws_c.cell(rr, cc).value = None
        for rr, ln in zip((256, 257), prolabore):
            ws_c.cell(rr, 3).value = ln['classe']
            ws_c.cell(rr, 4).value = round(ln['base'], 2)
            ws_c.cell(rr, 9).value = round(ln['final'], 2)

    # ---------- PREVISAO ----------
    ws_p = None
    for nome in wb.sheetnames:
        # nome da aba vem espacado: ' P R E V I S Ã O ' -> remover espacos p/ casar
        if 'REVIS' in nome.upper().replace(' ', '') and '(2)' not in nome:
            ws_p = wb[nome]
            break

    if ws_p:
        # Atualizar cabecalho: titulo com o ano (preserva acentos) e nome do
        # condominio onde houver placeholder (BARRAMARES / COND. ...).
        for r in range(1, 10):
            v = ws_p.cell(r, 1).value
            if isinstance(v, str) and ('PREVIS' in v.upper() or 'ORCAMENT' in v.upper()):
                ws_p.cell(r, 1).value = f'PREVISÃO ORÇAMENTÁRIA PARA {ano}'
            elif isinstance(v, str) and ('BARRAMARES' in v.upper()
                                         or v.upper().startswith('COND')):
                ws_p.cell(r, 1).value = nome_condominio

        # Linha de Obras/Benfeitorias MANTIDAS na revisão (R1 reprovado). O
        # template não tem linha de Obras na PREVISÃO; usamos uma linha livre
        # (dentro do range do SUBTOTAL) referenciando o grupo Obras da CONTAS.
        # Fica 0 quando nada é mantido.
        if ' C O N T A S ' in wb.sheetnames:
            ws_c2 = wb[' C O N T A S ']
            r_obras = None
            for rr in range(1, ws_c2.max_row + 1):
                nrc = _norm(ws_c2.cell(rr, 3).value)
                code_a = str(ws_c2.cell(rr, 1).value or '').strip()
                if 'obras' in nrc and re.match(r'^\d{2}$', code_a):
                    r_obras = rr
                    break
            if r_obras:
                # achar primeira linha de despesa livre (C vazio) antes do SUBTOTAL
                for rr in range(22, 48):
                    nome_l = _norm(ws_p.cell(rr, 3).value)
                    if not nome_l:
                        ws_p.cell(rr, 3).value = 'Despesas com Obras/Benfeitorias'
                        # Usa valor direto (nao formula) — openpyxl nao recalcula
                        obra_val = ws_c2.cell(r_obras, 9).value  # coluna I = final
                        if isinstance(obra_val, (int, float)):
                            ws_p.cell(rr, 4).value = round(float(obra_val), 2)
                        break

        # IMPORTANTE: nesta planilha a PREVISAO e inteiramente movida por
        # formulas que apontam para a CONTAS (ex.: D22 = CONTAS!I64,
        # contratos D33..D42 = CONTAS!I194..I203). Preencher a CONTAS
        # corretamente ja resolve a PREVISAO. So escrevemos um valor estatico
        # quando a celula NAO for formula (templates sem formula / do-zero),
        # nunca sobrescrevendo as formulas do modelo.
        def _set(r, c, valor):
            """Sobrescreve valores quando o dado calculado e confiavel.

            Linhas nao cobertas pelo match textual continuam com a formula do
            template, para que Excel/LibreOffice recalculem a partir da aba
            CONTAS e evitem lacunas em contas como Agua/Luz/Gas.
            """
            ws_p.cell(r, c).value = valor

        # compatibilidade: chamadas antigas a _set_se_nao_formula
        _set_se_nao_formula = _set

        # Receitas — reconstroi a area visivel com receitas recorrentes reais.
        # O template tem rotulos fixos (ex.: Fundo de Obra), mas cada condominio
        # pode ter fundos diferentes (ex.: Fundo de Verao). Evita receita ativa
        # sumir por nao existir no modelo.
        def _receita_entra(ln):
            nc = _norm(ln.get('classe'))
            if any(t in nc for t in ('credito', 'debito', 'debitos',
                                     'rendimento', 'outros', 'multa',
                                     'juros', 'acrescimo')):
                return False
            return (
                'condominio' in nc or nc.startswith('tx') or
                'fundo' in nc or 'agua' in nc or 'gas' in nc or
                'luz' in nc or 'tv' in nc or 'internet' in nc
            )

        def _label_receita(ln):
            nc = _norm(ln.get('classe'))
            if nc in ('tx. condominio', 'tx condominio', 'taxa condominio',
                      'taxas de condominio'):
                return 'Taxas de Condomínio'
            if 'compl' in nc and 'condominio' in nc:
                return 'Complemento Tx. Condomínio'
            # Expande abreviacoes comuns do Condominio21
            raw = str(ln.get('classe') or '').strip()
            return raw

        def _ordem_receita(item):
            nc = _norm(item[0])
            if 'condominio' in nc or nc.startswith('tx'):
                return 0
            if 'fundo reserva' in nc or 'fundo de reserva' in nc:
                return 1
            if 'fundo' in nc:
                return 2
            if 'agua' in nc:
                return 3
            if 'gas' in nc:
                return 4
            if 'luz' in nc:
                return 5
            if 'tv' in nc:
                return 6
            if 'internet' in nc:
                return 7
            return 8

        receitas_prev = []
        usados_receita = set()
        for ln in bal.get('receitas', []):
            val = _receita_mensal(ln)
            if val is None or abs(val) <= 0.005 or not _receita_entra(ln):
                continue
            label = _label_receita(ln)
            nlabel = _norm(label)
            if nlabel in usados_receita:
                continue
            receitas_prev.append((label, round(float(val), 2)))
            usados_receita.add(nlabel)
        receitas_prev.sort(key=_ordem_receita)

        total_rec_row = None
        for rr in range(10, 25):
            if _norm(ws_p.cell(rr, 3).value) == 'total':
                total_rec_row = rr
                break
        if total_rec_row is None:
            total_rec_row = 19

        rec_ini = 10
        rec_slots = max(total_rec_row - rec_ini, 0)
        if len(receitas_prev) > rec_slots:
            extra = len(receitas_prev) - rec_slots
            _insert_rows_for_dynamic_section(ws_p, total_rec_row, extra,
                                             max(total_rec_row - 1, rec_ini), (3, 4, 5, 6))
            total_rec_row += extra

        for r in range(rec_ini, total_rec_row):
            for c in (3, 4, 5, 6):
                ws_p.cell(r, c).value = None
        for r, (label, val) in zip(range(rec_ini, total_rec_row), receitas_prev):
            ws_p.cell(r, 3).value = label
            ws_p.cell(r, 4).value = val
            ws_p.cell(r, 5).value = val

        # Despesas — reconstroi a area visivel com os valores finais calculados.
        # Nao reaproveitamos formulas/rotulos do template aqui: elas podem apontar
        # para linhas antigas da CONTAS e gerar duplicidades ou subtotais falsos.
        subtotal_row = None
        despesas_header_row = None
        for rr in range(22, ws_p.max_row + 1):
            if _norm(ws_p.cell(rr, 3).value) == 'despesas':
                despesas_header_row = rr
            if 'subtotal' in _norm(ws_p.cell(rr, 3).value):
                subtotal_row = rr
                break
        if subtotal_row is None:
            subtotal_row = 48
        desp_ini = (despesas_header_row + 1) if despesas_header_row else 22
        desp_fim = subtotal_row - 1
        for rr in range(desp_ini, desp_fim + 1):
            for cc in (3, 4, 5, 6):
                ws_p.cell(rr, cc).value = None

        def _valor_linha(ln):
            return float(ln.get('final') or 0)

        despesas_ativas = [
            (idx, ln) for idx, ln in enumerate(linhas)
            if abs(_valor_linha(ln)) > 0.005
        ]
        consumidos = set()

        def _somar(label, pred):
            total = 0.0
            for idx, ln in despesas_ativas:
                if idx in consumidos:
                    continue
                if pred(ln):
                    total += _valor_linha(ln)
                    consumidos.add(idx)
            return (label, total) if abs(total) > 0.005 else None

        def _ng(ln):
            return _norm(ln.get('grupo'))

        def _nc(ln):
            return _norm(ln.get('classe'))

        def _tarifa_publica(ln):
            return 'tarifas publicas' in _ng(ln)

        linhas_prev = []
        categorias = [
            ('Despesas com Pessoal',
             lambda ln: 'pessoal' in _ng(ln)),
            ('Luz do Condomínio',
             lambda ln: _tarifa_publica(ln) and 'luz' in _nc(ln)),
            ('Água do Condomínio',
             lambda ln: _tarifa_publica(ln) and 'agua' in _nc(ln)),
            ('Telefone do Condomínio',
             lambda ln: _tarifa_publica(ln) and 'telefone' in _nc(ln)),
            ('Gás do Condomínio',
             lambda ln: _tarifa_publica(ln) and 'gas' in _nc(ln)),
            ('Material de Limpeza',
             lambda ln: 'material' in _nc(ln) and 'limpeza' in _nc(ln)),
            ('Gastos com conservação',
             lambda ln: 'conservacao' in _ng(ln)),
            ('Tarifas Bancárias',
             lambda ln: 'tarifas bancarias' in _ng(ln) or 'tarifas bancarias' in _nc(ln)),
            ('Seguro de Incêndio Obrigatório',
             lambda ln: 'seguro' in _nc(ln) and 'incendio' in _nc(ln)),
            ('Outras despesas diversas',
             lambda ln: 'diversas' in _ng(ln)),
        ]
        for label, pred in categorias:
            item = _somar(label, pred)
            if item:
                linhas_prev.append(item)

        labels_usados = set()
        for ln in linhas_contratuais:
            idx = ln.get('idx')
            if idx in consumidos:
                continue
            anual = float(ln.get('final') or 0)
            if abs(anual) <= 0.005:
                continue
            label_norm = _norm(ln['classe'])
            if label_norm in labels_usados:
                continue
            linhas_prev.append((ln['classe'], anual))
            labels_usados.add(label_norm)
            if idx is not None:
                consumidos.add(idx)

        categorias_finais = [
            ('Despesas Administrativas',
             lambda ln: 'administrativa' in _ng(ln) or 'administrativas' in _ng(ln)),
            ('Despesas Cartoriais e Honorários',
             lambda ln: any(t in _ng(ln) or t in _nc(ln) for t in ('cartori', 'honorari'))),
            ('Despesas com Obras/Benfeitorias',
             lambda ln: any(t in _ng(ln) for t in ('obras', 'benfeitoria'))),
        ]
        for label, pred in categorias_finais:
            item = _somar(label, pred)
            if item:
                linhas_prev.append(item)

        prov_laudo = float(R.get('prov_laudo') or 0)
        prov_incendio = float(R.get('prov_incendio') or 0)
        if abs(prov_laudo) > 0.005:
            linhas_prev.append(('Provisão Laudo Autovistoria', prov_laudo))
        if abs(prov_incendio) > 0.005:
            linhas_prev.append(('Provisão Sistema de Incêndio/Registro', prov_incendio))

        labels_existentes = {_norm(label) for label, _ in linhas_prev}
        for idx, ln in despesas_ativas:
            if idx in consumidos:
                continue
            label = str(ln.get('classe') or ln.get('grupo') or 'Despesa sem classificação').strip()
            nlabel = _norm(label)
            if nlabel in labels_existentes:
                label = f"{ln.get('grupo') or 'Outros'} - {label}"
                nlabel = _norm(label)
            linhas_prev.append((label, _valor_linha(ln)))
            labels_existentes.add(nlabel)
            consumidos.add(idx)

        alvo_subtotal = float(R.get('subtotal') or 0)
        soma_prev = sum(valor for _, valor in linhas_prev)
        diferenca = round(alvo_subtotal - soma_prev, 2)
        if abs(diferenca) > 0.05:
            linhas_prev.append(('Ajustes de previsão', diferenca))

        slots = max(desp_fim - desp_ini + 1, 0)
        if len(linhas_prev) > slots:
            extra = len(linhas_prev) - slots
            _insert_rows_for_dynamic_section(ws_p, subtotal_row, extra,
                                             max(subtotal_row - 1, desp_ini), (3, 4, 5, 6))
            subtotal_row += extra
            desp_fim += extra

        for rr, (label, anual) in zip(range(desp_ini, desp_fim + 1), linhas_prev):
            anual = round(float(anual), 2)
            ws_p.cell(rr, 3).value = label
            ws_p.cell(rr, 4).value = anual
            ws_p.cell(rr, 5).value = round(anual / num_frac, 2)
            ws_p.cell(rr, 6).value = round(anual / 12, 2)

        subtotal_val = round(sum(valor for _, valor in linhas_prev), 2)

        total_rec = sum(float(ws_p.cell(rr, 4).value or 0)
                        for rr in range(rec_ini, total_rec_row)
                        if isinstance(ws_p.cell(rr, 4).value, (int, float)))

        # SUBTOTAL, INFLACAO, TOTAL, SALDO (so para templates sem formula)
        fundo_mensal = sum(v for lbl, v in receitas_prev
                           if 'fundo' in _norm(lbl) and 'reserva' in _norm(lbl))
        saldo_row = None
        saldo_com_fundo = 0.0
        for r in range(1, ws_p.max_row + 1):
            n3 = _norm(ws_p.cell(r, 3).value)
            if 'subtotal' in n3:
                ws_p.cell(r, 3).value = 'SUBTOTAL'
                _set_se_nao_formula(r, 4, round(subtotal_val, 2))
                _set_se_nao_formula(r, 6, round(subtotal_val / 12, 2))
            elif 'infla' in n3 and 'previsao' in n3:
                pct_txt = f'{inflacao * 100:.2f}'.replace('.', ',').rstrip('0').rstrip(',')
                ws_p.cell(r, 3).value = f'PREVISÃO DE INFLAÇÃO - {pct_txt}%'
                _set_se_nao_formula(r, 4, round(subtotal_val * inflacao, 2))
                _set_se_nao_formula(r, 5, round(inflacao, 4))
            elif n3 == 'total':
                if r == total_rec_row:
                    _set_se_nao_formula(r, 4, round(total_rec, 2))
                    _set_se_nao_formula(r, 5, round(total_rec, 2))
                else:
                    ws_p.cell(r, 3).value = 'TOTAL'
                    total = subtotal_val * (1 + inflacao)
                    _set_se_nao_formula(r, 4, round(total, 2))
                    _set_se_nao_formula(r, 5, round(total / num_frac, 2))
            elif 'saldo' in n3 or 'deficit' in n3 or 'superavit' in n3:
                rec_anual = total_rec * 12
                total = subtotal_val * (1 + inflacao)
                saldo = round(rec_anual - total, 2)
                ws_p.cell(r, 3).value = 'SALDO ( SUPERÁVIT )' if saldo >= 0 else 'SALDO ( DÉFICIT )'
                _set_se_nao_formula(r, 4, saldo)
                _set_se_nao_formula(r, 5, round(saldo / 12, 2))
                saldo_row = r
                saldo_com_fundo = saldo

        # Cenario SEM fundo de reserva (feedback CEO 07/2026): a receita do
        # fundo nao cobre despesa ordinaria, entao o documento mostra tambem
        # o saldo excluindo essa arrecadacao.
        if saldo_row is not None and abs(fundo_mensal) > 0.005:
            ws_p.cell(saldo_row, 3).value = (
                'SALDO COM FUNDO DE RESERVA ( SUPERÁVIT )' if saldo_com_fundo >= 0
                else 'SALDO COM FUNDO DE RESERVA ( DÉFICIT )')
            alvo_p = saldo_row + 1
            if any(str(ws_p.cell(alvo_p, c).value or '').strip() for c in (3, 4, 5, 6)):
                # inserir desloca merges abaixo — evitar quando possivel
                ws_p.insert_rows(alvo_p)
            _copy_row_style(ws_p, saldo_row, alvo_p, (3, 4, 5, 6))
            saldo_sem = round(saldo_com_fundo - fundo_mensal * 12, 2)
            ws_p.cell(alvo_p, 3).value = (
                'SALDO SEM FUNDO DE RESERVA ( SUPERÁVIT )' if saldo_sem >= 0
                else 'SALDO SEM FUNDO DE RESERVA ( DÉFICIT )')
            ws_p.cell(alvo_p, 4).value = saldo_sem
            ws_p.cell(alvo_p, 5).value = round(saldo_sem / 12, 2)

        def _num_cell(row, col):
            val = ws_p.cell(row, col).value
            return float(val) if isinstance(val, (int, float)) else 0.0

        vis_rec = [
            _norm(ws_p.cell(rr, 3).value)
            for rr in range(rec_ini, total_rec_row)
            if str(ws_p.cell(rr, 3).value or '').strip()
        ]
        vis_desp = [
            _norm(ws_p.cell(rr, 3).value)
            for rr in range(desp_ini, subtotal_row)
            if str(ws_p.cell(rr, 3).value or '').strip()
        ]
        if len(vis_rec) != len(receitas_prev):
            raise ValueError(
                f'PREVISAO inconsistente: {len(receitas_prev)} receitas calculadas, '
                f'{len(vis_rec)} renderizadas'
            )
        if len(vis_desp) != len(linhas_prev):
            raise ValueError(
                f'PREVISAO inconsistente: {len(linhas_prev)} despesas calculadas, '
                f'{len(vis_desp)} renderizadas'
            )
        if len(set(vis_rec)) != len(vis_rec):
            raise ValueError('PREVISAO inconsistente: receitas com rotulos duplicados')
        if len(set(vis_desp)) != len(vis_desp):
            raise ValueError('PREVISAO inconsistente: despesas com rotulos duplicados')
        soma_rec_visivel = sum(_num_cell(rr, 4) for rr in range(rec_ini, total_rec_row))
        if abs(soma_rec_visivel - _num_cell(total_rec_row, 4)) > 0.05:
            raise ValueError(
                f'PREVISAO inconsistente: receitas visiveis={soma_rec_visivel:.2f}, '
                f'total={_num_cell(total_rec_row, 4):.2f}'
            )
        soma_desp_visivel = sum(_num_cell(rr, 4) for rr in range(desp_ini, subtotal_row))
        if abs(soma_desp_visivel - _num_cell(subtotal_row, 4)) > 0.05:
            raise ValueError(
                f'PREVISAO inconsistente: despesas visiveis={soma_desp_visivel:.2f}, '
                f'subtotal={_num_cell(subtotal_row, 4):.2f}'
            )

    # ---------- PREVISAO (2) ----------
    for nome in wb.sheetnames:
        if 'REVIS' in nome.upper().replace(' ', '') and '(2)' in nome:
            ws_p2 = wb[nome]
            ws_p2['F48'] = inflacao
            cab = str(ws_p2['A6'].value or '')
            if cab.startswith('=') or 'BARRAMARES' in cab.upper() or cab.upper().startswith('COND'):
                ws_p2['A6'] = nome_condominio
            if isinstance(ws_p2['A8'].value, str) and 'PREVIS' in ws_p2['A8'].value.upper():
                ws_p2['A8'] = f'PREVISÃO ORÇAMENTÁRIA PARA {ano}'

            # Constroi listas em ordem da PREVISAO (ignora B codes — sao formulas
            # que openpyxl nao calcula). A PREVISAO (2) tem a mesma ordem.
            prev_rec = []   # (nome, val_d, val_e) em ordem
            prev_desp = []
            receita_total_row = None
            despesas_header_row = None
            subtotal_prev_row = None
            for rp in range(1, ws_p.max_row + 1):
                nn = _norm(ws_p.cell(rp, 3).value)
                if nn == 'despesas':
                    despesas_header_row = rp
                elif nn == 'total' and despesas_header_row is None:
                    receita_total_row = rp
                elif 'subtotal' in nn and despesas_header_row is not None:
                    subtotal_prev_row = rp
                    break
            for rp in range(1, ws_p.max_row + 1):
                nome = str(ws_p.cell(rp, 3).value or '').strip()
                if not nome:
                    continue
                nn = _norm(nome)
                if nn in ('receitas', 'despesas', 'subtotal', 'total', ''):
                    continue
                if 'previsao de inflacao' in nn or 'aumento' in nn:
                    continue
                if 'saldo' in nn or 'deficit' in nn or 'superavit' in nn:
                    continue
                val_d = ws_p.cell(rp, 4).value
                val_e = ws_p.cell(rp, 5).value
                if isinstance(val_d, (int, float)) and abs(val_d) > 0.005:
                    if receita_total_row and 10 <= rp < receita_total_row:
                        prev_rec.append((nome, val_d, val_e))
                    elif (despesas_header_row and subtotal_prev_row and
                          despesas_header_row < rp < subtotal_prev_row):
                        prev_desp.append((nome, val_d, val_e))

            # Preenche PREVISAO (2) na mesma ordem, abrindo linhas quando a
            # PREVISAO dinamica tiver mais itens que o modelo antigo.
            rec_ini2 = 11
            rec_total_row2 = None
            for rr in range(rec_ini2, 30):
                if _norm(ws_p2.cell(rr, 3).value) == 'total':
                    rec_total_row2 = rr
                    break
            if rec_total_row2 is None:
                rec_total_row2 = 19
            rec_slots2 = max(rec_total_row2 - rec_ini2, 0)
            if len(prev_rec) > rec_slots2:
                extra = len(prev_rec) - rec_slots2
                _insert_rows_for_dynamic_section(ws_p2, rec_total_row2, extra,
                                                 max(rec_total_row2 - 1, rec_ini2), (3, 4, 5, 6, 7, 8))
                rec_total_row2 += extra
            for rr in range(rec_ini2, rec_total_row2):
                for cc in (3, 4, 5, 6, 7, 8):
                    ws_p2.cell(rr, cc).value = None
            for i, (nome, val_d, val_e) in enumerate(prev_rec):
                r2 = rec_ini2 + i
                ws_p2.cell(r2, 3).value = nome
                ws_p2.cell(r2, 4).value = round(float(val_d), 2) if isinstance(val_d, (int, float)) else val_d
                ws_p2.cell(r2, 5).value = round(float(val_e), 2) if isinstance(val_e, (int, float)) else val_e

            ws_p2['E11'] = num_frac

            desp_ini2 = 22
            subtotal_row2 = None
            for rr in range(desp_ini2, ws_p2.max_row + 1):
                if 'subtotal' in _norm(ws_p2.cell(rr, 3).value):
                    subtotal_row2 = rr
                    break
            if subtotal_row2 is None:
                subtotal_row2 = 48
            desp_slots2 = max(subtotal_row2 - desp_ini2, 0)
            if len(prev_desp) > desp_slots2:
                extra = len(prev_desp) - desp_slots2
                _insert_rows_for_dynamic_section(ws_p2, subtotal_row2, extra,
                                                 max(subtotal_row2 - 1, desp_ini2), (3, 4, 5, 6, 7, 8))
                subtotal_row2 += extra
            for rr in range(desp_ini2, subtotal_row2):
                for cc in (3, 4, 5, 6, 7, 8):
                    ws_p2.cell(rr, cc).value = None
            for i, (nome, val_d, val_e) in enumerate(prev_desp):
                r2 = desp_ini2 + i
                ws_p2.cell(r2, 3).value = nome
                val_d_div = round(float(val_d) / 12, 2) if isinstance(val_d, (int, float)) else val_d
                ws_p2.cell(r2, 4).value = val_d_div
                ws_p2.cell(r2, 5).value = round(float(val_e), 2) if isinstance(val_e, (int, float)) else val_e

            # Linhas de totais: identifica pelo texto da FORMULA (antes de ser limpa)
            # A formula contem "SUBTOTAL", "TOTAL", "SALDO" etc.
            fundo_mensal2 = sum(
                v for nome, v, _ in prev_rec
                if isinstance(v, (int, float))
                and 'fundo' in _norm(nome) and 'reserva' in _norm(nome))
            saldo_row2 = None
            for r2 in range(1, ws_p2.max_row + 1):
                val_c3 = ws_p2.cell(r2, 3).value
                val_text = str(val_c3 or '')
                # Tenta match pelo texto da formula ou pelo valor calculado
                if 'SUBTOTAL' in val_text.upper():
                    ws_p2.cell(r2, 3).value = 'SUBTOTAL'
                    ws_p2.cell(r2, 4).value = round(subtotal_val, 2)
                    ws_p2.cell(r2, 7).value = 'Inflação'
                elif 'AUMENTO' in val_text.upper() or ('infla' in _norm(val_text) and 'aumento' in _norm(val_text)):
                    ws_p2.cell(r2, 4).value = round(subtotal_val * inflacao, 2)
                    # F guarda a fracao usada pelo rotulo '="Aumento previsto..." & F48*100'
                    if isinstance(ws_p2.cell(r2, 6).value, (int, float)):
                        ws_p2.cell(r2, 6).value = round(inflacao, 4)
                elif val_text.upper().strip() in ('TOTAL', '="TOTAL"') or (_norm(val_text) == 'total' and 'aumento' not in _norm(val_text)):
                    ws_p2.cell(r2, 3).value = 'TOTAL'
                    total = subtotal_val * (1 + inflacao)
                    ws_p2.cell(r2, 4).value = round(total, 2)
                    ws_p2.cell(r2, 5).value = round(total / num_frac, 2)
                elif 'SALDO' in val_text.upper() or 'DEFICIT' in val_text.upper() or 'SUPERAVIT' in val_text.upper():
                    rec_mensal = sum(v for _, v, _ in prev_rec if isinstance(v, (int, float)))
                    rec_anual = rec_mensal * 12
                    total_anual = subtotal_val * (1 + inflacao)
                    saldo_val = rec_anual - total_anual
                    ws_p2.cell(r2, 3).value = ('SALDO ( SUPERÁVIT )' if saldo_val >= 0
                                               else 'SALDO ( DÉFICIT )')
                    ws_p2.cell(r2, 4).value = round(saldo_val, 2)
                    ws_p2.cell(r2, 5).value = round(saldo_val / 12, 2)
                    saldo_row2 = r2
                    saldo_val2 = saldo_val
                # Fallback: identifica pelo valor na coluna B (99 = inflacao)
                val_b = ws_p2.cell(r2, 2).value
                if val_b == 99:
                    ws_p2.cell(r2, 4).value = round(subtotal_val * inflacao, 2)

            # Cenario duplo COM/SEM fundo de reserva (feedback CEO 07/2026):
            # segunda linha de saldo excluindo a arrecadacao do fundo.
            if saldo_row2 is not None and abs(fundo_mensal2) > 0.005:
                ws_p2.cell(saldo_row2, 3).value = (
                    'SALDO COM FUNDO DE RESERVA ( SUPERÁVIT )' if saldo_val2 >= 0
                    else 'SALDO COM FUNDO DE RESERVA ( DÉFICIT )')
                alvo = saldo_row2 + 1
                ocupado = any(str(ws_p2.cell(alvo, c).value or '').strip()
                              for c in (3, 4, 5, 6))
                if ocupado:
                    # inserir desloca merges das Consideracoes — evitar quando possivel
                    ws_p2.insert_rows(alvo)
                _copy_row_style(ws_p2, saldo_row2, alvo, (3, 4, 5, 6))
                saldo_sem2 = round(saldo_val2 - fundo_mensal2 * 12, 2)
                ws_p2.cell(alvo, 3).value = (
                    'SALDO SEM FUNDO DE RESERVA ( SUPERÁVIT )' if saldo_sem2 >= 0
                    else 'SALDO SEM FUNDO DE RESERVA ( DÉFICIT )')
                ws_p2.cell(alvo, 4).value = saldo_sem2
                ws_p2.cell(alvo, 5).value = round(saldo_sem2 / 12, 2)

            # Colunas G/H do template manual: "Aumento c/ FR" e "s/ FR" —
            # substituir as formulas (que apontavam para abas removidas)
            # pelos valores calculados, em vez de apenas limpa-las.
            rec_mensal2 = sum(v for _, v, _ in prev_rec if isinstance(v, (int, float)))
            rec_anual2 = rec_mensal2 * 12
            total_anual2 = subtotal_val * (1 + inflacao)
            for r2 in range(1, ws_p2.max_row + 1):
                for c in (7, 8):
                    val = ws_p2.cell(r2, c).value
                    if isinstance(val, str) and val.startswith('='):
                        novo = None
                        if 'd50' in val.lower() and 'd19' in val.lower():
                            sem_fr = '(d19' in val.lower().replace(' ', '')
                            rec_base = (rec_anual2 - fundo_mensal2 * 12) if sem_fr else rec_anual2
                            if rec_base > 0.005:
                                novo = round(total_anual2 / rec_base - 1, 4)
                                ws_p2.cell(r2 - 1, c).value = (
                                    'Aumento s/ FR' if sem_fr else 'Aumento c/ FR')
                        ws_p2.cell(r2, c).value = novo
            _adicionar_consideracoes(ws_p2, ano)
            break

    # ---------- Recalculo ----------
    # Nao apagar formulas do template: varias linhas (ex.: Agua/Luz/Gas) usam
    # referencias corretas para a CONTAS, mesmo quando o match textual nao
    # encontra a conta equivalente. Marcamos o workbook para recalc completo ao
    # abrir, evitando manter os caches antigos do modelo.
    try:
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcOnSave = True
        wb.calculation.calcCompleted = False
    except Exception:
        pass

    _manter_apenas_previsao2(wb)
    wb.save(destino)
    return {'ok': True, 'modo': 'template'}


# ===========================================================================
# Geração do zero (fallback sem template)
# ===========================================================================
def _gerar_do_zero(destino, R, nome_condominio, ano, num_fracoes, inflacao,
                    inad_detalhe, inad_meta):
    """Gera Previsao do zero com layout padrao."""
    bal = R['bal']
    linhas = R['linhas']
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    num_frac = num_fracoes if num_fracoes is not None else 12
    data_ext = _data_extenso()

    ORDEM_GRUPOS = [
        'Despesas com Pessoal', 'Tarifas Publicas', 'Conservacao',
        'Tarifas Bancarias', 'Despesas Diversas', 'Contratos',
        'Despesas Cartoriais e Honorarios', 'Despesas Administrativas',
        'Reembolso/Pro-labore ao Sindico', 'Despesas com Obras/Benfeitorias',
    ]

    # CONTAS
    ws_c = wb.create_sheet(' C O N T A S ')
    for col, w in zip('ABCDEFGHI', [5, 12, 48, 16, 16, 14, 14, 14, 16]):
        ws_c.column_dimensions[col].width = w
    r = 1
    ws_c.cell(r, 3, 'Confronto Inicial').font = Font(bold=True, size=13)
    r += 2
    ws_c.cell(r, 3, 'Valor Transportado')
    ws_c.cell(r, 4, round(R['base_total'], 2)).number_format = MONEY
    r += 1
    ws_c.cell(r, 3, 'Desconsideracoes')
    ws_c.cell(r, 4, round(R['desconsideracoes'], 2)).number_format = MONEY
    r += 1
    ws_c.cell(r, 3, 'Subtotal Inicial').font = Font(bold=True)
    ws_c.cell(r, 4, round(R['base_total'] - R['desconsideracoes'], 2)).number_format = MONEY
    r += 2

    # Receitas no CONTAS
    r_ini = r
    cod = 1
    for ln in bal.get('receitas', []):
        val = _receita_mensal(ln)
        if val is None:
            continue
        ws_c.cell(r, 2, f'01.{cod:02d}')
        ws_c.cell(r, 3, ln['classe'])
        ws_c.cell(r, 4, val).number_format = MONEY
        ws_c.cell(r, 9, val).number_format = MONEY
        cod += 1
        r += 1
    r += 1

    # Despesas
    por_grupo = defaultdict(list)
    for l in linhas:
        por_grupo[l['grupo'] or 'Outros'].append(l)
    cod_grupo = 2
    for grupo_nome in ORDEM_GRUPOS:
        match_key = None
        for g in por_grupo:
            if _norm(grupo_nome) in _norm(g) or _norm(g) in _norm(grupo_nome):
                match_key = g
                break
        if match_key is None:
            continue
        linhas_grupo = por_grupo.pop(match_key)
        base_g = sum(l['base'] for l in linhas_grupo)
        final_g = sum(l['final'] for l in linhas_grupo)
        ws_c.cell(r, 2, f'{cod_grupo:02d}')
        ws_c.cell(r, 3, grupo_nome).font = Font(bold=True)
        ws_c.cell(r, 4, round(base_g, 2)).number_format = MONEY
        ws_c.cell(r, 9, round(final_g, 2)).number_format = MONEY
        for c in range(1, 10):
            ws_c.cell(r, c).fill = SUB_FILL
        r += 1
        cod_classe = 1
        for l in linhas_grupo:
            ws_c.cell(r, 2, f'{cod_grupo:02d}.{cod_classe:02d}')
            ws_c.cell(r, 3, l['classe'])
            ws_c.cell(r, 4, round(l['base'], 2)).number_format = MONEY
            if abs(l['deducao']) > 0.005:
                ws_c.cell(r, 5, round(l['deducao'], 2)).number_format = MONEY
            ws_c.cell(r, 9, round(l['final'], 2)).number_format = MONEY
            ws_c.cell(r, 6, l['n_meses'])
            cod_classe += 1
            r += 1
        cod_grupo += 1

    r += 1
    ws_c.cell(r, 3, 'Subtotal Atual').font = Font(bold=True, size=12)
    ws_c.cell(r, 4, round(R['subtotal'], 2)).number_format = MONEY

    # PREVISAO (2) — no fallback tambem entregamos somente a aba final.
    ws_p = wb.create_sheet('PREVISÃO (2)')
    for col, w in zip('ABCDEF', [5, 8, 52, 18, 18, 18]):
        ws_p.column_dimensions[col].width = w
    r = 6
    ws_p.cell(r, 1, nome_condominio).font = Font(bold=True, size=13)
    r += 1
    ws_p.cell(r, 1, f'PREVISÃO ORÇAMENTÁRIA PARA {ano}').font = Font(bold=True, size=13)
    r += 1
    ws_p.cell(r, 1, data_ext).font = Font(italic=True)
    r = 9
    ws_p.cell(r, 3, 'RECEITAS').font = Font(bold=True)
    ws_p.cell(r, 4, 'VALOR MENSAL').font = Font(bold=True)
    r += 1
    rec_total = 0
    for ln in bal.get('receitas', []):
        val = _receita_mensal(ln)
        if val is None:
            continue
        ws_p.cell(r, 3, ln['classe'])
        ws_p.cell(r, 4, val).number_format = MONEY
        ws_p.cell(r, 5, val).number_format = MONEY
        rec_total += val
        r += 1
    r += 1
    ws_p.cell(r, 3, 'TOTAL').font = Font(bold=True)
    ws_p.cell(r, 4, round(rec_total, 2)).number_format = MONEY
    r += 2
    ws_p.cell(r, 3, 'DESPESAS').font = Font(bold=True)
    ws_p.cell(r, 4, 'VALOR ANUAL').font = Font(bold=True)
    r += 1
    por_grupo2 = defaultdict(list)
    for l in linhas:
        por_grupo2[l['grupo'] or 'Outros'].append(l)
    despesa_total = 0
    idx = 1
    for grupo_nome in ORDEM_GRUPOS:
        match_key = None
        for g in por_grupo2:
            if _norm(grupo_nome) in _norm(g) or _norm(g) in _norm(grupo_nome):
                match_key = g
                break
        if match_key is None:
            continue
        linhas_grupo = por_grupo2.pop(match_key)
        final_g = sum(l['final'] for l in linhas_grupo)
        if abs(final_g) < 0.005:
            continue
        ws_p.cell(r, 2, idx)
        ws_p.cell(r, 3, grupo_nome)
        ws_p.cell(r, 4, round(final_g, 2)).number_format = MONEY
        ws_p.cell(r, 5, round(final_g / num_frac, 2)).number_format = MONEY
        ws_p.cell(r, 6, round(final_g / 12, 2)).number_format = MONEY
        despesa_total += final_g
        idx += 1
        r += 1
    r += 1
    ws_p.cell(r, 3, 'SUBTOTAL').font = Font(bold=True)
    ws_p.cell(r, 4, round(despesa_total, 2)).number_format = MONEY
    r += 1
    inflacao_val = despesa_total * inflacao
    ws_p.cell(r, 2, 99)
    ws_p.cell(r, 3, f'PREVISÃO DE INFLAÇÃO - {inflacao*100:.0f}%')
    ws_p.cell(r, 4, round(inflacao_val, 2)).number_format = MONEY
    r += 2
    total = despesa_total + inflacao_val
    ws_p.cell(r, 3, 'TOTAL').font = Font(bold=True, size=12)
    ws_p.cell(r, 4, round(total, 2)).number_format = MONEY
    r += 2
    rec_anual = rec_total * 12
    saldo = rec_anual - total
    fundo_mensal = sum(
        _receita_mensal(ln) or 0 for ln in bal.get('receitas', [])
        if 'fundo' in _norm(ln.get('classe')) and 'reserva' in _norm(ln.get('classe')))
    if abs(fundo_mensal) > 0.005:
        # Cenario duplo COM/SEM fundo de reserva (feedback CEO 07/2026)
        rotulo = ('SALDO COM FUNDO DE RESERVA ( DÉFICIT )' if saldo < 0
                  else 'SALDO COM FUNDO DE RESERVA ( SUPERÁVIT )')
        ws_p.cell(r, 3, rotulo).font = Font(bold=True, size=12)
        ws_p.cell(r, 4, round(saldo, 2)).number_format = MONEY
        r += 1
        saldo_sem = saldo - fundo_mensal * 12
        rotulo_sem = ('SALDO SEM FUNDO DE RESERVA ( DÉFICIT )' if saldo_sem < 0
                      else 'SALDO SEM FUNDO DE RESERVA ( SUPERÁVIT )')
        ws_p.cell(r, 3, rotulo_sem).font = Font(bold=True, size=12)
        ws_p.cell(r, 4, round(saldo_sem, 2)).number_format = MONEY
    else:
        ws_p.cell(r, 3, 'SALDO ( DÉFICIT )' if saldo < 0 else 'SALDO ( SUPERÁVIT )').font = Font(bold=True, size=12)
        ws_p.cell(r, 4, round(saldo, 2)).number_format = MONEY
    _adicionar_consideracoes(ws_p, ano)

    _manter_apenas_previsao2(wb)
    wb.save(destino)
    return {'ok': True, 'modo': 'zero'}
