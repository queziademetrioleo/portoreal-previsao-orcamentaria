#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VALIDADOR DE CORPUS — regressao da previsao contra os manuais da Quezia.

Para cada pasta <Condominio>/<Condominio ANO>/ que contenha os 4 relatorios de
entrada (balanual/desbai06/dessin02[/inad01]) E o manual "Previsao ANO.xlsx",
gera a previsao automatica e compara, linha a linha, com o manual.

Classifica cada linha de despesa como:
  - DETERMINISTICA: deve bater (Pessoal, contratos, agua, luz, tarifas, etc.)
  - REVISAO: esperado divergir (depende de decisao humana R3/R4) — conservacao,
    "outras despesas diversas".

Uso:
    python3 validar_corpus.py "/caminho/para/Quezia - Previsao Orcamentaria 3"
"""
import sys, os, re, warnings, unicodedata, tempfile, shutil
warnings.filterwarnings("ignore")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.chdir(os.path.dirname(os.path.abspath(__file__)))
import previsao as core
from gerador_previsao import gerar_previsao_adaptativa
import openpyxl

MODELO = os.path.join("templates", "modelo_previsao.xlsx")
REVIEW_KEYS = ['conserva', 'outras despesas diversas']


def _norm(s):
    s = str(s or '').lower().strip()
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')


def _evalI(c):
    """Avalia a coluna I (valor final) da CONTAS resolvendo as formulas do
    template (=SUM(range), =D-SUM(E:H); VLOOKUP/externas -> 0)."""
    def num(x):
        return float(x) if isinstance(x, (int, float)) else 0.0
    cache, stack = {}, set()

    def I(r):
        if r in cache:
            return cache[r]
        if r in stack:
            return 0.0
        stack.add(r)
        cache[r] = _calc(r)
        stack.discard(r)
        return cache[r]

    def _calc(r):
        v = c.cell(r, 9).value
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v or "")
        if not s.startswith("=") or "VLOOKUP" in s:
            return 0.0
        m = re.match(r"=SUM\((\w+?)(\d+):(\w+?)(\d+)\)$", s)
        if m:
            return sum(I(r2) for r2 in range(int(m.group(2)), int(m.group(4)) + 1))
        if "SUM(E" in s:
            return num(c.cell(r, 4).value) - sum(num(c.cell(r, col).value)
                                                 for col in (5, 6, 7, 8))
        return 0.0
    return I


def previsao_gerada(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    I = _evalI(wb[" C O N T A S "])
    pa = wb[" P R E V I S Ã O "]
    out = {}
    for r in range(22, 48):
        lbl = pa.cell(r, 3).value
        if not lbl:
            continue
        s = str(pa.cell(r, 4).value or "")
        if not s.startswith("="):
            v = pa.cell(r, 4).value
            v = float(v) if isinstance(v, (int, float)) else 0.0
        else:
            v = sum((-1 if m.group(1) == '-' else 1) * I(int(m.group(2)))
                    for m in re.finditer(r"([+\-]?)\s*' C O N T A S '!\$?[A-Z]+\$?(\d+)", s))
        out[_norm(lbl)] = (str(lbl), v)
    return out


def previsao_manual(path):
    ws = openpyxl.load_workbook(path, data_only=True)[" P R E V I S Ã O "]
    out = {}
    for r in range(20, 52):
        lbl = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        if lbl and isinstance(d, (int, float)):
            out[_norm(lbl)] = (str(lbl), float(d))
    return out


def _sheet_previsao(path, data_only=True):
    wb = openpyxl.load_workbook(path, data_only=data_only)
    for name in wb.sheetnames:
        nn = _norm(name).replace(' ', '')
        if 'previs' in nn and '(2)' not in name:
            return wb[name]
    raise ValueError(f'Aba PREVISAO nao encontrada em {path}')


def _fmt(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2) if abs(float(v)) > 0.005 else 0.0
    if v is None:
        return ''
    return str(v).strip()


def comparar_previsao_visivel(gerado, manual):
    """Compara o que aparece no documento final, nao apenas linhas tecnicas."""
    ws_g = _sheet_previsao(gerado, data_only=True)
    ws_m = _sheet_previsao(manual, data_only=True)
    diffs = []
    for r in range(9, 53):
        for c in (3, 4, 5, 6):
            gv = _fmt(ws_g.cell(r, c).value)
            mv = _fmt(ws_m.cell(r, c).value)
            if gv == mv:
                continue
            if isinstance(gv, float) and isinstance(mv, float) and abs(gv - mv) < 0.01:
                continue
            diffs.append((r, c, gv, mv))
    return diffs


def main(root):
    from collections import defaultdict
    byline = defaultdict(lambda: [0, 0.0, 0.0])
    casos = tot_ok = tot_dif = 0
    print(f"{'CONDO/ANO':28} {'detOK':>6} {'detDIF':>6} {'Δdet R$':>11}")
    print("-" * 56)
    for condo in sorted(os.listdir(root)):
        cp = os.path.join(root, condo)
        if not os.path.isdir(cp):
            continue
        for yd in sorted(os.listdir(cp)):
            yp = os.path.join(cp, yd)
            if not os.path.isdir(yp):
                continue
            if not all(os.path.exists(os.path.join(yp, n))
                       for n in ['balanual.xls', 'desbai06.xls', 'dessin02.xls']):
                continue
            man = [os.path.join(yp, f) for f in os.listdir(yp)
                   if f.lower().endswith(".xlsx") and "revis" in _norm(f)]
            if not man:
                continue
            ano = (re.search(r'(\d{4})', yd) or [None, '2026'])[1]
            tmp = f"/tmp/_val_{_norm(condo)}_{ano}.xlsx".replace(" ", "_")
            try:
                with tempfile.TemporaryDirectory() as workdir:
                    for fname in ('balanual.xls', 'desbai06.xls', 'dessin02.xls', 'inad01.xls'):
                        src = os.path.join(yp, fname)
                        if os.path.exists(src):
                            shutil.copy2(src, os.path.join(workdir, fname))
                    R = core.analisar(workdir)
                gerar_previsao_adaptativa(tmp, R, condo, int(ano), num_fracoes=None,
                                          inflacao=0.10, referencia=MODELO)
                mine, manu = previsao_gerada(tmp), previsao_manual(man[0])
                diffs_visiveis = comparar_previsao_visivel(tmp, man[0])
            except Exception as e:
                print(f"{(condo + ' ' + ano)[:28]:28}  ERRO: {e}")
                continue
            ok = dif = 0
            ddet = 0.0
            for k, (lbl, mv) in mine.items():
                if k not in manu or any(rk in k for rk in REVIEW_KEYS):
                    continue
                d = mv - manu[k][1]
                if abs(d) < 1.0:
                    ok += 1
                else:
                    dif += 1
                    ddet += d
                    e = byline[lbl]
                    e[0] += 1
                    e[1] += abs(d)
                    e[2] += d
            casos += 1
            tot_ok += ok
            dif_total = dif + len(diffs_visiveis)
            tot_dif += dif_total
            print(f"{(condo + ' ' + ano)[:28]:28} {ok:>6} {dif_total:>6} {ddet:>11.2f}")
            for rr, cc, gv, mv in diffs_visiveis[:8]:
                print(f"  VIS row={rr} col={cc}: gerado={gv!r} manual={mv!r}")
    print("-" * 56)
    print(f"TOTAL: casos={casos}  linhas OK={tot_ok}  DIF={tot_dif}")
    print(f"\n{'PADROES (linha deterministica divergente)':44}{'#':>4}{'Σ|Δ|':>12}{'netΔ':>12}")
    for lbl, (n, sa, net) in sorted(byline.items(), key=lambda x: -x[1][1]):
        print(f"{lbl[:44]:44}{n:>4}{sa:>12.2f}{net:>12.2f}")


if __name__ == "__main__":
    root = sys.argv[1] if len(sys.argv) > 1 else \
        "/Users/Usuario/Downloads/Quezia - Previsão Orçamentária 3"
    main(root)
