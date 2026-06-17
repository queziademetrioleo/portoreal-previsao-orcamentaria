# -*- coding: utf-8 -*-
"""
Gera o documento final "Previsão <ano>.xlsx" preenchendo o template com
mapeamento inteligente (IA-powered) das contas do condomínio para as
linhas do template.

Diferente da versão anterior (que dependia de matching frágil por nome),
esta versão:
  1. Usa IA para mapear semanticamente cada conta do balanual à linha
     correspondente no template (ou criar nova linha se necessário).
  2. Preenche a PREVISÃO diretamente com os valores de grupo (bypass
     das fórmulas que dependem de estrutura fixa).
  3. Mantém a aritmética R1-R8 100% determinística.

Mecânica do template (decifrada do manual):
  - aba ' C O N T A S ': cada conta tem  I = D - SUM(E:H)
      D = base   E = dedução aprovada   I = valor final
    grupos somam as contas; o restante do arquivo puxa tudo por fórmula.
  - aba ' P R E V I S Ã O ': D22..47 = CONTAS!I<linha do grupo>
  - aba ' P R E V I S Ã O  (2)': VLOOKUPs na PREVISAO + numero de fracoes
  - As fórmulas '[1]DesSin02.rpt' (link externo) são substituídas por literais.
"""
import re, json, datetime, unicodedata, warnings, os
warnings.filterwarnings('ignore')
import openpyxl

ABA_CONTAS = ' C O N T A S '
ABA_PREV = ' P R E V I S Ã O '
ABA_PREV2 = ' P R E V I S Ã O  (2)'

MESES_PT = ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
            'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']


def _norm(s):
    s = str(s or '').lower().strip()
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn')


def _e_formula_externa(v):
    return isinstance(v, str) and v.startswith('=') and 'DesSin02' in v


# ---------------------------------------------------------------------------
# IA-powered mapping: contas do balanual → linhas do template
# ---------------------------------------------------------------------------
def _ia_mapear_contas(linhas, template_path):
    """Usa IA para mapear semanticamente cada conta do balanual
    para uma linha do template CONTAS. Retorna dict: {idx_linha: row_no_template}
    ou None se IA indisponível (usa fallback por nome)."""
    # Construir inventário do template
    wb = openpyxl.load_workbook(template_path)
    ws = wb[ABA_CONTAS]
    template_rows = []
    for r in range(1, ws.max_row + 1):
        code = str(ws.cell(r, 2).value or '').strip()
        nome = str(ws.cell(r, 3).value or '').strip()
        if code or nome:
            template_rows.append({'row': r, 'code': code, 'name': nome})
    wb.close()

    # Prompt para IA
    tpl_inv = "\n".join(
        f"R{tr['row']}: code={tr['code']} | nome={tr['name']}"
        for tr in template_rows if tr['code'] or tr['name']
    )
    contas_inv = "\n".join(
        f"[{i}] grupo={l['grupo']} | classe={l['classe']} | base={l['base']:,.2f} | regra={l['regra']}"
        for i, l in enumerate(linhas)
    )

    prompt = f"""Mapeie cada conta do condomínio para a linha correspondente no template.

CONTAS DO CONDOMÍNIO (a serem mapeadas):
{contas_inv}

LINHAS DO TEMPLATE (destino):
{tpl_inv}

Regras de mapeamento:
- Cada conta do condomínio deve ser mapeada para UMA linha do template
- Use o nome da classe para encontrar o match semântico (ex: "Salário Empregado(s)" casa com "Salário Empregado(s)" no template)
- Se não houver correspondência exata, use a linha mais próxima (mesmo grupo/categoria)
- Se não houver NENHUMA linha adequada, retorne row: null
- Contas com final=0 (totalmente deduzidas) NÃO precisam ser mapeadas
- O código XX.YY do template é a autoridade — se o nome for parecido, é match

Retorne APENAS JSON:
{{"mapeamento": [{{"idx": 0, "row": 66}}, {{"idx": 1, "row": 67}}, ...]}}
"""

    try:
        import sys
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..'))
        import previsao as core
        if not core._ia_disponivel():
            return None
        resp = core._claude_chat(
            "Você é um mapeador de plano de contas contábil. Retorne APENAS JSON válido.",
            prompt, max_tokens=4000
        )
        # Extrair JSON
        m = re.search(r'\{.*\}', resp or '', re.DOTALL)
        if m:
            data = json.loads(m.group(0))
            return {m['idx']: m['row'] for m in data.get('mapeamento', []) if m.get('row')}
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Gerador principal
# ---------------------------------------------------------------------------
def gerar_documento_final(template, destino, nome_condominio, ano, R,
                          impacto_receita_mensal=0.0, num_fracoes=None,
                          inflacao=None, inad_detalhe=None, inad_meta=None):
    """Preenche o template com os números de R (resultado do recálculo
    com decisões humanas) e salva em `destino`.

    Usa IA para mapear contas → template; fallback para matching por nome."""
    wb = openpyxl.load_workbook(template)
    ws = wb[ABA_CONTAS]

    # ------------------------------------------------------------------
    # 1) Indexar o plano de contas do template: código XX.YY -> linha
    # ------------------------------------------------------------------
    contas_rows = {}        # nome normalizado -> linha
    contas_by_code = {}     # código XX.YY -> linha
    receitas_rows = {}      # nome normalizado -> linha (seção 01)
    for r in range(1, ws.max_row + 1):
        code = str(ws.cell(r, 2).value or '').strip()
        if re.match(r'^\d{2}\.\d{2}$', code):
            nome = _norm(ws.cell(r, 3).value)
            contas_by_code[code] = r
            if code.startswith('01.'):
                receitas_rows[nome] = r
            else:
                contas_rows.setdefault(nome, r)

    # ------------------------------------------------------------------
    # 1b) IA mapping: mapeia semanticamente cada conta do balanual → template row
    # ------------------------------------------------------------------
    ia_map = _ia_mapear_contas(R['linhas'], template) if R['linhas'] else {}

    # ------------------------------------------------------------------
    # 1c) LIMPAR valores herdados do template (só as linhas que vamos usar)
    #     Contas sem movimento ficam zeradas.
    # ------------------------------------------------------------------
    linhas_para_limpar = set()
    for i, l in enumerate(R['linhas']):
        row = ia_map.get(i) if ia_map else None
        if row is None:
            nc = _norm(l['classe'])
            row = contas_rows.get(nc)
            if row is None:
                cand = [r for nome, r in contas_rows.items()
                        if nc and (nc in nome or nome in nc)]
                row = cand[0] if cand else None
        if row:
            linhas_para_limpar.add(row)

    # Limpa TODAS as linhas de conta do template (código XX.YY)
    # para evitar carryover de valores do condomínio original do template
    for r in range(1, ws.max_row + 1):
        code = str(ws.cell(r, 2).value or '').strip()
        if re.match(r'^\d{2}\.\d{2}$', code):
            ws.cell(r, 4).value = 0
            for c in (5, 6, 7, 8):
                ws.cell(r, c).value = None

    # ------------------------------------------------------------------
    # 2) DESPESAS: col D = base, col E = dedução
    # ------------------------------------------------------------------
    usado = set()
    nao_mapeadas = []

    for i, l in enumerate(R['linhas']):
        nc = _norm(l['classe'])
        ng = _norm(l['grupo'])

        # Tenta IA primeiro, depois fallback por nome
        row = ia_map.get(i)
        if row is None:
            row = contas_rows.get(nc)
        if row is None:
            cand = [r for nome, r in contas_rows.items()
                    if nome not in usado and (nc in nome or nome in nc)]
            row = cand[0] if cand else None

        if row is None:
            nao_mapeadas.append(l)
            continue

        usado.add(_norm(ws.cell(row, 3).value))
        ws.cell(row, 4).value = round(l['base'], 2)

        # Contas de OBRAS ficam CHEIAS no I (entram no Valor Transportado)
        # e são subtraídas via "Desconsiderações" — não lançar E aqui.
        if 'obras' in ng or 'benfeitoria' in ng:
            continue
        ded = round(l['deducao'], 2)
        ws.cell(row, 5).value = ded if abs(ded) > 0.005 else None

    # Limpar fórmulas externas remanescentes
    for r in range(1, ws.max_row + 1):
        code = str(ws.cell(r, 2).value or '').strip()
        if re.match(r'^\d{2}(\.\d{2})?$', code):
            for c in (4, 5, 6, 7, 8):
                if _e_formula_externa(ws.cell(r, c).value):
                    ws.cell(r, c).value = 0

    # ------------------------------------------------------------------
    # 3) PROVISÕES (R4/R5): col E negativa nas contas de destino
    # ------------------------------------------------------------------
    def _acha(*termos):
        for nome, r in contas_rows.items():
            if all(t in nome for t in termos):
                return r
        # também busca em todas as linhas do template (incluindo grupos)
        for rr in range(1, ws.max_row + 1):
            n = _norm(ws.cell(rr, 3).value)
            if all(t in n for t in termos):
                return rr
        return None

    r_laudo = _acha('laudo', 'autovistoria') or _acha('autovistoria')
    if r_laudo and R['prov_laudo'] > 0.005:
        ws.cell(r_laudo, 4).value = ws.cell(r_laudo, 4).value or 0
        ws.cell(r_laudo, 5).value = -round(R['prov_laudo'], 2)
    r_inc = _acha('sistema', 'combate', 'incendio') or _acha('registro', 'convencao')
    if r_inc and R['prov_incendio'] > 0.005:
        ws.cell(r_inc, 4).value = ws.cell(r_inc, 4).value or 0
        ws.cell(r_inc, 5).value = -round(R['prov_incendio'], 2)

    # Se não achou a conta de provisão, cria linha extra no CONTAS
    nao_mapeadas_prov = []
    if R['prov_laudo'] > 0.005 and not r_laudo:
        nao_mapeadas_prov.append(('Provisão Laudo Autovistoria (R4)', R['prov_laudo']))
    if R['prov_incendio'] > 0.005 and not r_inc:
        nao_mapeadas_prov.append(('Provisão Sist. Incêndio/Registro (R5)', R['prov_incendio']))

    # ------------------------------------------------------------------
    # 4) RECEITAS (seção 01): média mensal por classe
    # ------------------------------------------------------------------
    bal = R['bal']
    for ln in bal['receitas']:
        nc = _norm(ln['classe'])
        r = receitas_rows.get(nc)
        if r is None:
            cand = [row for nome, row in receitas_rows.items()
                    if nc and (nc in nome or nome in nc)]
            r = cand[0] if cand else None
        if r is not None:
            ws.cell(r, 4).value = round(ln['total'] / 12.0, 2)

    # Abate da receita (inadimplência crítica ≥ 3 meses)
    if impacto_receita_mensal > 0.005:
        r_tx = receitas_rows.get(_norm('Tx. Condomínio'))
        if r_tx is None:
            cand = [row for nome, row in receitas_rows.items() if 'condominio' in nome]
            r_tx = cand[0] if cand else None
        if r_tx is not None:
            atual = ws.cell(r_tx, 4).value or 0
            ws.cell(r_tx, 4).value = round(float(atual) - impacto_receita_mensal, 2)

    # Limpar receitas com link externo
    for nome, r in receitas_rows.items():
        if _e_formula_externa(ws.cell(r, 4).value):
            ws.cell(r, 4).value = 0

    # ------------------------------------------------------------------
    # 5) CONFRONTO INICIAL (cabeçalho do CONTAS)
    # ------------------------------------------------------------------
    total_relatorio = (R.get('sin') or {}).get('grand_total') \
        or R['bal'].get('total_despesas') or R['base_total']
    for r in range(1, 10):
        rotulo = _norm(ws.cell(r, 3).value)
        if rotulo == 'valor transportado':
            # Manter fórmula original em D3 (=SUM(I64:I270)/2)
            # Só preencher E3 (Total do relatório) se for fórmula externa
            if _e_formula_externa(ws.cell(r, 5).value) or ws.cell(r, 5).value in (None, 0):
                ws.cell(r, 5).value = round(float(total_relatorio), 2)
        elif rotulo in ('desconsideracoes', 'desconsiderações'):
            ws.cell(r, 4).value = round(R['desconsideracoes'], 2)

    # ------------------------------------------------------------------
    # 6) PREVISÃO: preencher valores diretamente (bypass fórmulas fixas)
    #    Primeiro tenta match por CLASSE (ex: "Luz do Condomínio"),
    #    depois agrupa o restante por GRUPO.
    #    SUBTOTAL e TOTAL são escritos diretamente (fórmulas não calculam no openpyxl).
    # ------------------------------------------------------------------
    if ABA_PREV in wb.sheetnames:
        wsp = wb[ABA_PREV]
        from collections import defaultdict

        # 6a) Construir índice de linhas da PREVISÃO por nome
        #     Só mapeamos linhas da seção de DESPESAS (após cabeçalho "DESPESAS")
        prev_rows = {}       # nome normalizado -> row (contas de despesa)
        summary_rows = {}    # nome -> row (SUBTOTAL, TOTAL, INFLAÇÃO, SALDO)
        despesas_start = None
        for r in range(1, wsp.max_row + 1):
            nome = str(wsp.cell(r, 3).value or '').strip()
            c1 = str(wsp.cell(r, 1).value or '').strip()
            if not nome:
                continue
            nn = _norm(nome)
            # Detectar início da seção de despesas
            if nn == 'despesas' and c1 == '':
                despesas_start = r
                continue
            # Pular cabeçalhos
            if nn in ('receitas', 'valor mensal', 'valor anual'):
                continue
            # Linhas de sumário (após despesas)
            if any(k in nn for k in ('subtotal', 'total', 'infla', 'saldo', 'deficit', 'superavit')):
                summary_rows[nn] = r
                continue
            # Só mapear linhas após o início das despesas
            if despesas_start and r > despesas_start:
                prev_rows[nn] = r

        # 6b) Para cada linha do balanual, tentar match por CLASSE primeiro,
        #     depois por GRUPO. Acumula valores.
        class_final = defaultdict(float)   # {prev_row: valor}
        class_resto = defaultdict(float)   # {grupo_norm: valor} — classes sem row própria
        usado = set()

        for l in R['linhas']:
            nc = _norm(l['classe'])
            ng = _norm(l['grupo'] or '')

            # Tenta match direto por CLASSE na PREVISÃO
            row = prev_rows.get(nc)
            if row is None:
                # Tenta match parcial por classe
                for nome, r in prev_rows.items():
                    if nc and len(nc) > 4 and (nc in nome or nome in nc):
                        row = r
                        break
            if row is not None:
                class_final[row] += l['final']
                usado.add(nc)
            else:
                # Acumula no grupo para distribuir depois
                class_resto[ng] += l['final']

        # 6c) Para classes não mapeadas, tentar match por GRUPO
        for ng, valor in class_resto.items():
            row = prev_rows.get(ng)
            if row is None:
                for nome, r in prev_rows.items():
                    if ng and len(ng) > 4 and (ng in nome or nome in ng):
                        row = r
                        break
            if row is not None:
                class_final[row] += valor

        # 6d) Escrever valores na PREVISÃO D
        prev_preenchidos = set()
        for row, valor in class_final.items():
            if abs(valor) > 0.005:
                wsp.cell(row, 4).value = round(valor, 2)
                prev_preenchidos.add(row)

        # 6e) PROVISÕES (R4/R5) — adicionar na PREVISÃO
        if R['prov_laudo'] > 0.005:
            r = prev_rows.get(_norm('laudo de autovistoria'))
            if not r:
                r = prev_rows.get(_norm('outras despesas diversas'))
            if r:
                atual = wsp.cell(r, 4).value or 0
                if isinstance(atual, (int, float)):
                    wsp.cell(r, 4).value = round(float(atual) + R['prov_laudo'], 2)
                else:
                    wsp.cell(r, 4).value = round(R['prov_laudo'], 2)
                prev_preenchidos.add(r)

        if R['prov_incendio'] > 0.005:
            r = prev_rows.get(_norm('sistema de combate a incendio'))
            if not r:
                r = prev_rows.get(_norm('despesas cartoriais e honorarios'))
            if r:
                atual = wsp.cell(r, 4).value or 0
                if isinstance(atual, (int, float)):
                    wsp.cell(r, 4).value = round(float(atual) + R['prov_incendio'], 2)
                else:
                    wsp.cell(r, 4).value = round(R['prov_incendio'], 2)
                prev_preenchidos.add(r)

        # 6f) SUBTOTAL, INFLAÇÃO e TOTAL — escrever diretamente
        #     (as fórmulas originais não calculam no openpyxl)
        subtotal = R.get('subtotal', sum(v for v in class_final.values()))
        inflacao = R.get('inflacao', 0.10)

        r_subtotal = None
        for nn, r in summary_rows.items():
            if 'subtotal' in nn:
                r_subtotal = r
                break
        if r_subtotal:
            wsp.cell(r_subtotal, 4).value = round(subtotal, 2)
            wsp.cell(r_subtotal, 6).value = round(subtotal / 12, 2)  # mensal

        r_inflacao = None
        for nn, r in summary_rows.items():
            if 'infla' in nn:
                r_inflacao = r
                break
        if r_inflacao:
            wsp.cell(r_inflacao, 4).value = round(subtotal * inflacao, 2)

        r_total = None
        for nn, r in summary_rows.items():
            if nn == 'total':
                r_total = r
                break
        if r_total:
            total = subtotal * (1 + inflacao)
            wsp.cell(r_total, 4).value = round(total, 2)
            wsp.cell(r_total, 5).value = round(total / (
                wb[ABA_PREV2]['E11'].value if ABA_PREV2 in wb.sheetnames else 1), 2)

        # SALDO: receita total - despesa total
        # Receita total está em D19 (TOTAL das receitas)
        r_saldo = None
        for nn, r in summary_rows.items():
            if 'saldo' in nn or 'deficit' in nn or 'superavit' in nn:
                r_saldo = r
                break
        if r_saldo:
            receita = wsp.cell(19, 4).value  # row 19 = TOTAL receitas
            receita_total = float(receita) if isinstance(receita, (int, float)) else 0
            despesa_total = subtotal * (1 + inflacao)
            wsp.cell(r_saldo, 4).value = round(receita_total - despesa_total, 2)

    # ------------------------------------------------------------------
    # 7) TEXTOS: nome do condomínio, ano, data, inflação, frações
    # ------------------------------------------------------------------
    hoje = datetime.date.today()
    data_ext = f'{hoje.day} de {MESES_PT[hoje.month - 1]} de {hoje.year}'

    if ABA_PREV2 in wb.sheetnames:
        ws2 = wb[ABA_PREV2]
        for r in range(1, min(ws2.max_row, 100) + 1):
            v = ws2.cell(r, 1).value
            if isinstance(v, str) and v.startswith('=VLOOKUP(Cadastro'):
                ws2.cell(r, 1).value = nome_condominio
            if isinstance(v, str) and 'PREVISÃO ORÇAMENTÁRIA PARA' in v:
                ws2.cell(r, 1).value = f'PREVISÃO ORÇAMENTÁRIA PARA {ano}'
            for c in range(1, 8):
                vv = ws2.cell(r, c).value
                if isinstance(vv, str) and re.match(r'^\d+ de \w+ de \d{4}$', vv.strip()):
                    ws2.cell(r, c).value = data_ext
        if num_fracoes:
            ws2['E11'] = num_fracoes
        if inflacao is not None:
            ws2['F48'] = inflacao

    if ABA_PREV in wb.sheetnames:
        wsp = wb[ABA_PREV]
        for r in range(1, min(wsp.max_row, 90) + 1):
            v = wsp.cell(r, 1).value
            if isinstance(v, str) and v.startswith('=VLOOKUP(Cadastro'):
                wsp.cell(r, 1).value = nome_condominio
            for c in range(1, 8):
                vv = wsp.cell(r, c).value
                if isinstance(vv, str) and re.match(r'^\d+ de \w+ de \d{4}$', vv.strip()):
                    wsp.cell(r, c).value = data_ext

    # ------------------------------------------------------------------
    # 8) ABA INADIMPLÊNCIA — detalhe por unidade
    # ------------------------------------------------------------------
    if inad_detalhe:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        thin = Side(style='thin', color='B0B7C3')
        borda = Border(left=thin, right=thin, top=thin, bottom=thin)
        azul = PatternFill('solid', fgColor='1F3864')
        amarelo = PatternFill('solid', fgColor='FFF2CC')
        branco_negrito = Font(bold=True, color='FFFFFF')

        wsi = wb.create_sheet('Inadimplência')
        wsi['A1'] = f'INADIMPLÊNCIA — {nome_condominio}'
        wsi['A1'].font = Font(bold=True, size=13)
        if inad_meta:
            wsi['A2'] = (f"Data-base: {inad_meta.get('data_base', '')}  |  "
                         f"Critério: a partir de 3 meses de atraso a taxa da unidade "
                         f"é abatida da receita mensal prevista")
            wsi['A2'].font = Font(italic=True, color='5A6472')

        cab = ['Unidade / Devedor', 'Classe', 'Mês Ref.', 'Vencimento', 'Valor (R$)',
               'Meses de atraso', 'Situação', 'Decisão aplicada']
        for j, h in enumerate(cab, 1):
            c = wsi.cell(4, j, h)
            c.fill = azul; c.font = branco_negrito; c.border = borda
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for w, col in zip([34, 22, 10, 12, 12, 13, 18, 26], 'ABCDEFGH'):
            wsi.column_dimensions[col].width = w

        r = 5
        for it in sorted(inad_detalhe, key=lambda x: -(x.get('meses_atraso') or 0)):
            wsi.cell(r, 1, it.get('unidade')).border = borda
            wsi.cell(r, 2, it.get('classe')).border = borda
            wsi.cell(r, 3, it.get('mes_ref')).border = borda
            wsi.cell(r, 4, str(it.get('vencimento') or '')).border = borda
            c = wsi.cell(r, 5, it.get('valor')); c.number_format = '#,##0.00'; c.border = borda
            wsi.cell(r, 6, it.get('meses_atraso')).border = borda
            critica = bool(it.get('critica'))
            wsi.cell(r, 7, 'CRÍTICA (≥ 3 meses)' if critica else 'recente (< 3 meses)').border = borda
            dec = it.get('decisao')
            wsi.cell(r, 8, 'Abatida da receita prevista' if dec == 'abater'
                     else 'Mantida na receita (não abate)').border = borda
            if critica:
                for j in range(1, 9):
                    wsi.cell(r, j).fill = amarelo
            r += 1

        r += 1
        tot = sum(i.get('valor') or 0 for i in inad_detalhe)
        crit = sum(i.get('valor') or 0 for i in inad_detalhe if i.get('critica'))
        resumo_rows = [
            ('Total em aberto', tot),
            ('Crítica (≥ 3 meses)', crit),
            ('Recente (< 3 meses)', tot - crit),
            ('Impacto mensal na receita prevista (unidades abatidas)', impacto_receita_mensal),
        ]
        for rotulo, val in resumo_rows:
            wsi.cell(r, 4, rotulo).font = Font(bold=True)
            c = wsi.cell(r, 5, round(val, 2)); c.number_format = '#,##0.00'; c.font = Font(bold=True)
            r += 1
        wsi.cell(r + 1, 1, 'A taxa mensal de cada unidade abatida foi estimada pela média '
                           'das parcelas em aberto da própria unidade.').font = Font(italic=True, color='5A6472')

    wb.save(destino)
    return {'nao_mapeadas': [(l['grupo'], l['classe'], l['base']) for l in nao_mapeadas],
            'nao_mapeadas_prov': nao_mapeadas_prov}
