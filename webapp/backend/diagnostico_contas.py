#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Diagnostico da aba CONTAS de uma previsao manual.

Objetivo: transformar a planilha manual em uma memoria de calculo auditavel,
preservando formulas, valores salvos pelo Excel, dependencias e linhas usadas
para alimentar a previsao.
"""
from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries


METRICAS_CHAVE = {
    'valor transportado',
    'desconsideracoes',
    'desconsiderações',
    'subtotal inicial',
    'subtotal atual',
    'inflacao',
    'inflação',
    'total previsto',
    'saldo',
}


def _norm(s: Any) -> str:
    s = str(s or '').lower().strip()
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )


def _serializar_valor(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def _numero(v: Any) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def _achar_aba_contas(wb) -> str:
    for name in wb.sheetnames:
        if 'contas' in _norm(name).replace(' ', ''):
            return name
    raise ValueError('Aba CONTAS nao encontrada no arquivo')


def _refs_formula(formula: str) -> dict[str, list[str]]:
    """Extrai referencias simples de celulas/ranges de uma formula Excel."""
    if not formula or not formula.startswith('='):
        return {'cells': [], 'ranges': [], 'sheets': []}

    sheets = []
    for quoted in re.findall(r"'([^']+)'!", formula):
        if quoted not in sheets:
            sheets.append(quoted)
    for plain in re.findall(r"(?<![A-Z0-9_])([A-Za-z0-9_ ]+)!", formula):
        sheet = plain.strip()
        if sheet and sheet not in sheets and not re.match(r'^[A-Z]+$', sheet):
            sheets.append(sheet)

    ranges = []
    for ref in re.findall(r'(\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)', formula):
        clean = ref.replace('$', '')
        if clean not in ranges:
            ranges.append(clean)

    formula_sem_ranges = formula
    for ref in ranges:
        formula_sem_ranges = formula_sem_ranges.replace(ref, '')
        formula_sem_ranges = formula_sem_ranges.replace(ref.replace('$', ''), '')

    cells = []
    for ref in re.findall(r'(?<![A-Z])\$?([A-Z]{1,3})\$?(\d+)', formula_sem_ranges):
        clean = f'{ref[0]}{ref[1]}'
        if clean not in cells:
            cells.append(clean)

    return {'cells': cells, 'ranges': ranges, 'sheets': sheets}


def _avaliador_simples(ws_formula):
    """Avaliador conservador para formulas comuns da aba CONTAS.

    Nao pretende substituir o Excel. Serve para diagnostico quando o arquivo nao
    tem valores recalculados salvos. Formulas desconhecidas retornam None.
    """
    cache: dict[str, float | None] = {}
    stack: set[str] = set()

    def valor_cell(addr: str) -> float | None:
        addr = addr.replace('$', '')
        if addr in cache:
            return cache[addr]
        if addr in stack:
            return None
        stack.add(addr)
        v = ws_formula[addr].value
        if isinstance(v, (int, float)):
            out = float(v)
        elif isinstance(v, str) and v.startswith('='):
            out = avaliar(v)
        else:
            out = 0.0 if v in (None, '') else None
        cache[addr] = out
        stack.discard(addr)
        return out

    def soma_range(ref: str) -> float | None:
        min_col, min_row, max_col, max_row = range_boundaries(ref.replace('$', ''))
        total = 0.0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                v = valor_cell(f'{get_column_letter(col)}{row}')
                if v is None:
                    return None
                total += v
        return total

    def avaliar(formula: str) -> float | None:
        expr = formula.strip()
        if not expr.startswith('='):
            return None
        expr = expr[1:]
        if '!' in expr or re.search(r'\b(VLOOKUP|PROCV|IF|SE|SUBTOTAL)\b', expr, re.I):
            return None

        def repl_sum(match):
            total = soma_range(match.group(1))
            return 'None' if total is None else str(total)

        expr = re.sub(r'SUM\((\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)\)', repl_sum, expr, flags=re.I)
        expr = re.sub(r'SOMA\((\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+)\)', repl_sum, expr, flags=re.I)

        def repl_cell(match):
            v = valor_cell(match.group(0))
            return 'None' if v is None else str(v)

        expr = re.sub(r'\$?[A-Z]{1,3}\$?\d+', repl_cell, expr)
        if 'None' in expr:
            return None
        if not re.fullmatch(r'[0-9\.\+\-\*\/\(\) ]+', expr):
            return None
        try:
            return float(eval(expr, {'__builtins__': {}}, {}))
        except Exception:
            return None

    return avaliar


def extrair_contas(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    wb_valores = openpyxl.load_workbook(path, data_only=True)
    sheet_name = _achar_aba_contas(wb_formula)
    ws_f = wb_formula[sheet_name]
    ws_v = wb_valores[sheet_name]
    avaliar = _avaliador_simples(ws_f)

    merged = [str(rng) for rng in ws_f.merged_cells.ranges]
    colunas_ocultas = [
        col for col, dim in ws_f.column_dimensions.items()
        if getattr(dim, 'hidden', False)
    ]
    linhas_ocultas = [
        row for row, dim in ws_f.row_dimensions.items()
        if getattr(dim, 'hidden', False)
    ]

    celulas = []
    formulas = []
    metricas = []
    linhas = []

    for r in range(1, ws_f.max_row + 1):
        linha_cells = {}
        tem_conteudo = False
        for c in range(1, ws_f.max_column + 1):
            addr = f'{get_column_letter(c)}{r}'
            vf = ws_f.cell(r, c).value
            vv = ws_v.cell(r, c).value
            if vf is None and vv is None:
                continue
            tem_conteudo = True
            formula = vf if isinstance(vf, str) and vf.startswith('=') else None
            calculado = avaliar(formula) if formula else None
            payload = {
                'addr': addr,
                'row': r,
                'col': c,
                'col_letter': get_column_letter(c),
                'value': _serializar_valor(vf if not formula else None),
                'formula': formula,
                'cached_value': _serializar_valor(vv),
                'calculated_value': round(calculado, 8) if isinstance(calculado, float) else None,
                'number_format': ws_f.cell(r, c).number_format,
                'hidden_row': r in linhas_ocultas,
                'hidden_col': get_column_letter(c) in colunas_ocultas,
            }
            celulas.append(payload)
            linha_cells[get_column_letter(c)] = payload
            if formula:
                refs = _refs_formula(formula)
                formulas.append({
                    'addr': addr,
                    'formula': formula,
                    'cached_value': _serializar_valor(vv),
                    'calculated_value': payload['calculated_value'],
                    'dependencies': refs,
                    'resolved_by_simple_engine': payload['calculated_value'] is not None,
                })

        if not tem_conteudo:
            continue

        label = ws_f.cell(r, 3).value
        nlabel = _norm(label)
        if nlabel in METRICAS_CHAVE:
            metricas.append({
                'row': r,
                'label': str(label).strip(),
                'formula_or_value': _serializar_valor(ws_f.cell(r, 4).value),
                'cached_value': _serializar_valor(ws_v.cell(r, 4).value),
                'dependencies': _refs_formula(ws_f.cell(r, 4).value)
                if isinstance(ws_f.cell(r, 4).value, str) else {'cells': [], 'ranges': [], 'sheets': []},
            })

        codigo = ws_f.cell(r, 2).value
        if codigo is not None or label is not None:
            linhas.append({
                'row': r,
                'codigo': str(codigo).strip() if codigo is not None else '',
                'label': str(label).strip() if label is not None else '',
                'D_base': _serializar_valor(ws_v.cell(r, 4).value),
                'E_ajuste': _serializar_valor(ws_v.cell(r, 5).value),
                'F_ajuste': _serializar_valor(ws_v.cell(r, 6).value),
                'G_ajuste': _serializar_valor(ws_v.cell(r, 7).value),
                'H_ajuste': _serializar_valor(ws_v.cell(r, 8).value),
                'I_final': _serializar_valor(ws_v.cell(r, 9).value),
                'I_formula': ws_f.cell(r, 9).value
                if isinstance(ws_f.cell(r, 9).value, str) and ws_f.cell(r, 9).value.startswith('=')
                else None,
            })

    valor_transportado = next(
        (m for m in metricas if _norm(m['label']) == 'valor transportado'),
        None,
    )
    soma_i = sum(
        _numero(l.get('I_final'))
        for l in linhas
        if isinstance(l.get('I_final'), (int, float))
    )

    return {
        'arquivo': str(path),
        'sheet': sheet_name,
        'dimensions': {
            'max_row': ws_f.max_row,
            'max_column': ws_f.max_column,
            'merged_ranges': merged,
            'hidden_rows': linhas_ocultas,
            'hidden_columns': colunas_ocultas,
        },
        'summary': {
            'non_empty_cells': len(celulas),
            'formulas': len(formulas),
            'rows_with_labels': len(linhas),
            'key_metrics': len(metricas),
            'valor_transportado': valor_transportado,
            'soma_coluna_i_visivel_extraida': round(soma_i, 2),
        },
        'metricas': metricas,
        'linhas': linhas,
        'formulas': formulas,
        'celulas': celulas,
    }


def salvar_diagnostico(path: str | Path, destino_json: str | Path) -> dict[str, Any]:
    diag = extrair_contas(path)
    destino_json = Path(destino_json)
    destino_json.parent.mkdir(parents=True, exist_ok=True)
    destino_json.write_text(json.dumps(diag, ensure_ascii=False, indent=2), encoding='utf-8')
    return diag


def gerar_markdown(diag: dict[str, Any]) -> str:
    linhas = [
        f"# Diagnóstico CONTAS",
        '',
        f"Arquivo: `{diag['arquivo']}`",
        f"Aba: `{diag['sheet']}`",
        '',
        '## Resumo',
        f"- Células preenchidas: {diag['summary']['non_empty_cells']}",
        f"- Fórmulas: {diag['summary']['formulas']}",
        f"- Linhas com rótulo/código: {diag['summary']['rows_with_labels']}",
        f"- Métricas-chave: {diag['summary']['key_metrics']}",
        '',
        '## Métricas-chave',
    ]
    for m in diag['metricas']:
        linhas.append(
            f"- Linha {m['row']}: **{m['label']}** = `{m['formula_or_value']}` "
            f"=> `{m['cached_value']}`"
        )
    linhas.extend(['', '## Fórmulas principais'])
    for f in diag['formulas'][:80]:
        deps = f["dependencies"]
        linhas.append(
            f"- `{f['addr']}` `{f['formula']}` => `{f['cached_value']}` "
            f"refs células={deps['cells']} ranges={deps['ranges']} sheets={deps['sheets']}"
        )
    if len(diag['formulas']) > 80:
        linhas.append(f"- ... {len(diag['formulas']) - 80} fórmulas adicionais no JSON.")
    return '\n'.join(linhas) + '\n'
