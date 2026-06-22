#!/usr/bin/env python3
"""Debug script to compare auto-generated PREVISAO tab with manual file."""
import sys, os, warnings, tempfile
warnings.filterwarnings('ignore')
sys.path.insert(0, '/Users/Usuario/PrevisaoOrcamentaria')
sys.path.insert(0, '/Users/Usuario/PrevisaoOrcamentaria/webapp/backend')
os.chdir('/Users/Usuario/PrevisaoOrcamentaria')
import previsao as core
from gerador_previsao import gerar_previsao_adaptativa
import openpyxl

folder = '/Users/Usuario/Downloads/Quezia - Previsão Orçamentária 3/Sophia I/Sophia I 2024'
template = '/Users/Usuario/PrevisaoOrcamentaria/webapp/backend/templates/modelo_previsao.xlsx'
manual_path = '/Users/Usuario/Downloads/Quezia - Previsão Orçamentária 3/Sophia I/Sophia I 2024/Previsão 2024.xlsx'

R = core.analisar(folder)
inad = R.get('inad')
inad_itens = inad['itens'] if inad and isinstance(inad, dict) else None
inad_meta = {'total': inad['total'], 'critica': 0, 'data_base': str(inad.get('data_base',''))} if inad and isinstance(inad, dict) else None

with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
    destino = tmp.name

gerar_previsao_adaptativa(destino, R, 'Sophia I', 2024,
    num_fracoes=None, inflacao=core.INFLACAO,
    inad_detalhe=inad_itens, inad_meta=inad_meta, referencia=template)

print(f'Generated: {destino}')

wb_a = openpyxl.load_workbook(destino, data_only=True)
wb_m = openpyxl.load_workbook(manual_path, data_only=True)

ws_a = wb_a[' P R E V I S Ã O ']
ws_m = wb_m[' P R E V I S Ã O ']

print('=== AUTO-GENERATED ===')
for row in ws_a.iter_rows(min_row=1, max_row=60, max_col=10, values_only=False):
    vals = [(c.column, c.value) for c in row if c.value is not None]
    if vals:
        print(f'Row {row[0].row}: {vals}')

print()
print('=== MANUAL ===')
for row in ws_m.iter_rows(min_row=1, max_row=60, max_col=10, values_only=False):
    vals = [(c.column, c.value) for c in row if c.value is not None]
    if vals:
        print(f'Row {row[0].row}: {vals}')
