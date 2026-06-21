#!/usr/bin/env python3
# -*- coding: utf-8 -*-
CRITERIOS = """
GERADOR DE PREVISAO ORCAMENTARIA
Uso:  python3 gerar_previsao.py "Sophia I/Sophia I 2026"
      python3 gerar_previsao.py --todos     (processa as 4 pastas 2026)

Le os relatorios do Condominio21 (balanual, desbai06, dessin02, inad01) e gera
"Previsao AUTO - <condominio>.xlsx" replicando a metodologia manual aprendida
dos arquivos Previsao 20XX.xlsx historicos:

REGRAS APRENDIDAS DO HISTORICO (2022-2026, 4 condominios):
 R1. Grupo "Despesas com Obras/Benfeitorias" -> desconsiderado integralmente
     (linha Desconsideracoes do Confronto Inicial).
 R2. Rescisao, Indenizacao Trabalhista e Pensao Alimenticia -> deduzidas 100%
     (eventos pontuais de pessoal; ratio 100% em todos os casos historicos).
 R3. Manutencoes "lumpy" (pintura, portao, eletrica, hidraulica, pequenas
     reformas, cameras, antena, interfone, maquinas, jardim, dedetizacao,
     fossa) -> deduzir a parcela extraordinaria. Calculo: soma dos lancamentos
     do desbai06 classificados como extraordinarios; fallback = fator mediano
     historico de deducao da conta.
 R4. Grupo "Despesas Diversas" exceto Seguro Incendio -> deduzido integralmente
     e REALOCADO como provisao para Laudo de Autovistoria Predial (mesmo valor).
 R5. Grupo "Despesas Cartoriais e Honorarios" -> deduzido e REALOCADO como
     provisao para Sistema de Combate a Incendio / Registro da Convencao.
 R6. Contratos, Pro-labore e Taxa de Administracao -> previsao = ultimo valor
     mensal vigente x 12 (anualizacao da tarifa atual, nao a soma historica).
 R7. Reajuste de inflacao: +IPCA 4.72% sobre o subtotal (ou PREVISAO_INFLACAO_PCT).
 R8. Inadimplencia: nao entra como despesa; reportada como risco de caixa,
     com regua de criticidade > 3 meses da data-base do inad01.
"""
"""
PREVISAO ORCAMENTARIA DE CONDOMINIOS — ferramenta autonoma
Como usar:
    python3 previsao.py
    (o programa pergunta a pasta — arraste a pasta do condominio para o terminal e de Enter)
Ou direto:
    python3 previsao.py "/caminho/da/pasta do condominio"
A pasta deve conter os relatorios exportados do Condominio21:
    balanual.xls  desbai06.xls  [dessin02.xls]  [inad01.xls]  [Previsao XXXX.xlsx]
(entre colchetes = opcionais)
"""
import sys, os, glob, re, datetime, shlex, time
import warnings
warnings.filterwarnings('ignore')
import logging
logger = logging.getLogger('previsao')
import xlrd, openpyxl, unicodedata
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===================== CAMADA DE IA (Anthropic OU OpenAI) =====================
# A IA AUXILIA: sugere classificacao p/ itens ambiguos ("Revisar") e escreve um
# parecer executivo. As regras R1-R8 continuam deterministicas — a IA nao
# altera numeros, apenas adiciona sugestoes e analise.
#
# Provedores suportados (escolha automatica pela chave disponivel):
#   Anthropic: ANTHROPIC_API_KEY ou arquivo chave_claude.txt  (prioridade)
#   OpenAI:    OPENAI_API_KEY    ou arquivo chave_openai.txt
# Forcar provedor: PREVISAO_IA_PROVEDOR=anthropic|openai
# Modelo: PREVISAO_IA_MODELO (default: claude-opus-4-8 / gpt-5.4)
CLAUDE_MODEL = os.environ.get('PREVISAO_IA_MODELO', 'claude-opus-4-8')
OPENAI_MODEL = os.environ.get('PREVISAO_IA_MODELO_OPENAI',
                              os.environ.get('PREVISAO_IA_MODELO', '') or 'gpt-5.4')
if OPENAI_MODEL.startswith('claude'):
    OPENAI_MODEL = 'gpt-5.4'


def _le_chave_arquivo(nome):
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), nome)
    if os.path.exists(p):
        for line in open(p, encoding='utf-8'):
            line = line.strip()
            if line and not line.startswith('#'):
                return line
    return None


def _claude_key():
    return os.environ.get('ANTHROPIC_API_KEY', '').strip() or _le_chave_arquivo('chave_claude.txt')


def _openai_key():
    return os.environ.get('OPENAI_API_KEY', '').strip() or _le_chave_arquivo('chave_openai.txt')


def _ia_provedor():
    """Retorna 'anthropic', 'openai' ou None (sem chave)."""
    forcado = os.environ.get('PREVISAO_IA_PROVEDOR', '').strip().lower()
    if forcado in ('anthropic', 'openai'):
        return forcado
    if _claude_key():
        return 'anthropic'
    if _openai_key():
        return 'openai'
    return None


def _ia_modelo():
    """Modelo do provedor EFETIVO (considera falhas de billing ja detectadas)."""
    prov = _ia_provedor()
    if prov in _PROVEDOR_FALHOU:
        alt = 'openai' if prov == 'anthropic' else 'anthropic'
        if alt not in _PROVEDOR_FALHOU and (_openai_key() if alt == 'openai' else _claude_key()):
            prov = alt
    return CLAUDE_MODEL if prov == 'anthropic' else OPENAI_MODEL


def _ia_disponivel():
    return _ia_provedor() is not None

def _anthropic_chat(system, user, max_tokens, temperature=None):
    """POST /v1/messages da Anthropic (sem SDK). Retorna texto ou levanta excecao."""
    import json as _json, urllib.request
    body = {'model': CLAUDE_MODEL, 'max_tokens': max_tokens,
            'system': system,
            'messages': [{'role': 'user', 'content': user}]}
    if temperature is not None:
        body['temperature'] = temperature
    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=_json.dumps(body).encode('utf-8'),
        headers={'Content-Type': 'application/json',
                 'x-api-key': _claude_key(),
                 'anthropic-version': '2023-06-01'})
    with urllib.request.urlopen(req, timeout=300) as r:
        data = _json.loads(r.read().decode('utf-8'))
    return ''.join(b.get('text', '') for b in data.get('content', [])
                   if b.get('type') == 'text')


def _openai_chat(system, user, max_tokens, temperature=None):
    """POST /v1/chat/completions da OpenAI (sem SDK). Retorna texto ou levanta excecao."""
    import json as _json, urllib.request
    body = {'model': OPENAI_MODEL,
            'max_completion_tokens': max_tokens,
            'messages': [{'role': 'system', 'content': system},
                         {'role': 'user', 'content': user}]}
    if temperature is not None:
        body['temperature'] = temperature
    req = urllib.request.Request(
        'https://api.openai.com/v1/chat/completions',
        data=_json.dumps(body).encode('utf-8'),
        headers={'Content-Type': 'application/json',
                 'Authorization': f'Bearer {_openai_key()}'})
    with urllib.request.urlopen(req, timeout=300) as r:
        data = _json.loads(r.read().decode('utf-8'))
    return data['choices'][0]['message']['content'] or ''


_PROVEDOR_FALHOU = set()   # provedores que falharam nesta execucao (ex.: sem credito)


def _claude_chat(system, user, max_tokens=4000, temperature=None):
    """Chat com o provedor de IA ativo (Anthropic ou OpenAI). Retorna texto ou None.
    Nome mantido por compatibilidade — roteia conforme _ia_provedor().
    Se o provedor preferido falhar (ex.: sem credito), tenta o outro automaticamente."""
    prov = _ia_provedor()
    if prov is None:
        return None
    ordem = [prov] + [p for p in ('anthropic', 'openai') if p != prov]
    for p in ordem:
        if p in _PROVEDOR_FALHOU:
            continue
        if p == 'anthropic' and not _claude_key():
            continue
        if p == 'openai' and not _openai_key():
            continue
        try:
            fn = _openai_chat if p == 'openai' else _anthropic_chat
            return fn(system, user, max_tokens, temperature)
        except Exception as e:
            _PROVEDOR_FALHOU.add(p)
            logger.warning('IA %s falhou (%s: %s) — tentando alternativa...', p, type(e).__name__, e)
    logger.warning('Nenhum provedor de IA disponivel — seguindo so com as regras.')
    return None

def _extrai_json(texto):
    """Extrai objeto JSON da resposta (tolera cercas ```json ... ```)."""
    import json as _json
    t = texto.strip()
    if t.startswith('```'):
        t = re.sub(r'^```[a-zA-Z]*\s*', '', t)
        t = re.sub(r'\s*```$', '', t)
    ini, fim = t.find('{'), t.rfind('}')
    if ini >= 0 and fim > ini:
        t = t[ini:fim + 1]
    return _json.loads(t)

def _recupera_itens_json(texto):
    """Recupera objetos {id,sugestao,justificativa} de um JSON truncado/malformado.
    Varre o texto por objetos individuais bem-formados, ignorando o que quebrou."""
    import json as _json
    itens = []
    for m in re.finditer(r'\{[^{}]*?"id"\s*:\s*\d+[^{}]*?\}', texto):
        try:
            itens.append(_json.loads(m.group(0)))
        except Exception:
            continue
    return itens

IA_SISTEMA_CLASSIF = """\
Voce e um analista financeiro de condominios no Brasil especializado em previsao orcamentaria.

TAREFA: Para cada lancamento, classificar como:
- "Recorrente": mantem na base — despesa periodica/recorrente do condominio
- "Extraordinaria": remove da base — evento pontual claro, nao se repetira
- "Revisar": ambiguo, impossivel decidir sem contexto humano

REGRA PRINCIPAL — SEJA CONSERVADOR:
Na duvida, prefira "Recorrente". So marque "Extraordinaria" quando ha evidencia clara e inequivoca de que e um evento unico.

EXTRAORDINARIA (evidencia inequivoca):
- Empresa de pintura/reforma contratada para projeto especifico com NFs em serie ou "Sinal"/"Parcela X/Y"/"Restante"
- Compra de EQUIPAMENTO NOVO identificavel pelo nome: amplificador, camera completa, motor de portao, bomba nova, placa de ramais, central nova
- Parcelas de grande reparo documentado (ex: "Reparo PC de energia blocos A e B parcela X/Y")
- Material de obra em volume claramente acima do normal (argamassa, pastilhas, estrutural)
- Rescisao trabalhista, indenizacao, processo judicial

RECORRENTE (manter na base — incluindo casos ambiguos):
- "Aquisicao de material" sem indicar equipamento novo especifico = provavelmente reposicao/manutencao = RECORRENTE
- Material de pintura (tinta, rolo, lixa, sika) independente do valor = manutencao preventiva corriqueira = RECORRENTE
- Visitas mensais/periodicas de qualquer tipo de manutencao
- Pequenas compras de material eletrico, hidraulico, de seguranca
- Revisoes periodicas (trimestral, semestral, anual) de qualquer sistema

CALIBRACAO:
Se houver valores de referencia do calculo manual do especialista, use-os para calibrar.
Quando uma classe tem muitos itens e o total de "Extraordinaria" que voce calcularia excederia
significativamente o que o especialista historicamente remove, seja mais conservador e
reclassifique itens ambiguos para "Recorrente".

Responda APENAS JSON: {"itens": [{"id": <numero>, "sugestao": "Recorrente|Extraordinaria|Revisar", "justificativa": "<=15 palavras"}]}
"""

def ia_classificar_revisar(itens, nome_condo, manual=None):
    """Sugestao da IA para itens 'Revisar'. Retorna {indice_do_item: (sugestao, justificativa)}.
    manual: resultado de parse_previsao() — usado como referencia de calibracao."""
    rev = [(i, it) for i, it in enumerate(itens) if it['cat'] == 'Revisar']
    if not rev or not _ia_disponivel():
        return {}
    logger.info('🤖 IA analisando %d itens ambiguos (%s)...', len(rev), _ia_modelo())

    # Contexto de referencia: valores de cada classe no calculo manual do especialista
    ref_ctx = ''
    if manual and manual.get('contas'):
        linhas_ref = [ct for ct in manual['contas'] if ct['base'] and abs(ct['base']) > 0.5]
        if linhas_ref:
            ref_ctx = ('\n\nREFERENCIA DO ESPECIALISTA (calculo manual aprovado para este condominio):\n'
                       'Estes sao os valores finais que o especialista definiu para cada classe '
                       '(base = total do periodo, ajuste = reducao aplicada pelo especialista):\n')
            for ct in linhas_ref[:30]:
                ajuste_str = f'  ajuste={ct["ajuste"]:,.2f}' if ct['ajuste'] is not None else ''
                ref_ctx += f'  {ct["nome"]}: base={ct["base"]:,.2f}{ajuste_str}\n'
            ref_ctx += ('Use esses valores como calibracao: se sua soma de "Extraordinaria" '
                        'para uma classe exceder o ajuste do especialista, seja mais conservador '
                        'e reclassifique itens ambiguos para "Recorrente".\n')

    out = {}
    LOTE = 25  # lotes menores -> menos risco de truncar o JSON da resposta
    n_lotes = (len(rev) + LOTE - 1) // LOTE
    for nlote, k in enumerate(range(0, len(rev), LOTE), 1):
        lote = rev[k:k + LOTE]
        linhas = '\n'.join(
            f"id={j} | grupo={it['grupo']} | classe={it['classe']} | "
            f"valor=R${it['valor_pago']:.2f} | meses_com_gasto_na_classe={it.get('n_meses','?')} | "
            f"descricao={(it['descricao'] or '')[:100]}"
            for j, (i, it) in enumerate(lote))
        prompt = f'Condominio: {nome_condo}{ref_ctx}\n\nLancamentos:\n{linhas}'
        # Cada lote e independente: ate 3 tentativas; NUNCA aborta os demais lotes.
        itens_lote = []
        for tentativa in range(3):
            resp = _claude_chat(IA_SISTEMA_CLASSIF, prompt, max_tokens=8000)
            if not resp:
                continue
            try:
                itens_lote = _extrai_json(resp).get('itens', [])
            except Exception:
                itens_lote = _recupera_itens_json(resp)  # JSON truncado -> recupera o que der
            if itens_lote:
                logger.info('Lote %d/%d: %d sugestoes recebidas', nlote, n_lotes, len(itens_lote))
                break
        if not itens_lote:
            logger.warning('Lote %d/%d sem resposta valida apos 3 tentativas (%d itens ficam para revisao humana).',
                           nlote, n_lotes, len(lote))
        for d in itens_lote:
            try:
                j = int(d.get('id', -1))
            except (ValueError, TypeError):
                continue
            if 0 <= j < len(lote):
                idx = lote[j][0]
                sug = str(d.get('sugestao', '')).strip().capitalize()
                if sug not in ('Recorrente', 'Extraordinaria', 'Revisar'):
                    sug = 'Revisar'
                out[idx] = (sug, str(d.get('justificativa', ''))[:160])
    return out

def ia_parecer(R, nome_condo):
    """Parecer executivo escrito pela IA a partir dos numeros calculados."""
    if not _ia_disponivel():
        return None
    logger.info('🤖 IA escrevendo parecer executivo...')
    ex_top = defaultdict(float)
    for it in R['des']['itens']:
        if it['cat'] == 'Extraordinaria':
            ex_top[f"{it['grupo']} > {it['classe']}"] += it['valor_pago']
    ex_txt = '\n'.join(f'  - {k}: R${v:,.2f}' for k, v in
                       sorted(ex_top.items(), key=lambda x: -x[1])[:8])
    inad_txt = 'sem inadimplencia registrada'
    if R['inad']:
        inad_txt = (f"total R${R['inad']['total']:,.2f}; critica (>=3 meses) "
                    f"R${R['inad']['critica']:,.2f} ({R['inad']['unidades_criticas']} unidade(s)); "
                    f"impacto mensal na receita R${R['inad']['impacto_mensal_receita']:,.2f}")
    man_txt = 'sem previsao manual para comparar'
    if R['manual']:
        cf = R['manual']['confronto']
        man = cf.get('subtotal atual') or cf.get('subtotal inicial') or 0
        if man:
            man_txt = f"subtotal manual R${man:,.2f} vs automatico R${R['subtotal']:,.2f}"
    bal = R['bal']
    resumo = (
        f"Condominio: {nome_condo}\n"
        f"Despesa total 12 meses: R${R['base_total']:,.2f}\n"
        f"Receita total 12 meses: R${bal['total_receitas'] or 0:,.2f}\n"
        f"Saldo inicial: R${bal['saldo_inicial'] or 0:,.2f} | final: R${bal['saldo_final'] or 0:,.2f}\n"
        f"Desconsideracoes (obras): R${R['desconsideracoes']:,.2f}\n"
        f"Provisao laudo autovistoria: R${R['prov_laudo']:,.2f}\n"
        f"Provisao sist. incendio/registro: R${R['prov_incendio']:,.2f}\n"
        f"Base recorrente ajustada (subtotal): R${R['subtotal']:,.2f}\n"
        f"Previsao anual (+{INFLACAO:.0%}): R${R['total_previsto']:,.2f} "
        f"(mensal R${R['total_previsto']/12:,.2f})\n"
        f"Receita mensal media atual: R${(bal['total_receitas'] or 0)/12:,.2f}\n"
        f"Principais despesas extraordinarias removidas:\n{ex_txt or '  (nenhuma)'}\n"
        f"Inadimplencia: {inad_txt}\n"
        f"Comparacao com previsao manual: {man_txt}\n")
    return _claude_chat(
        'Voce e um consultor financeiro de condominios. Escreva um PARECER EXECUTIVO '
        'em portugues do Brasil (350-500 palavras), claro para sindicos leigos, sobre a '
        'previsao orcamentaria a seguir. Estruture em: 1) Situacao atual; 2) O que foi '
        'ajustado e por que; 3) Riscos (inadimplencia, deficit, despesas volateis); '
        '4) Recomendacoes praticas (taxa condominial, fundo de reserva, provisoes). '
        'Use os numeros fornecidos; nao invente valores. Responda apenas com o texto '
        'do parecer, sem markdown.',
        resumo, max_tokens=2000)


# ----------------------------------------------------------------------------
def _xldate(wb, v):
    try:
        return datetime.datetime(*xlrd.xldate_as_tuple(v, wb.datemode)).date()
    except Exception:
        return None

def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return 0.0
    s = str(v).strip().replace('.', '').replace(',', '.') if ',' in str(v) else str(v).strip()
    try:
        return float(s)
    except Exception:
        return 0.0

def _norm(s):
    s = str(s or '').lower().strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s

# ----------------------------------------------------------------------------
# DESBAI06 - analitico de despesas baixadas (base principal de classificacao)
# ----------------------------------------------------------------------------
def parse_desbai(path):
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    # localizar linha de cabecalho ("Classe da Conta")
    start = None
    for r in range(sh.nrows):
        if 'classe da conta' in _norm(sh.cell_value(r, 2)):
            start = r + 1
            break
    if start is None:
        start = 22
    itens = []
    classe_totais = {}   # (grupo,classe) -> total relatorio
    grupo_totais = {}
    grand_total = None
    cur_grupo = None
    cur_classe = None
    next_is_group = True
    for r in range(start, sh.nrows):
        c0 = str(sh.cell_value(r, 0)).strip()
        c4 = sh.cell_value(r, 4)
        c6 = str(sh.cell_value(r, 6)).strip()
        c7 = sh.cell_value(r, 7)
        c8 = sh.cell_value(r, 8)
        c9 = str(sh.cell_value(r, 9)).strip()
        # totais
        n6 = _norm(c6)
        if n6.startswith('total classe'):
            classe_totais[(cur_grupo, cur_classe)] = _num(c7)
            continue
        if n6.startswith('total grupo'):
            grupo_totais[cur_grupo] = _num(c7)
            next_is_group = True
            continue
        if n6.startswith('total condominio') or n6 == 'total:':
            grand_total = _num(c7)
            continue
        # item (tem data de pagamento em col4)
        if isinstance(c4, float) and c4 > 30000:
            forn = c0.lstrip('.').strip()
            itens.append({
                'grupo': cur_grupo, 'classe': cur_classe,
                'fornecedor': forn,
                'data': _xldate(wb, c4),
                'tipo_pgto': str(sh.cell_value(r, 5)).strip(),
                'valor_lcto': _num(c7),
                'valor_pago': _num(c8) if _num(c8) else _num(c7),
                'descricao': c9 or forn,
            })
            continue
        # linha estrutural (texto em col0, sem data)
        if c0 and not c0.lower().startswith('cond.'):
            if next_is_group:
                cur_grupo = c0
                next_is_group = False
                cur_classe = None
            else:
                cur_classe = c0
        # se for a linha entidade "COND. ..." apenas ignora (mantem next_is_group)
    return {'itens': itens, 'classe_totais': classe_totais,
            'grupo_totais': grupo_totais, 'grand_total': grand_total,
            'periodo': (min((i['data'] for i in itens if i['data']), default=None),
                        max((i['data'] for i in itens if i['data']), default=None))}

# ----------------------------------------------------------------------------
# BALANUAL - demonstrativo mensal de receitas e despesas
# ----------------------------------------------------------------------------
def parse_balanual(path):
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    def rownorm(r):
        return [_norm(sh.cell_value(r, c)) for c in range(sh.ncols)]
    # localizar diretamente os cabecalhos das secoes (col2 = Receitas/Despesas + meses)
    sec_rec = sec_desp = None
    saldo_ini = saldo_fim = None
    for r in range(sh.nrows):
        c2 = _norm(sh.cell_value(r, 2))
        has_month = any(re.match(r'\d{2}/\d{4}', str(sh.cell_value(r, c)).strip()) for c in range(sh.ncols))
        if not has_month:
            continue
        if c2 == 'receitas' and sec_rec is None:
            sec_rec = r
        elif c2 == 'despesas' and sec_desp is None:
            sec_desp = r
    def parse_section(head):
        if head is None:
            return [], None, [], []
        # colunas de meses = colunas entre col4 e a col 'total'
        total_col = None
        media_col = None
        month_cols = []
        for c in range(sh.ncols):
            v = _norm(sh.cell_value(head, c))
            if v == 'total':
                total_col = c
            elif v == 'media':
                media_col = c
            elif re.match(r'\d{2}/\d{4}', str(sh.cell_value(head, c)).strip()):
                month_cols.append(c)
        mlabels = [str(sh.cell_value(head, c)).strip() for c in month_cols]
        nmcols = max(len(month_cols), 1)
        linhas = []
        cur_grp = None
        total_geral = None
        r = head + 1
        while r < sh.nrows:
            c0 = str(sh.cell_value(r, 0)).strip()
            c3 = _norm(sh.cell_value(r, 3))
            monthly = [_num(sh.cell_value(r, c)) for c in month_cols]
            tot = _num(sh.cell_value(r, total_col)) if total_col is not None else sum(monthly)
            # fim da secao
            if c3 == 'total:':
                total_geral = tot if total_col is not None else sum(monthly)
                break
            if 'demonstrativo' in _norm(c0):
                break
            has_month_vals = any(abs(x) > 0.005 for x in monthly)
            if c0 and not has_month_vals and abs(tot) < 0.005:
                cur_grp = c0  # cabecalho de grupo
            elif c0 and (has_month_vals or abs(tot) > 0.005):
                nmonths = sum(1 for x in monthly if abs(x) > 0.005)
                media = _num(sh.cell_value(r, media_col)) if media_col is not None else tot / nmcols
                linhas.append({'grupo': cur_grp, 'classe': c0,
                               'monthly': monthly, 'total': tot,
                               'media': media, 'n_meses': nmonths})
            r += 1
        return linhas, total_geral, mlabels, month_cols
    receitas, tot_rec, mlabels, _ = parse_section(sec_rec) if sec_rec is not None else ([], None, [], [])
    despesas, tot_desp, _, _ = parse_section(sec_desp) if sec_desp is not None else ([], None, [], [])
    # saldos
    for r in range(sh.nrows):
        rn = ' '.join(rownorm(r))
        if 'saldos iniciais' in rn:
            # valores na proxima linha (col4 = primeiro mes)
            for rr in range(r, min(r+3, sh.nrows)):
                vals = [sh.cell_value(rr, c) for c in range(4, sh.ncols) if isinstance(sh.cell_value(rr, c), float)]
                if vals:
                    saldo_ini = vals[0]
                    break
        if 'saldos finais' in rn:
            for rr in range(r, min(r+3, sh.nrows)):
                vals = [sh.cell_value(rr, c) for c in range(4, sh.ncols) if isinstance(sh.cell_value(rr, c), float)]
                if vals:
                    saldo_fim = vals[-1]
                    break
    if not mlabels and despesas:
        # se receitas falhou, derivar meses da secao de despesas
        mlabels = [f'M{i+1}' for i in range(len(despesas[0]['monthly']))]
    return {'receitas': receitas, 'despesas': despesas,
            'total_receitas': tot_rec, 'total_despesas': tot_desp,
            'saldo_inicial': saldo_ini, 'saldo_final': saldo_fim,
            'meses': mlabels, 'n_meses': len(mlabels)}

# ----------------------------------------------------------------------------
# DESSIN02 - sintetico de despesas (cross-check)
# ----------------------------------------------------------------------------
def parse_dessin(path):
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    grand = None
    for r in range(sh.nrows):
        rn = [_norm(sh.cell_value(r, c)) for c in range(sh.ncols)]
        if any(x.startswith('total geral') for x in rn):
            nums = [_num(sh.cell_value(r, cc)) for cc in range(sh.ncols)]
            grand = max(nums)
    if grand is None:  # fallback: maior "Total:" da planilha
        best = 0.0
        for r in range(sh.nrows):
            rn = [_norm(sh.cell_value(r, c)) for c in range(sh.ncols)]
            if any(x == 'total:' for x in rn):
                best = max(best, max(_num(sh.cell_value(r, cc)) for cc in range(sh.ncols)))
        grand = best or None
    return {'grand_total': grand}

# ----------------------------------------------------------------------------
# INAD01 - inadimplencia
# ----------------------------------------------------------------------------
def parse_inad(path):
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    data_base = None
    for r in range(sh.nrows):
        t = str(sh.cell_value(r, 0))
        m = re.search(r'até\s+(\d{2}/\d{2}/\d{4})', t) or re.search(r'ate\s+(\d{2}/\d{2}/\d{4})', _norm(t))
        if m:
            try:
                data_base = datetime.datetime.strptime(m.group(1), '%d/%m/%Y').date()
            except Exception:
                pass
        if data_base:
            break
    itens = []
    cur_unidade = None
    total = None
    # achar header
    start = 0
    for r in range(sh.nrows):
        if 'mes ref' in _norm(sh.cell_value(r, 4)) or 'vencimento' in _norm(sh.cell_value(r, 5)):
            start = r + 1
            break
    for r in range(start, sh.nrows):
        c0 = str(sh.cell_value(r, 0)).strip()
        c1 = str(sh.cell_value(r, 1)).strip()
        c4 = str(sh.cell_value(r, 4)).strip()
        c5 = sh.cell_value(r, 5)
        c6 = sh.cell_value(r, 6)
        c10 = sh.cell_value(r, 10)
        n5 = _norm(sh.cell_value(r, 5))
        # codigo da unidade isolado na col1 (ex.: "B-402")
        if c1 and not c0 and c4 == '' and not isinstance(c5, float):
            cur_unidade = c1
            continue
        # nome do devedor: "301 - NOME" ou "402 B - FERNANDO ROBERTO TORRE"
        m = re.match(r'^(.+?)\s*-\s*(.+)$', c0)
        if m and c4 == '' and not isinstance(c5, float) and not _norm(c0).startswith('total'):
            # anexa o nome ao codigo ja detectado, sem duplicar (ex.: "203 - NOME" ja contem "203")
            if cur_unidade and '—' not in cur_unidade and not c0.startswith(cur_unidade):
                cur_unidade = f'{cur_unidade} — {c0}'
            else:
                cur_unidade = c0
            continue
        if n5.startswith('total') and 'total' in n5:
            if "total:" == n5 or n5 == 'total:':
                total = _num(c6)
            continue
        # item: tem mes ref (MM/AAAA) e vencimento serial
        if re.match(r'\d{2}/\d{4}', c4) and isinstance(c5, float) and c5 > 30000:
            venc = _xldate(wb, c5)
            valor = _num(c6)
            proj = _num(c10)
            meses_atraso = None
            if data_base and venc:
                meses_atraso = (data_base.year - venc.year) * 12 + (data_base.month - venc.month)
            itens.append({'unidade': cur_unidade, 'classe': c0.replace('³', '').strip(),
                          'mes_ref': c4, 'vencimento': venc, 'valor': valor,
                          'proj_rec': proj, 'meses_atraso': meses_atraso})
    if total is None:
        total = sum(i['valor'] for i in itens)
    return {'data_base': data_base, 'itens': itens, 'total': total}

# ----------------------------------------------------------------------------
# PREVISAO xlsx (manual)
# ----------------------------------------------------------------------------
def parse_previsao(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if 'contas' in _norm(name):
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    contas = []
    confronto = {}
    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        c5 = ws.cell(r, 5).value
        n3 = _norm(c3)
        if n3 in ('valor transportado', 'desconsiderações', 'desconsideracoes', 'subtotal inicial', 'subtotal atual'):
            confronto[n3] = _num(c4)
        code = str(c2).strip() if c2 is not None else ''
        if re.match(r'^\d{2}\.\d{2}$', code):
            contas.append({'codigo': code, 'nome': str(c3).strip() if c3 else '',
                           'base': _num(c4), 'ajuste': (None if c5 is None else _num(c5))})
    return {'contas': contas, 'confronto': confronto}

# ----------------------------------------------------------------------------
# CLASSIFICACAO  (conservadora)
#   EXTRAORDINARIA -> removida da base recorrente (so com forte indicio pontual)
#   REVISAR        -> mantida na base, mas sinalizada p/ decisao humana
#   RECORRENTE     -> mantida na base
# ----------------------------------------------------------------------------
EXTRA_CLASSES = [
    # Classes SEMPRE extraordinárias (evento pontual inequívoco)
    # 'reparo no elevador' e 'conserto de bomba' NÃO são mais extra automáticas —
    #   os manuais mostram que são mantidos na base (E=0, I=D).
]
REVISAR_CLASSES = [
    'reparo no elevador', 'conserto de bomba',  # manutenções corretivas: IA/humano decide
    'manutencao pintura', 'pequenas reformas', 'manutencao eletrica',
    'manutencao hidraulica', 'manutencao de cameras', 'manutencao interfone',
    'manutencao antena', 'manutencao de porta', 'manutencao portao',
    'manutencao de sauna', 'manutencao piscina', 'dedetizac', 'descupiniz',
    'limpeza de fossa', 'caixa de gordura', 'limpeza caixa', 'caixa d',
    'extintor', 'mangueira', 'honorarios advocat', 'custas judiciais',
    'despesas cartorio', 'cartorial', 'registro da convencao', 'habite',
    'sistema de combate a incendio', 'laudo', 'autovistoria',
    'rescisao trabalhista', 'indenizacao trabalhista', 'indenizacao judicial',
    'acordo judicial', 'taxa sobre realizacao', 'poco semi', 'confeccao',
    'pensao aliment',  # julgamento humano: deduzir (pontual) ou manter (folha)
]
RECORRENTE_HINTS = [
    'salario', 'inss', 'fgts', 'pis', 'cofins', 'csll', 'ferias',
    '13', 'vale transporte', 'vale aliment', 'cesta basica', 'pensao',
    'plano de assist', 'consult', 'medicina', 'seg. do trabalho',
    'contribuic', 'uniforme', 'adiantamento', 'certificado', 'recolhimento',
    'piso salarial', 'homologac', 'agua do', 'luz do', 'gas do', 'energia',
    'telefone', 'contrato', 'seguro', 'tarifa', 'iof', 'cpmf', 'juros',
    'material de limpeza', 'material de expediente', 'correio', 'xerox',
    'impressao boleto', 'taxa de administrac', 'pro-labore', 'pro labore',
    'reembolso', 'aplicacao financeira', 'manutencao de gas', 'manutencao de jardim',
    'servico de faxina',
]
GENERIC_CLASSES = ['outras despesas', 'outros materiais', 'outros', 'estorno']
# Baldes genéricos no grupo Despesas Diversas que NÃO devem virar provisão de
# Laudo (R4): são heterogêneos e a decisão de manter/excluir é humana (revisão).
DIVERSAS_GENERICAS = ['outras despesa', 'outros', 'estorno', 'diversas']
# Termos inequivocamente de CAPITAL/obra (nunca aparecem em compras rotineiras de consumo)
CAPITAL_KW = ['reforma', 'benfeitoria', 'laudo', 'projeto',
              'reconstruc', 'autovistoria']
# Valor a partir do qual um item dentro de classe generica recorrente e' sinalizado p/ revisao
BIG_ITEM = 2000.0

def _has(text, keys):
    for k in keys:
        if k in text:
            return k
    return None

def classify(grupo, classe, descricao='', n_meses_classe=None, valor=0.0):
    """Retorna (categoria, motivo). Conservador: so remove (Extraordinaria) com forte indicio pontual."""
    g = _norm(grupo); c = _norm(classe); d = _norm(descricao)
    # 1) Grupo de Obras/Benfeitorias -> sempre extraordinaria
    if 'obras' in g or 'benfeitoria' in g:
        return ('Extraordinaria', 'Grupo Obras/Benfeitorias — evento pontual de capital')
    # 2) classes explicitamente extraordinarias (reparo/conserto corretivo pontual)
    k = _has(c, EXTRA_CLASSES)
    if k:
        return ('Extraordinaria', f'Classe "{classe}" — reparo/conserto pontual (corretiva)')
    # 3) classes para revisar (periodicas/ambiguas)
    k = _has(c, REVISAR_CLASSES)
    if k:
        return ('Revisar', f'Classe "{classe}" — despesa periodica/ambigua, mantida na base p/ revisao')
    # 4) classes genericas -> frequencia mensal decide; capital na descricao escala
    if _has(c, GENERIC_CLASSES):
        ke = _has(d, CAPITAL_KW)
        if ke:
            return ('Extraordinaria', f'Conta generica com descricao de capital ("{ke}")')
        recorrente = (n_meses_classe is not None and n_meses_classe >= 6)
        if recorrente:
            if valor and valor >= BIG_ITEM:
                return ('Revisar', f'Item de valor alto (R$ {valor:,.2f}) em conta generica recorrente — revisar')
            return ('Recorrente', f'Conta generica recorrente ({n_meses_classe}/12 meses) — consumo operacional')
        return ('Revisar', f'Conta generica pouco frequente ({n_meses_classe if n_meses_classe is not None else "?"}/12 meses)')
    # 5) recorrentes conhecidas
    k = _has(c, RECORRENTE_HINTS)
    if k:
        return ('Recorrente', f'Despesa recorrente de funcionamento ("{k}")')
    # 6) descricao com forte indicio de capital -> revisar (nao remove automaticamente)
    ke = _has(d, CAPITAL_KW)
    if ke:
        return ('Revisar', f'Descricao com possivel obra/capital ("{ke}") — revisar')
    # 7) default conservador por frequencia
    if n_meses_classe is not None and n_meses_classe >= 6:
        return ('Recorrente', f'Classe sem regra; recorrente por frequencia ({n_meses_classe}/12)')
    return ('Revisar', 'Classe sem regra explicita — revisar')


# ===================== GERADOR =====================


# ---------------------------------------------------------------------------
# Fatores medianos de deducao aprendidos do historico (fallback da R3,
# quando o desbai nao permite identificar a parcela extraordinaria).
# fator = parcela DEDUZIDA / base  (mediana dos ajustes manuais 2022-2026)
FATOR_DEDUCAO_HIST = {
    'manutencao eletrica':            0.53,
    'manutencao hidraulica':          0.56,
    'pequenas reformas':              0.42,   # mediana mantida ~58%
    'manutencao portao':              0.46,
    'manutencao pintura':             0.48,
    'manutencao de cameras':          0.25,
    'manutencao antena':              0.21,
    'manutencao interfone':           0.05,
    'manutencao de maquinas':         0.44,
    'manutencao de jardim':           0.17,
    'dedetizac':                      0.33,
    'limpeza de fossa':               0.50,
}
LUMPY_KEYS = list(FATOR_DEDUCAO_HIST.keys())

PESSOAL_PONTUAL = ['rescisao', 'indenizacao trabalhista', 'pensao aliment']
ANUALIZAR = ['contrato', 'pro-labore', 'pro labore', 'taxa de administrac',
             '13. taxa de administrac', '13o taxa']
INFLACAO = float(os.environ.get('PREVISAO_INFLACAO_PCT', '0.0472'))

THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H_FILL = PatternFill('solid', fgColor='1F3864')
H_FONT = Font(bold=True, color='FFFFFF')
SUB_FILL = PatternFill('solid', fgColor='D6E4F0')
WARN_FILL = PatternFill('solid', fgColor='FFF2CC')
EXTRA_FILL = PatternFill('solid', fgColor='FCE4EC')
OK_FILL = PatternFill('solid', fgColor='E2EFDA')
MONEY = '#,##0.00'


def detectar_outliers_estatisticos(des, lumpy_keys):
    """Camada 1 da R3: deteccao estatistica de outliers dentro de cada classe lumpy.

    Usa o metodo MAD (Median Absolute Deviation) que e robusto a distribuicoes
    assimetricas — comum em classes de manutencao onde ha muitos itens pequenos
    e alguns grandes.

    Para cada classe, NFs com valor > mediana + 3*MAD*1.4826 sao consideradas
    outliers (provavel extraordinario).

    Retorna: {(grupo_norm, classe_norm): valor_total_outliers}
    Requer minimo de 3 NFs na classe para ter significancia estatistica.
    """
    import statistics

    # Agrupa NFs por (grupo, classe) para classes lumpy E genericas de Diversas
    GENERIC_DIVERSAS = ['outras despesa', 'outros', 'estorno', 'diversas']
    nfs_por_classe = defaultdict(list)
    for it in des['itens']:
        ng = _norm(it['grupo'] or '')
        nc = _norm(it['classe'] or '')
        if any(k in nc for k in lumpy_keys) or ('diversas' in ng and any(k in nc for k in GENERIC_DIVERSAS)):
            nfs_por_classe[(ng, nc)].append(it['valor_pago'])

    outliers = {}
    for (ng, nc), valores in nfs_por_classe.items():
        n = len(valores)
        if n < 3:
            continue  # poucos dados = sem significancia estatistica
        try:
            mediana = statistics.median(valores)
        except Exception:
            continue
        if mediana < 0.01:
            continue  # valores zerados
        # MAD = median absolute deviation
        desvios = [abs(v - mediana) for v in valores]
        try:
            mad = statistics.median(desvios)
        except Exception:
            continue
        if mad < 0.01:
            # MAD zero = valores identicos — usa media + 2*desvio padrao como fallback
            try:
                media = statistics.mean(valores)
                std = statistics.stdev(valores)
            except Exception:
                continue
            if std < 0.01:
                continue
            threshold = media + 2.0 * std
        else:
            # 1.4826 = fator de consistencia para distribuicao normal
            threshold = mediana + 3.0 * mad * 1.4826

        total_outliers = sum(v for v in valores if v > threshold)
        if total_outliers > 0:
            outliers[(ng, nc)] = round(total_outliers, 2)

    return outliers


IA_SISTEMA_LUMPY = """\
Voce e um analista financeiro de condominios no Brasil especializado em previsao orcamentaria.

TAREFA: Para cada classe de manutencao abaixo, analise o padrao de gastos e determine qual
percentual do total deve ser considerado EXTRAORDINARIO (nao recorrente) para o orcamento.

CONTEXTO — O QUE E EXTRAORDINARIO:
- Reparos de grande porte: pintura externa completa, reforma de fachada, troca de sistema inteiro
- Substituicao de equipamentos: bomba nova, motor de portao novo, central de cameras nova
- Obras pontuais: impermeabilizacao, reconstrucao de calcada, troca de encanamento
- Manutencoes corretivas de alto valor, claramente acima do padrao mensal da classe

O QUE E RECORRENTE (manter na base):
- Pequenos reparos e retoques mensais/bimestrais/trimestrais
- Material de consumo para manutencao preventiva (tinta, rolo, lixa, conectores, lampadas)
- Visitas tecnicas periodicas de rotina
- Substituicao de pecas de desgaste normal (rolamentos, vedacoes, filtros)

OUTLIERS ESTATISTICOS ja foram detectados e serao deduzidos automaticamente.
Sua funcao e encontrar itens extraordinarios que o metodo estatistico NAO pegou
(ex.: valores medios mas descricao indica evento pontual).

IMPORTANTE: Seja conservador. Na duvida, mantenha como recorrente.
Itens de baixo valor ou claramente rotineiros devem ficar na base.

Responda APENAS com JSON: {"classes": {"nome da classe": {"pct": 0.XX, "justificativa": "<=25 palavras"}, ...}}
"""


def ia_analisar_classes_lumpy(itens, nome_condo, outliers_por_classe):
    """Camada 2 da R3: IA analisa cada classe lumpy e recomenda % de deducao.

    Diferente de ia_classificar_revisar() que classifica NF por NF,
    esta funcao analisa a CLASSE INTEIRA e sugere um percentual de deducao,
    exatamente como o especialista humano faz.

    Retorna: {(grupo_norm, classe_norm): pct_deducao}
    """
    if not _ia_disponivel():
        return {}

    # Agrupa NFs por classe lumpy
    classes = defaultdict(list)
    for it in itens:
        ng = _norm(it['grupo'] or '')
        nc = _norm(it['classe'] or '')
        if any(k in nc for k in LUMPY_KEYS):
            classes[(ng, nc)].append(it)

    if not classes:
        return {}

    logger.info('Enviando %d classes para IA...', len(classes))
    logger.debug('Classes para IA:')
    for (ng, nc), nfs in classes.items():
        total_classe = sum(it['valor_pago'] for it in nfs)
        logger.debug('  %s: %d NFs, total=R$ %.2f', nc, len(nfs), total_classe)

    # Monta o prompt com todas as classes
    blocos = []
    for (ng, nc), nfs in classes.items():
        total_classe = sum(it['valor_pago'] for it in nfs)
        n_nfs = len(nfs)
        extra_est = outliers_por_classe.get((ng, nc), 0.0)
        # Lista as NFs (ordenadas por valor decrescente, top 15)
        nfs_ord = sorted(nfs, key=lambda x: x['valor_pago'], reverse=True)[:15]
        nfs_txt = '\n'.join(
            f"    R${it['valor_pago']:,.2f} | {it['data']} | "
            f"{(it['descricao'] or it['fornecedor'] or '')[:100]}"
            for it in nfs_ord)
        blocos.append(
            f"CLASSE: {nfs[0]['classe']} (grupo: {nfs[0]['grupo']})\n"
            f"  Total 12 meses: R${total_classe:,.2f} | NFs: {n_nfs}\n"
            f"  Outliers estatisticos ja detectados: R${extra_est:,.2f}\n"
            f"  Lancamentos (maiores valores):\n{nfs_txt}"
        )

    prompt = (f'Condominio: {nome_condo}\n\n'
              f'Analise cada classe de manutencao e determine o % extraordinario:\n\n'
              + '\n\n'.join(blocos))

    resp = _claude_chat(IA_SISTEMA_LUMPY, prompt, max_tokens=4000)
    if not resp:
        return {}

    try:
        data = _extrai_json(resp)
    except Exception:
        logger.warning('Falha ao extrair JSON da resposta IA para classes lumpy')
        return {}

    logger.info('IA retornou %% para %d classes', len(data.get('classes', {})))
    resultado = {}
    for nome_classe, info in data.get('classes', {}).items():
        try:
            pct = float(info.get('pct', 0))
        except (ValueError, TypeError):
            pct = 0.0
        pct = max(0.0, min(1.0, pct))  # clamp 0..1
        # Busca o par (ng, nc) que casa com o nome da classe
        nc_norm = _norm(nome_classe)
        for (ng, nc) in classes:
            if nc_norm in nc or nc in nc_norm:
                resultado[(ng, nc)] = pct
                break

    return resultado


def _ws_header(ws, row, headers, widths=None):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.fill = H_FILL; c.font = H_FONT; c.border = BORDER
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w


def analisar(folder, progress_callback=None):
    """Executa parsing + classificacao + regras; retorna dict com tudo.
    Usa IA (Claude API) como parser principal; fallback para parsers rigidos."""
    t0 = time.time()
    nome = os.path.basename(folder.rstrip('/'))
    logger.info('=== INICIO analisar(): %s ===', folder)

    # --- IA-powered parsing (substitui os 4 parsers rigidos) ---
    logger.info('Passo 1/6: Carregando relatorios...')
    ia_dados = None
    try:
        from ia_parser import ia_parse_pasta
        ia_dados = ia_parse_pasta(folder)
    except Exception as e:
        logger.warning('ia_parser indisponivel (%s) — usando parsers rigidos', e)

    if ia_dados:
        bal = ia_dados['bal']
        des = ia_dados['des']
        sin = ia_dados['sin']
        ina = ia_dados['inad'] if ia_dados['inad'] and ia_dados['inad'].get('itens') else None
    else:
        # --- Fallback: parsers rigidos originais ---
        bal = parse_balanual(os.path.join(folder, 'balanual.xls'))
        des = parse_desbai(os.path.join(folder, 'desbai06.xls'))
        sin = parse_dessin(os.path.join(folder, 'dessin02.xls'))
        inad_path = os.path.join(folder, 'inad01.xls')
        ina = parse_inad(inad_path) if os.path.exists(inad_path) else None

    logger.info('balanual: %d receitas, %d despesas, total=R$ %.2f',
                len(bal['receitas']), len(bal['despesas']), bal.get('total_despesas', 0))
    logger.info('desbai06: %d itens, periodo=%s a %s',
                len(des['itens']),
                des['periodo'][0] if des['periodo'][0] else '?',
                des['periodo'][1] if des['periodo'][1] else '?')
    logger.info('dessin02: total=R$ %.2f', sin['grand_total'] if sin and sin['grand_total'] else 0)
    if ina and ina['itens']:
        logger.info('inad01: %d unidades inadimplentes, total=R$ %.2f',
                    len(set(i.get('unidade', '?') for i in ina['itens'])), ina['total'])

    # --- Cross-check entre relatorios ---
    tot_bal = bal.get('total_despesas', 0)
    tot_des = des.get('grand_total', 0)
    tot_sin = sin.get('grand_total', 0) if sin else 0
    divergencias = []
    if tot_bal > 0 and tot_des > 0:
        pct = abs(tot_bal - tot_des) / max(tot_bal, tot_des)
        if pct > 0.05:
            divergencias.append(f'balanual (R${tot_bal:,.2f}) vs desbai06 (R${tot_des:,.2f}) = {pct:.1%}')
    if tot_bal > 0 and tot_sin > 0:
        pct = abs(tot_bal - tot_sin) / max(tot_bal, tot_sin)
        if pct > 0.05:
            divergencias.append(f'balanual (R${tot_bal:,.2f}) vs dessin02 (R${tot_sin:,.2f}) = {pct:.1%}')
    if divergencias:
        logger.warning('DIVERGENCIA ENTRE RELATORIOS (>5%%):')
        for d in divergencias:
            logger.warning('  %s', d)

    prev_paths = glob.glob(os.path.join(folder, 'Previs*.xlsx'))
    manual = parse_previsao(prev_paths[0]) if prev_paths else None

    freq = {_norm(l['classe']): l['n_meses'] for l in bal['despesas']}
    media_ult = {}   # media dos ultimos 3 meses nao-zero por classe (p/ R6)
    for l in bal['despesas']:
        nz = [v for v in l['monthly'] if abs(v) > 0.005]
        if nz:
            ultimos = nz[-3:]  # ate os ultimos 3 meses
            media_ult[_norm(l['classe'])] = sum(ultimos) / len(ultimos)
        else:
            media_ult[_norm(l['classe'])] = 0.0

    if progress_callback:
        progress_callback({'fase': 'Relatórios carregados', 'passo': 1, 'total': 6, 'detalhe': f'{len(des["itens"])} itens encontrados'})

    # --- classificar lancamentos do desbai ---
    logger.info('Passo 2/6: Classificando %d itens do desbai06...', len(des['itens']))
    for it in des['itens']:
        nm = freq.get(_norm(it['classe']))
        cat, mot = classify(it['grupo'], it['classe'], it['descricao'], nm, it['valor_pago'])
        it['cat'], it['motivo'], it['n_meses'] = cat, mot, nm
    # Contagens
    n_extra = sum(1 for it in des['itens'] if it['cat'] == 'Extraordinaria')
    n_rec = sum(1 for it in des['itens'] if it['cat'] == 'Recorrente')
    n_rev = sum(1 for it in des['itens'] if it['cat'] == 'Revisar')
    logger.info('  Extraordinaria: %d | Recorrente: %d | Revisar: %d', n_extra, n_rec, n_rev)

    if progress_callback:
        progress_callback({'fase': 'Classificação inicial concluída', 'passo': 2, 'total': 6, 'detalhe': f'{n_extra} extraordinários, {n_rev} em revisão'})

    # --- IA: reclassifica itens "Revisar" ANTES de montar extra_por_classe ---
    # As sugestoes da IA alimentam o calculo do R3; nao sao so para exibicao.
    # O manual (quando presente) e passado como referencia de calibracao.
    logger.info('Passo 3/6: IA — reclassificando itens Revisar...')
    sugestoes_ia = ia_classificar_revisar(des['itens'], nome, manual=manual)
    logger.info('  IA sugeriu reclassificar %d itens', len(sugestoes_ia))
    for idx, (sug, just) in sugestoes_ia.items():
        it = des['itens'][idx]
        if sug in ('Extraordinaria', 'Recorrente'):
            it['cat'] = sug
            it['motivo'] = f'IA: {just}'

    if progress_callback:
        progress_callback({'fase': 'IA analisando itens ambíguos', 'passo': 3, 'total': 6, 'detalhe': f'{len(sugestoes_ia)} sugestões recebidas'})

    # --- R3 Camada 1: deteccao estatistica de outliers ---
    logger.info('Passo 4/6: R3 Camada 1 — deteccao estatistica (MAD)...')
    outliers_estatisticos = detectar_outliers_estatisticos(des, LUMPY_KEYS)
    if outliers_estatisticos:
        logger.info('  %d classes com outliers', len(outliers_estatisticos))
        for (ng, nc), total in sorted(outliers_estatisticos.items()):
            logger.debug('  %s: R$ %.2f em outliers', nc, total)
        # Marca apenas as NFs que sao outliers individuais (para o frontend)
        # Recalcula thresholds por classe para identificar quais NFs especificas
        import statistics as _st
        nfs_por_classe = defaultdict(list)
        for it in des['itens']:
            ng = _norm(it['grupo'] or ''); nc = _norm(it['classe'] or '')
            if (ng, nc) in outliers_estatisticos:
                nfs_por_classe[(ng, nc)].append(it)
            # Tambem marca outliers em classes genericas de Diversas
            elif 'diversas' in ng and any(k in nc for k in ('outras despesa', 'outros', 'estorno', 'diversas')):
                nfs_por_classe[(ng, nc)].append(it)
        for (ng, nc), nfs in nfs_por_classe.items():
            valores = [it['valor_pago'] for it in nfs]
            mediana = _st.median(valores)
            mad = _st.median([abs(v - mediana) for v in valores])
            if mad < 0.01:
                continue
            threshold = mediana + 3.0 * mad * 1.4826
            for it in nfs:
                if it['valor_pago'] > threshold and it['cat'] != 'Extraordinaria':
                    it['cat'] = 'Extraordinaria'
                    it['motivo'] = (f'Outlier estatistico (MAD): R${it["valor_pago"]:,.2f} > '
                                    f'mediana + 3*MAD (R${threshold:,.2f}) na classe {it["classe"]}')

    if progress_callback:
        progress_callback({'fase': 'Detectando outliers estatísticos', 'passo': 4, 'total': 6, 'detalhe': f'{len(outliers_estatisticos)} classes com outliers'})

    # --- R3 Camada 2: IA analisa cada classe lumpy e sugere % de deducao ---
    logger.info('Passo 5/6: R3 Camada 2 — IA por classe...')
    pct_ia_por_classe = ia_analisar_classes_lumpy(des['itens'], nome, outliers_estatisticos)
    n_classes_ia = len(pct_ia_por_classe)
    n_com_deducao = sum(1 for v in pct_ia_por_classe.values() if v > 0)
    logger.info('  IA analisou %d classes, %d com deducao recomendada', n_classes_ia, n_com_deducao)

    if progress_callback:
        progress_callback({'fase': 'IA analisando classes de manutenção', 'passo': 5, 'total': 6, 'detalhe': f'{len(pct_ia_por_classe)} classes avaliadas'})

    # --- Acumula NFs classificadas como Extraordinaria (NF por NF, método tradicional) ---
    extra_por_classe = defaultdict(float)
    for it in des['itens']:
        if it['cat'] == 'Extraordinaria':
            extra_por_classe[(_norm(it['grupo'] or ''), _norm(it['classe'] or ''))] += it['valor_pago']

    # --- montar plano de contas com deducoes (a partir do balanual) ---
    logger.info('Passo 6/6: Aplicando regras R1-R8...')
    linhas = []          # uma por classe do balanual
    desconsider = 0.0    # R1 (obras)
    prov_laudo = 0.0     # R4
    prov_incendio = 0.0  # R5
    for l in bal['despesas']:
        g, c = l['grupo'] or '', l['classe']
        ng, nc = _norm(g), _norm(c)
        base = l['total']
        ded = 0.0
        regra = ''
        final = base
        # R1 obras
        if 'obras' in ng or 'benfeitoria' in ng:
            desconsider += base
            ded = base; final = 0.0
            regra = 'R1: grupo Obras/Benfeitorias desconsiderado'
        # R4 diversas (exceto seguro e manutenções)
        elif 'diversas' in ng and 'seguro' not in nc:
            # Itens de manutenção/reparo no grupo Diversas NÃO são "diversas" —
            # estão mal classificados no grupo. Manter na base (como os manuais fazem).
            if any(k in nc for k in ('reparo', 'conserto', 'manutencao', 'bomba', 'portao', 'elevador')):
                regra = 'Recorrente: manutencao em grupo Diversas — mantida integral'
            elif any(k in nc for k in ('outras despesa', 'outros', 'estorno')):
                # Classes genéricas (ex.: "Outras Despesas"): aplicar MAD + IA
                # para classificar o que é extraordinário vs recorrente.
                # A maioria são equipamentos/compras pontuais → extraordinário.
                # Itens de consumo que caíram na conta errada → recorrente.
                extra_est = outliers_estatisticos.get((ng, nc), 0.0)
                ex_ia = extra_por_classe.get((ng, nc), 0.0)
                ex = max(extra_est, ex_ia)
                if ex > 0:
                    ded = min(ex, base)
                    prov_laudo += ded
                    regra = f'R4: {ded/base:.0%} extraordinario (MAD+IA) — provisionado Laudo'
                else:
                    ded = 0.0
                    regra = 'R4: sem itens extraordinarios — mantido integral'
                final = base - ded
            else:
                ded = base; final = 0.0
                prov_laudo += base
                regra = 'R4: Diversas zeradas -> provisao Laudo Autovistoria'
        # R5 cartoriais
        elif 'cartoriais' in ng or 'honorarios' in ng:
            ded = base; final = 0.0
            prov_incendio += base
            regra = 'R5: Cartoriais -> provisao Sist. Incendio/Registro'
        # R2 pessoal pontual — pensao alimenticia CONTINUA (6+ meses) e
        # desconto em folha repassado, nao custo extra: fica na base
        # (aprendido do manual 2023: Quezia manteve pensao de 11 meses)
        elif any(k in nc for k in PESSOAL_PONTUAL):
            if 'pensao' in nc and (l['n_meses'] or 0) >= 6:
                regra = 'Recorrente: pensao alimenticia continua (desconto em folha)'
            else:
                ded = base; final = 0.0
                regra = 'R2: evento pontual de pessoal deduzido 100%'
        # R6 anualizacao
        elif any(k in nc for k in ANUALIZAR) and '13' not in nc:
            final = round(media_ult.get(nc, base / 12.0) * 12, 2)
            ded = base - final
            regra = 'R6: media ultimos 3 meses x 12'
        # R3 lumpy — 2 camadas: outliers estatisticos + decisoes humanas
        elif any(k in nc for k in LUMPY_KEYS):
            # Camada 1: outliers estatisticos (NFs > media + 2σ na propria classe)
            extra_est = outliers_estatisticos.get((ng, nc), 0.0)

            # NF por NF: NFs marcadas como Extraordinaria pelo MAD + IA NF-by-NF
            extra_humano = extra_por_classe.get((ng, nc), 0.0)

            # Camada 2: IA recomenda % de deducao para a classe inteira
            pct_ia = pct_ia_por_classe.get((ng, nc), 0.0)
            extra_ia = round(base * pct_ia, 2) if pct_ia > 0 else 0.0

            # Combina: NF-por-NF tem prioridade; fallback para max(estatistica, IA)
            ex = extra_humano if extra_humano > 0 else max(extra_est, extra_ia)

            if ex > 0:
                ded = min(ex, base)
                if extra_humano > 0:
                    regra = f'R3: decisoes humanas R${extra_humano:,.2f}'
                else:
                    partes = []
                    if extra_est > 0:
                        partes.append(f'outliers R${extra_est:,.2f}')
                    if extra_ia > 0:
                        partes.append(f'IA {pct_ia:.0%}')
                    regra = 'R3: ' + ' + '.join(partes) if partes else 'R3: mantido integral'
            else:
                ded = 0.0
                regra = 'R3: sem itens extraordinarios — mantido integral'
            final = base - ded
        else:
            regra = 'Recorrente: mantida integral'
        linhas.append({'grupo': g, 'classe': c, 'base': base, 'deducao': ded,
                       'final': final, 'regra': regra, 'n_meses': l['n_meses'],
                       'monthly': l['monthly']})

    base_total = sum(l['base'] for l in linhas)
    subtotal = sum(l['final'] for l in linhas) + prov_laudo + prov_incendio
    total_previsto = subtotal * (1 + INFLACAO)

    # Log R1-R8 summary
    r1 = [(l['classe'], l['deducao']) for l in linhas if l['regra'].startswith('R1')]
    r2 = [(l['classe'], l['deducao']) for l in linhas if l['regra'].startswith('R2')]
    r3 = [(l['classe'], l['deducao']) for l in linhas if l['regra'].startswith('R3')]
    r4 = [(l['classe'], l['deducao']) for l in linhas if l['regra'].startswith('R4')]
    r5 = [(l['classe'], l['deducao']) for l in linhas if l['regra'].startswith('R5')]
    r6 = [(l['classe'], l['deducao']) for l in linhas if l['regra'].startswith('R6')]
    if r1:
        logger.info('  R1 (Obras): %d classes, R$ %.2f desconsiderados', len(r1), sum(v for _, v in r1))
    if r2:
        logger.info('  R2 (Pessoal): %d classes, R$ %.2f deduzidos', len(r2), sum(v for _, v in r2))
    if r3:
        logger.info('  R3 (Manutencao): %d classes, R$ %.2f deduzidos', len(r3), sum(v for _, v in r3))
    if r4:
        logger.info('  R4 (Diversas/Laudo): R$ %.2f provisionados', prov_laudo)
    if r5:
        logger.info('  R5 (Cartoriais/SCIP): R$ %.2f provisionados', prov_incendio)
    if r6:
        logger.info('  R6 (Anualizacao): %d contratos anualizados', len(r6))

    logger.info('=== RESULTADO ===')
    logger.info('  Base total:    R$ %12.2f', base_total)
    logger.info('  Desconsideracoes: R$ %10.2f', desconsider)
    logger.info('  Subtotal:      R$ %12.2f', subtotal)
    logger.info('  Inflacao (%.1f%%): R$ %10.2f', INFLACAO * 100, subtotal * INFLACAO)
    logger.info('  Total previsto: R$ %12.2f', total_previsto)
    logger.info('  Provisao Laudo:  R$ %10.2f', prov_laudo)
    logger.info('  Provisao SCIP:   R$ %10.2f', prov_incendio)
    if divergencias:
        for d in divergencias:
            logger.warning('  %s', d)
    logger.info('Tempo total: %.1fs', time.time() - t0)
    logger.info('=== FIM analisar() ===')

    # --- inadimplencia (R8) ---
    # Regra: unidade e critica se ficou >= 3 meses CONSECUTIVOS sem pagar.
    # Impacto: abate da receita a taxa mensal da unidade × meses consecutivos devidos.
    # O arquivo inad01 ja traz o total calculado.
    inad_res = None
    if ina and ina['itens']:
        data_base = ina['data_base']
        # Agrupar por unidade e extrair meses consecutivos
        unidade_meses = defaultdict(set)   # unidade -> set de meses (MM/AAAA)
        unidade_itens = defaultdict(list)  # unidade -> itens
        for i in ina['itens']:
            u = i['unidade'] or 'desconhecida'
            unidade_meses[u].add(i['mes_ref'])
            unidade_itens[u].append(i)

        # Determinar quais unidades tem >= 3 meses CONSECUTIVOS de divida
        unidades_criticas = {}
        for u, meses_str in unidade_meses.items():
            # Converter MM/AAAA para (ano, mes) ordenado
            meses_tuples = []
            for m in meses_str:
                parts = m.split('/')
                if len(parts) == 2:
                    try:
                        mm, aa = int(parts[0]), int(parts[1])
                        meses_tuples.append((aa, mm))
                    except ValueError:
                        pass
            if not meses_tuples:
                continue
            meses_tuples.sort()
            # Verificar se há >= 3 meses consecutivos
            consec = 1
            max_consec = 1
            for k in range(1, len(meses_tuples)):
                prev_aa, prev_mm = meses_tuples[k-1]
                cur_aa, cur_mm = meses_tuples[k]
                # Consecutivo: mesmo ano + mes seguinte, ou dez→jan do ano seguinte
                if (cur_aa == prev_aa and cur_mm == prev_mm + 1) or \
                   (cur_aa == prev_aa + 1 and prev_mm == 12 and cur_mm == 1):
                    consec += 1
                    max_consec = max(max_consec, consec)
                else:
                    consec = 1
            if max_consec >= 3:
                # Unidade critica: impacto = media da Tx. Condominio mensal × meses consecutivos
                tx_items = [i for i in unidade_itens[u]
                           if 'condominio' in _norm(i.get('classe', ''))]
                if tx_items:
                    tx_media = sum(i['valor'] for i in tx_items) / len(tx_items)
                else:
                    tx_media = sum(i['valor'] for i in unidade_itens[u]) / len(unidade_itens[u])
                unidades_criticas[u] = {
                    'meses_consecutivos': max_consec,
                    'tx_mensal_media': tx_media,
                    'impacto': tx_media * max_consec,  # impacto total na receita
                }

        # Totais
        todos_itens = ina['itens']
        critica_ids = set()
        for u, info in unidades_criticas.items():
            for i in unidade_itens[u]:
                critica_ids.add(id(i))

        crit = sum(i['valor'] for i in todos_itens if id(i) in critica_ids)
        rec = sum(i['valor'] for i in todos_itens if id(i) not in critica_ids)
        impacto_mensal = sum(info['impacto'] for info in unidades_criticas.values())

        # Marcar cada item como critico ou nao
        for i in todos_itens:
            i['critica'] = id(i) in critica_ids

        inad_res = {'total': ina['total'], 'critica': crit, 'recente': rec,
                    'data_base': ina['data_base'], 'itens': todos_itens,
                    'unidades_criticas': len(unidades_criticas),
                    'unidades_criticas_detalhe': unidades_criticas,
                    'impacto_mensal_receita': impacto_mensal}

    if progress_callback:
        progress_callback({'fase': 'Cálculos finalizados', 'passo': 6, 'total': 6, 'detalhe': f'Total previsto: R$ {total_previsto:,.2f}'})

    return {'bal': bal, 'des': des, 'sin': sin, 'inad': inad_res,
            'manual': manual, 'linhas': linhas, 'sugestoes_ia': sugestoes_ia,
            'desconsideracoes': desconsider, 'prov_laudo': prov_laudo,
            'prov_incendio': prov_incendio, 'base_total': base_total,
            'subtotal': subtotal, 'total_previsto': total_previsto,
            'outliers_estatisticos': outliers_estatisticos,
            'pct_ia_por_classe': pct_ia_por_classe,
            'divergencias': divergencias}


# ---------------------------------------------------------------------------
def recalcular(R):
    """Reaplica as regras R1-R8 com a classificacao ja decidida pelo humano.
    Usado pelo webapp apos o usuario aprovar/reprovar itens na interface.
    Retorna o mesmo dict R, atualizado com as novas linhas e subtotais."""
    # --- recalcular com as mesmas regras do analisar() ---
    # extra_por_classe reflete as decisoes humanas (itens aprovados como Extraordinaria)
    # keep_por_classe reflete NFs que o humano REPROVOU (decidiu manter na base)
    extra_por_classe = defaultdict(float)
    keep_por_classe = defaultdict(float)
    for it in R['des']['itens']:
        key = (_norm(it['grupo'] or ''), _norm(it['classe'] or ''))
        if it['cat'] == 'Extraordinaria':
            extra_por_classe[key] += it['valor_pago']
        elif it['cat'] == 'Recorrente':
            keep_por_classe[key] += it['valor_pago']

    # Recupera dados da R3 em 2 camadas (calculados no analisar)
    outliers_est = R.get('outliers_estatisticos', {})
    pct_ia_classe = R.get('pct_ia_por_classe', {})

    bal = R['bal']
    media_ult = {}   # media dos ultimos 3 meses nao-zero por classe (p/ R6)
    for l in bal['despesas']:
        nz = [v for v in l['monthly'] if abs(v) > 0.005]
        if nz:
            ultimos = nz[-3:]  # ate os ultimos 3 meses
            media_ult[_norm(l['classe'])] = sum(ultimos) / len(ultimos)
        else:
            media_ult[_norm(l['classe'])] = 0.0

    linhas, desconsider, prov_laudo, prov_incendio = [], 0.0, 0.0, 0.0
    for l in bal['despesas']:
        g, c = l['grupo'] or '', l['classe']
        ng, nc = _norm(g), _norm(c)
        base = l['total']
        ded, regra, final = 0.0, '', base
        if 'obras' in ng or 'benfeitoria' in ng:
            # R1: capital, excluído por padrão. Mas se o humano REPROVOU itens
            # (decidiu manter), essa parte volta para a base (e aparece numa
            # linha de Obras na PREVISÃO via gerador).
            kept = min(keep_por_classe.get((ng, nc), 0.0), base)
            final = round(kept, 2)
            ded = base - final
            desconsider += ded
            regra = ('R1: Obras — parte mantida na revisão' if kept > 0.005
                     else 'R1: Obras/Benfeitorias')
        elif 'diversas' in ng and 'seguro' not in nc:
            # Mesma lógica do analisar(): manutenção mal-classificada fica na base;
            # balde genérico ("Outras Despesas") NÃO vira provisão (revisar); o
            # resto vira provisão de Laudo (R4).
            if any(k in nc for k in ('reparo', 'conserto', 'manutencao', 'bomba', 'portao', 'elevador')):
                regra = 'Recorrente: manutencao em grupo Diversas — mantida integral'
            elif any(k in nc for k in DIVERSAS_GENERICAS):
                extra_est = outliers_est.get((ng, nc), 0.0)
                ex_ia = extra_por_classe.get((ng, nc), 0.0)
                ex = max(extra_est, ex_ia)
                if ex > 0:
                    ded = min(ex, base)
                    prov_laudo += ded
                    regra = f'R4: {ded/base:.0%} extraordinario — provisionado Laudo'
                else:
                    regra = 'R4: sem extraordinarios — mantido integral'
                final = base - ded
            else:
                ded, final = base, 0.0
                prov_laudo += base
                regra = 'R4: Diversas -> provisao Laudo'
        elif 'cartoriais' in ng or 'honorarios' in ng:
            ded, final = base, 0.0
            prov_incendio += base
            regra = 'R5: Cartoriais -> provisao Incendio/Registro'
        elif any(k in nc for k in PESSOAL_PONTUAL):
            if 'pensao' in nc and (l['n_meses'] or 0) >= 6:
                regra = 'Pessoal pontual: pensao continua — revisar'
            else:
                ded, final, regra = base, 0.0, 'R2: pessoal pontual deduzido — revisar'
        elif any(k in nc for k in ANUALIZAR) and '13' not in nc:
            final = round(media_ult.get(nc, base / 12.0) * 12, 2)
            ded, regra = base - final, 'R6: media ultimos 3 meses x 12'
        elif any(k in nc for k in LUMPY_KEYS):
            # R3 lumpy — 2 camadas: outliers estatisticos + decisoes humanas
            extra_est = outliers_est.get((ng, nc), 0.0)
            extra_humano = extra_por_classe.get((ng, nc), 0.0)
            pct_ia = pct_ia_classe.get((ng, nc), 0.0)
            extra_ia = round(base * pct_ia, 2) if pct_ia > 0 else 0.0
            # Camada humana tem prioridade; fallback para estatistica e IA
            ex = extra_humano if extra_humano > 0 else max(extra_est, extra_ia)
            if ex > 0:
                ded = min(ex, base)
                if extra_humano > 0:
                    regra = f'R3: decisoes humanas R${extra_humano:,.2f}'
                else:
                    partes = []
                    if extra_est > 0:
                        partes.append(f'outliers R${extra_est:,.2f}')
                    if extra_ia > 0:
                        partes.append(f'IA {pct_ia:.0%}')
                    regra = 'R3: ' + ' + '.join(partes) if partes else 'R3: mantido integral'
            else:
                ded = 0.0
                regra = 'R3: sem itens extraordinarios — mantido integral'
            final = base - ded
        else:
            ex = extra_por_classe.get((ng, nc), 0.0)
            if ex > 0:
                ded = min(ex, base)
                regra = 'Deducao aprovada na revisao (NFs extraordinarias)'
            final = base - ded
            if not regra:
                regra = 'Recorrente'
        linhas.append({'grupo': g, 'classe': c, 'base': base, 'deducao': ded,
                       'final': final, 'regra': regra, 'n_meses': l['n_meses'],
                       'monthly': l['monthly']})

    base_total = sum(l['base'] for l in linhas)
    subtotal = sum(l['final'] for l in linhas) + prov_laudo + prov_incendio
    R.update({'linhas': linhas, 'desconsideracoes': desconsider,
              'prov_laudo': prov_laudo, 'prov_incendio': prov_incendio,
              'base_total': base_total, 'subtotal': subtotal,
              'total_previsto': subtotal * (1 + INFLACAO)})
    return R


# ---------------------------------------------------------------------------
def gerar_xlsx(folder, out_path=None):
    nome = os.path.basename(folder.rstrip('/'))
    R = analisar(folder)
    bal, des = R['bal'], R['des']
    wb = openpyxl.Workbook()

    # ============== ABA RESUMO ==============
    ws = wb.active; ws.title = 'Resumo'
    ws['A1'] = f'PREVISAO ORCAMENTARIA AUTOMATICA — {nome}'
    ws['A1'].font = Font(bold=True, size=14)
    p = des['periodo']
    ws['A2'] = f"Base: 12 meses de {p[0]} a {p[1]}  |  gerado em {datetime.date.today()}"
    rows = [
        ('Despesa total do periodo (Valor Transportado)', R['base_total']),
        ('(-) Desconsideracoes — Obras/Benfeitorias (R1)', -R['desconsideracoes']),
        ('(-) Deducoes pontuais/extraordinarias (R2,R3,R4,R5)',
         -(sum(l['deducao'] for l in R['linhas']) - R['desconsideracoes'])),
        ('(+) Provisao Laudo de Autovistoria (R4)', R['prov_laudo']),
        ('(+) Provisao Sist. Incendio / Registro (R5)', R['prov_incendio']),
        ('(=) SUBTOTAL base recorrente ajustada', R['subtotal']),
        (f'(+) Inflacao {INFLACAO:.0%} (R7)', R['subtotal'] * INFLACAO),
        ('(=) PREVISAO ANUAL DE DESPESAS', R['total_previsto']),
        ('Media mensal original', R['base_total'] / 12),
        ('Media mensal ajustada', R['subtotal'] / 12),
        ('MEDIA MENSAL PREVISTA (c/ reajuste)', R['total_previsto'] / 12),
        ('Receita anual do periodo', bal['total_receitas'] or 0),
        ('Receita media mensal', (bal['total_receitas'] or 0) / 12),
    ]
    r = 4
    for label, val in rows:
        ws.cell(r, 1, label).border = BORDER
        c = ws.cell(r, 2, round(val, 2)); c.number_format = MONEY; c.border = BORDER
        if label.startswith('(='):
            ws.cell(r, 1).font = Font(bold=True); c.font = Font(bold=True)
            ws.cell(r, 1).fill = SUB_FILL; c.fill = SUB_FILL
        r += 1
    if R['inad']:
        r += 1
        ws.cell(r, 1, 'INADIMPLENCIA — abate da receita (nao e despesa)').font = Font(bold=True)
        r += 1
        inad = R['inad']
        for label, val, critica in [
            ('Total em aberto (todos os devedores)', inad['total'], False),
            (f'Critica (>= 3 meses) — {inad["unidades_criticas"]} unidade(s)', inad['critica'], True),
            ('Recente (< 3 meses) — nao abate receita', inad['recente'], False),
            ('(-) Impacto mensal na receita prevista (taxa das unidades criticas)',
             inad['impacto_mensal_receita'], True),
        ]:
            ws.cell(r, 1, label).border = BORDER
            c = ws.cell(r, 2, round(val, 2)); c.number_format = MONEY; c.border = BORDER
            if critica:
                ws.cell(r, 1).fill = WARN_FILL; c.fill = WARN_FILL
            r += 1
        rec_ajustada = (bal['total_receitas'] or 0) / 12 - inad['impacto_mensal_receita']
        ws.cell(r, 1, 'Receita mensal liquida (descontada inadimplencia critica)').font = Font(bold=True)
        c = ws.cell(r, 2, round(rec_ajustada, 2)); c.number_format = MONEY; c.font = Font(bold=True)
        c.fill = (WARN_FILL if rec_ajustada < R['total_previsto'] / 12 else OK_FILL)
        r += 1
        ws.cell(r, 1, f"Data-base inad01: {inad['data_base']}")
    ws.column_dimensions['A'].width = 52; ws.column_dimensions['B'].width = 16

    # ============== ABA CONTAS (auto) ==============
    ws = wb.create_sheet('CONTAS (auto)')
    _ws_header(ws, 1, ['Grupo', 'Classe', 'Base 12m (D)', 'Deducao sugerida (E)',
                       'Previsao (D-E)', 'Meses c/ gasto', 'Regra aplicada'],
               [26, 36, 14, 16, 14, 9, 52])
    r = 2
    for l in R['linhas']:
        ws.cell(r, 1, l['grupo']).border = BORDER
        ws.cell(r, 2, l['classe']).border = BORDER
        for j, v in ((3, l['base']), (4, l['deducao']), (5, l['final'])):
            c = ws.cell(r, j, round(v, 2)); c.number_format = MONEY; c.border = BORDER
        ws.cell(r, 6, l['n_meses']).border = BORDER
        ws.cell(r, 7, l['regra']).border = BORDER
        if l['deducao'] > 0.005:
            for j in range(1, 8):
                ws.cell(r, j).fill = WARN_FILL if l['final'] > 0 else EXTRA_FILL
        r += 1
    for label, val in [('(+) Provisao Laudo Autovistoria (R4)', R['prov_laudo']),
                       ('(+) Provisao Sist. Incendio/Registro (R5)', R['prov_incendio'])]:
        ws.cell(r, 2, label).font = Font(bold=True)
        c = ws.cell(r, 5, round(val, 2)); c.number_format = MONEY; c.font = Font(bold=True)
        for j in range(1, 8): ws.cell(r, j).fill = OK_FILL
        r += 1
    ws.cell(r, 2, 'SUBTOTAL').font = Font(bold=True)
    c = ws.cell(r, 5, round(R['subtotal'], 2)); c.number_format = MONEY; c.font = Font(bold=True)
    for j in range(1, 8): ws.cell(r, j).fill = SUB_FILL

    # ============== ABA PREVISAO (auto) — formato do relatorio final ==============
    ws = wb.create_sheet('PREVISAO (auto)')
    ws['A1'] = f'PREVISAO ORCAMENTARIA — {nome}'
    ws['A1'].font = Font(bold=True, size=13)
    # receitas
    rec_media = (bal['total_receitas'] or 0) / 12
    ws['A3'] = 'RECEITAS'; ws['A3'].font = Font(bold=True); ws['A3'].fill = SUB_FILL
    ws['B3'] = 'VALOR MENSAL'; ws['B3'].font = Font(bold=True); ws['B3'].fill = SUB_FILL
    r = 4
    for ln in bal['receitas']:
        ws.cell(r, 1, ln['classe'])
        c = ws.cell(r, 2, round(ln['total'] / 12, 2)); c.number_format = MONEY
        r += 1
    ws.cell(r, 1, 'TOTAL').font = Font(bold=True)
    c = ws.cell(r, 2, round(rec_media, 2)); c.number_format = MONEY; c.font = Font(bold=True)
    r += 2
    ws.cell(r, 1, 'DESPESAS').font = Font(bold=True); ws.cell(r, 1).fill = SUB_FILL
    ws.cell(r, 2, 'VALOR ANUAL').font = Font(bold=True); ws.cell(r, 2).fill = SUB_FILL
    ws.cell(r, 3, 'VALOR MENSAL').font = Font(bold=True); ws.cell(r, 3).fill = SUB_FILL
    r += 1
    # agrupar previsao final por grupo (com provisoes destacadas)
    por_grupo = defaultdict(float)
    for l in R['linhas']:
        por_grupo[l['grupo']] += l['final']
    for g, v in por_grupo.items():
        if abs(v) < 0.005: continue
        ws.cell(r, 1, g)
        c = ws.cell(r, 2, round(v, 2)); c.number_format = MONEY
        c = ws.cell(r, 3, round(v / 12, 2)); c.number_format = MONEY
        r += 1
    for label, val in [('Provisao Laudo de Autovistoria Predial', R['prov_laudo']),
                       ('Provisao Sist. Combate Incendio / Registro', R['prov_incendio'])]:
        if val > 0.005:
            ws.cell(r, 1, label).fill = OK_FILL
            c = ws.cell(r, 2, round(val, 2)); c.number_format = MONEY; c.fill = OK_FILL
            c = ws.cell(r, 3, round(val / 12, 2)); c.number_format = MONEY; c.fill = OK_FILL
            r += 1
    ws.cell(r, 1, 'SUBTOTAL').font = Font(bold=True)
    c = ws.cell(r, 2, round(R['subtotal'], 2)); c.number_format = MONEY; c.font = Font(bold=True)
    r += 1
    ws.cell(r, 1, f'PREVISAO DE INFLACAO — {INFLACAO:.0%}')
    c = ws.cell(r, 2, round(R['subtotal'] * INFLACAO, 2)); c.number_format = MONEY
    r += 1
    ws.cell(r, 1, 'TOTAL').font = Font(bold=True); ws.cell(r, 1).fill = SUB_FILL
    c = ws.cell(r, 2, round(R['total_previsto'], 2)); c.number_format = MONEY
    c.font = Font(bold=True); c.fill = SUB_FILL
    c = ws.cell(r, 3, round(R['total_previsto'] / 12, 2)); c.number_format = MONEY
    c.font = Font(bold=True); c.fill = SUB_FILL
    r += 2
    ws.cell(r, 1, 'Taxa condominial mensal necessaria (despesa prevista / 12):').font = Font(bold=True)
    c = ws.cell(r, 2, round(R['total_previsto'] / 12, 2)); c.number_format = MONEY; c.font = Font(bold=True)
    r += 1
    ws.cell(r, 1, 'Receita mensal atual (media do periodo):')
    c = ws.cell(r, 2, round(rec_media, 2)); c.number_format = MONEY
    r += 1
    dif = R['total_previsto'] / 12 - rec_media
    ws.cell(r, 1, 'Ajuste mensal necessario (+ aumenta taxa / - sobra):').font = Font(bold=True)
    c = ws.cell(r, 2, round(dif, 2)); c.number_format = MONEY; c.font = Font(bold=True)
    c.fill = WARN_FILL if dif > 0 else OK_FILL
    ws.column_dimensions['A'].width = 46
    for col in 'BC': ws.column_dimensions[col].width = 15

    # ============== ABA BASE DESBAI CLASSIFICADA ==============
    ws = wb.create_sheet('Base_Desbai_Classificada')
    _ws_header(ws, 1, ['Grupo', 'Classe', 'Data Pgto', 'Fornecedor/Historico',
                       'Valor Pago', 'Categoria', 'Motivo'],
               [24, 30, 11, 46, 12, 14, 50])
    r = 2
    for it in des['itens']:
        ws.cell(r, 1, it['grupo']).border = BORDER
        ws.cell(r, 2, it['classe']).border = BORDER
        ws.cell(r, 3, str(it['data'] or '')).border = BORDER
        ws.cell(r, 4, it['descricao'][:120]).border = BORDER
        c = ws.cell(r, 5, it['valor_pago']); c.number_format = MONEY; c.border = BORDER
        ws.cell(r, 6, it['cat']).border = BORDER
        ws.cell(r, 7, it['motivo']).border = BORDER
        if it['cat'] == 'Extraordinaria':
            for j in range(1, 8): ws.cell(r, j).fill = EXTRA_FILL
        elif it['cat'] == 'Revisar':
            for j in range(1, 8): ws.cell(r, j).fill = WARN_FILL
        r += 1

    # ============== ABAS EXTRAORDINARIAS / REVISAR (com sugestao da IA) ==============
    # sugestoes_ia ja foi calculado em analisar() e aplicado nos itens.
    # Aqui usamos para exibir a justificativa nos itens que continuam "Revisar".
    sugestoes_ia = R.get('sugestoes_ia', {})
    for aba, cat in (('Extraordinarias', 'Extraordinaria'), ('Revisar', 'Revisar')):
        ws = wb.create_sheet(aba)
        heads = ['Grupo', 'Classe', 'Data', 'Fornecedor/Historico', 'Valor', 'Motivo']
        widths = [24, 30, 11, 46, 12, 52]
        if cat == 'Revisar':
            heads += ['Sugestao IA', 'Justificativa IA']
            widths += [14, 44]
        _ws_header(ws, 1, heads, widths)
        r = 2; tot = 0.0
        for idx, it in enumerate(des['itens']):
            if it['cat'] != cat: continue
            ws.cell(r, 1, it['grupo']).border = BORDER
            ws.cell(r, 2, it['classe']).border = BORDER
            ws.cell(r, 3, str(it['data'] or '')).border = BORDER
            ws.cell(r, 4, it['descricao'][:120]).border = BORDER
            c = ws.cell(r, 5, it['valor_pago']); c.number_format = MONEY; c.border = BORDER
            ws.cell(r, 6, it['motivo']).border = BORDER
            if cat == 'Revisar':
                sug, just = sugestoes_ia.get(idx, ('', ''))
                cs = ws.cell(r, 7, sug); cs.border = BORDER
                ws.cell(r, 8, just).border = BORDER
                if sug == 'Extraordinaria':
                    cs.fill = EXTRA_FILL
                elif sug == 'Recorrente':
                    cs.fill = OK_FILL
            tot += it['valor_pago']; r += 1
        ws.cell(r, 4, 'TOTAL').font = Font(bold=True)
        c = ws.cell(r, 5, round(tot, 2)); c.number_format = MONEY; c.font = Font(bold=True)
        if cat == 'Revisar':
            r += 2
            ws.cell(r, 4, 'Itens acima necessitam revisao humana. '
                          'Os demais ja foram incorporados ao calculo pelo modulo de IA.').font = Font(italic=True)

    # ============== ABA COMPARATIVO RELATORIOS ==============
    ws = wb.create_sheet('Comparativo_Relatorios')
    _ws_header(ws, 1, ['Relatorio', 'Total Despesas', 'Diferenca vs desbai06', 'Status'],
               [30, 16, 18, 24])
    base = des['grand_total'] or 0
    for r, (nm_, v) in enumerate([('desbai06 (analitico)', base),
                                  ('balanual (demonstrativo)', bal['total_despesas'] or 0),
                                  ('dessin02 (sintetico)', R['sin']['grand_total'] or 0)], 2):
        ws.cell(r, 1, nm_).border = BORDER
        c = ws.cell(r, 2, round(v, 2)); c.number_format = MONEY; c.border = BORDER
        d = v - base
        c = ws.cell(r, 3, round(d, 2)); c.number_format = MONEY; c.border = BORDER
        ok = abs(d) < max(50, base * 0.001)
        st = ws.cell(r, 4, 'OK' if ok else 'VERIFICAR (juros/estornos?)')
        st.border = BORDER; st.fill = OK_FILL if ok else WARN_FILL

    # ============== ABA INADIMPLENCIA ==============
    ws = wb.create_sheet('Inadimplencia')
    if R['inad']:
        ws['A1'] = f"Data-base: {R['inad']['data_base']}  |  Critica = vencido ha mais de 3 meses"
        ws['A1'].font = Font(bold=True)
        _ws_header(ws, 3, ['Unidade/Devedor', 'Classe', 'Mes Ref', 'Vencimento',
                           'Valor', 'Meses de atraso', 'Criticidade'],
                   [34, 20, 10, 12, 12, 14, 14])
        r = 4
        for i in sorted(R['inad']['itens'], key=lambda x: -(x['meses_atraso'] or 0)):
            ws.cell(r, 1, i['unidade']).border = BORDER
            ws.cell(r, 2, i['classe']).border = BORDER
            ws.cell(r, 3, i['mes_ref']).border = BORDER
            ws.cell(r, 4, str(i['vencimento'] or '')).border = BORDER
            c = ws.cell(r, 5, i['valor']); c.number_format = MONEY; c.border = BORDER
            ws.cell(r, 6, i['meses_atraso']).border = BORDER
            crit = (i['meses_atraso'] or 0) >= 3
            cc = ws.cell(r, 7, 'CRITICA (>= 3 meses)' if crit else 'recente (< 3 meses)')
            cc.border = BORDER
            if crit:
                for j in range(1, 8): ws.cell(r, j).fill = WARN_FILL
            r += 1
        r += 1
        inad = R['inad']
        for label, val in [('Inadimplencia total', inad['total']),
                           ('Critica (>= 3 meses)', inad['critica']),
                           ('Recente (< 3 meses) — nao abate receita', inad['recente']),
                           (f'Impacto mensal na receita ({inad["unidades_criticas"]} unidade(s) critica(s))',
                            inad['impacto_mensal_receita'])]:
            ws.cell(r, 4, label).font = Font(bold=True)
            c = ws.cell(r, 5, round(val, 2)); c.number_format = MONEY; c.font = Font(bold=True)
            r += 1
    else:
        ws['A1'] = 'Sem relatorio inad01.xls nesta pasta ou sem inadimplencia na data da previsao.'

    # ============== ABA FLUXO DE CAIXA ==============
    ws = wb.create_sheet('Fluxo_Caixa')
    meses = bal['meses']
    _ws_header(ws, 1, ['Mes'] + meses + ['Total'], [22] + [11] * (len(meses) + 1))
    rec_m = [sum(l['monthly'][i] for l in bal['receitas']) for i in range(len(meses))]
    dsp_m = [sum(l['monthly'][i] for l in bal['despesas']) for i in range(len(meses))]
    for r, (nm_, vals) in enumerate([('Receitas', rec_m), ('Despesas', dsp_m),
                                     ('Resultado', [a - b for a, b in zip(rec_m, dsp_m)])], 2):
        ws.cell(r, 1, nm_).border = BORDER; ws.cell(r, 1).font = Font(bold=True)
        for j, v in enumerate(vals, 2):
            c = ws.cell(r, j, round(v, 2)); c.number_format = MONEY; c.border = BORDER
            if nm_ == 'Resultado' and v < 0: c.fill = WARN_FILL
        c = ws.cell(r, len(meses) + 2, round(sum(vals), 2))
        c.number_format = MONEY; c.font = Font(bold=True); c.border = BORDER
    r = 6
    ws.cell(r, 1, 'Saldo inicial do periodo:').font = Font(bold=True)
    c = ws.cell(r, 2, bal['saldo_inicial']); c.number_format = MONEY
    ws.cell(r + 1, 1, 'Saldo final do periodo:').font = Font(bold=True)
    c = ws.cell(r + 1, 2, bal['saldo_final']); c.number_format = MONEY

    # ============== ABA COMPARACAO PREVISAO MANUAL ==============
    ws = wb.create_sheet('Comparacao_Previsao_Manual')
    if R['manual']:
        cf = R['manual']['confronto']
        _ws_header(ws, 1, ['Metrica', 'Manual (Previsao.xlsx)', 'Automatico', 'Diferenca'],
                   [42, 18, 18, 14])
        man_sub = cf.get('subtotal atual') or cf.get('subtotal inicial') or 0
        pares = [('Valor Transportado (base)', cf.get('valor transportado', 0), R['base_total']),
                 ('Desconsideracoes (obras)', cf.get('desconsiderações', cf.get('desconsideracoes', 0)), R['desconsideracoes']),
                 ('Subtotal ajustado', man_sub, R['subtotal']),
                 (f'Total com inflacao {INFLACAO:.0%}', man_sub * (1 + INFLACAO), R['total_previsto'])]
        for r, (label, m, a) in enumerate(pares, 2):
            ws.cell(r, 1, label).border = BORDER
            c = ws.cell(r, 2, round(m, 2)); c.number_format = MONEY; c.border = BORDER
            c = ws.cell(r, 3, round(a, 2)); c.number_format = MONEY; c.border = BORDER
            d = a - m
            c = ws.cell(r, 4, round(d, 2)); c.number_format = MONEY; c.border = BORDER
            c.fill = OK_FILL if (abs(d) <= max(0.02 * max(abs(m), 1), 50)) else WARN_FILL
        # contas com ajuste manual vs deducao automatica
        r = len(pares) + 3
        ws.cell(r, 1, 'CONTAS COM AJUSTE MANUAL (col E do seu arquivo) vs DEDUCAO AUTOMATICA').font = Font(bold=True)
        r += 1
        _ws_header(ws, r, ['Conta', 'Base (D)', 'Ajuste manual (E)', 'Deducao automatica', 'Dif'],
                   [42, 14, 16, 17, 12])
        r += 1
        ded_auto = {_norm(l['classe']): l['deducao'] for l in R['linhas']}
        for ct in R['manual']['contas']:
            if ct['ajuste'] is None or abs(ct['ajuste']) < 0.005: continue
            da = ded_auto.get(_norm(ct['nome']), None)
            ws.cell(r, 1, f"{ct['codigo']} {ct['nome']}").border = BORDER
            c = ws.cell(r, 2, ct['base']); c.number_format = MONEY; c.border = BORDER
            c = ws.cell(r, 3, ct['ajuste']); c.number_format = MONEY; c.border = BORDER
            if da is not None:
                c = ws.cell(r, 4, round(da, 2)); c.number_format = MONEY; c.border = BORDER
                c = ws.cell(r, 5, round(da - ct['ajuste'], 2)); c.number_format = MONEY; c.border = BORDER
            else:
                ws.cell(r, 4, '—').border = BORDER; ws.cell(r, 5, '').border = BORDER
            r += 1
    else:
        ws['A1'] = 'Sem planilha Previsao manual nesta pasta para comparar.'

    # ============== ABA PARECER IA ==============
    parecer = ia_parecer(R, nome)
    if parecer:
        ws = wb.create_sheet('Parecer IA')
        ws.column_dimensions['A'].width = 110
        ws['A1'] = f'PARECER EXECUTIVO (gerado por IA — {_ia_modelo()})'
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = 'Revisao humana recomendada. Os numeros vem do calculo deterministico das abas anteriores.'
        ws['A2'].font = Font(italic=True, color='808080')
        r = 4
        import textwrap
        for par in parecer.split('\n'):
            if not par.strip():
                r += 1
                continue
            for line in textwrap.wrap(par.strip(), width=110) or ['']:
                ws.cell(r, 1, line)
                r += 1

    # ============== ABA CRITERIOS ==============
    ws = wb.create_sheet('Criterios')
    ws.column_dimensions['A'].width = 110
    doc = CRITERIOS.strip().split('\n')
    for r, line in enumerate(doc, 1):
        ws.cell(r, 1, line)
        if line.strip().startswith('R'):
            ws.cell(r, 1).font = Font(bold=True)
    rr = len(doc) + 2
    ws.cell(rr, 1, 'CAMADA DE IA: itens "Revisar" recebem sugestao automatica (aba Revisar, colunas '
                   'Sugestao IA/Justificativa) e um parecer executivo e gerado (aba Parecer IA). '
                   'A IA nao altera os numeros — decisao final e humana.').font = Font(bold=True)

    if out_path is None:
        out_path = os.path.join(folder, f'Previsao AUTO - {nome}.xlsx')
    wb.save(out_path)
    return out_path, R




# ===================== MODO INTERATIVO =====================
def _limpa_caminho(s):
    s = s.strip().strip('"').strip("'")
    # caminhos arrastados no Terminal vem com espacos escapados (\ )
    try:
        parts = shlex.split(s)
        if parts:
            s = ' '.join(parts) if len(parts) > 1 and not os.path.exists(parts[0]) else parts[0]
    except ValueError:
        pass
    return os.path.expanduser(s.rstrip('/'))


def _valida_pasta(pasta):
    if not os.path.isdir(pasta):
        return None, f"'{pasta}' nao e uma pasta."
    tem_desbai = glob.glob(os.path.join(pasta, 'desbai*.xls'))
    tem_bal = glob.glob(os.path.join(pasta, 'balanual*.xls'))
    if tem_desbai and tem_bal:
        return [pasta], None
    # talvez seja uma pasta-mae com varias subpastas de condominio
    subs = []
    for d in sorted(glob.glob(os.path.join(pasta, '*'))):
        if os.path.isdir(d) and glob.glob(os.path.join(d, 'desbai*.xls')) and glob.glob(os.path.join(d, 'balanual*.xls')):
            subs.append(d)
        # nivel 2 (Condominio/Condominio ANO)
        for d2 in sorted(glob.glob(os.path.join(d, '*'))):
            if os.path.isdir(d2) and glob.glob(os.path.join(d2, 'desbai*.xls')) and glob.glob(os.path.join(d2, 'balanual*.xls')):
                subs.append(d2)
    if subs:
        return subs, None
    return None, ("Nao encontrei balanual.xls + desbai06.xls nessa pasta "
                  "(nem em subpastas). Confira se exportou os relatorios.")


def processar(pasta):
    logger.info('📂 Processando: %s', pasta)
    try:
        out, R = gerar_xlsx(pasta)
    except Exception as e:
        logger.error('Erro: %s', e)
        return
    logger.info('✅ Gerado: %s', out)
    logger.info('Despesa 12m: R$ %.2f', R['base_total'])
    logger.info('Base recorrente ajustada: R$ %.2f', R['subtotal'])
    logger.info('Previsao anual (+%.0f%%): R$ %.2f   |   mensal: R$ %.2f',
                INFLACAO * 100, R['total_previsto'], R['total_previsto'] / 12)
    if R['inad'] and R['inad']['critica'] > 0.005:
        logger.warning('Inadimplencia critica (>=3 meses): R$ %.2f  |  impacto mensal na receita: -R$ %.2f',
                       R['inad']['critica'], R['inad']['impacto_mensal_receita'])
    if R['manual']:
        cf = R['manual']['confronto']
        man = cf.get('subtotal atual') or cf.get('subtotal inicial') or 0
        if man:
            d = R['subtotal'] - man
            logger.info('Comparacao c/ previsao manual: R$ %.2f  (dif %+.2f = %+.1f%%)',
                        man, d, 100 * d / man)


def main():
    logger.info('%s', '=' * 62)
    logger.info('  PREVISAO ORCAMENTARIA DE CONDOMINIOS  (Condominio21)')
    logger.info('%s', '=' * 62)
    if _ia_disponivel():
        logger.info('🤖 IA ativa (%s: %s) — sugestoes + parecer executivo', _ia_provedor(), _ia_modelo())
    else:
        logger.info('ℹ️  IA desligada — para ativar, crie chave_claude.txt OU chave_openai.txt')
        logger.info('     (ao lado deste programa), ou defina ANTHROPIC_API_KEY / OPENAI_API_KEY.')
    if len(sys.argv) > 1:
        entrada = ' '.join(sys.argv[1:])
    else:
        logger.info('Arraste a pasta do condominio para esta janela e de Enter')
        logger.info('(pode ser a pasta de um condominio/ano, ou uma pasta-mae com varios):')
        try:
            entrada = input('Pasta: ')
        except (EOFError, KeyboardInterrupt):
            logger.info('Cancelado.'); return
    pasta = _limpa_caminho(entrada)
    pastas, erro = _valida_pasta(pasta)
    if erro:
        logger.error(erro)
        sys.exit(1)
    if len(pastas) > 1:
        logger.info('Encontrei %d pastas de condominio:', len(pastas))
        for p in pastas:
            logger.info('   • %s', os.path.relpath(p, pasta))
        try:
            ok = input('\nProcessar todas? (s/n) ')
        except (EOFError, KeyboardInterrupt):
            return
        if ok.strip().lower() not in ('s', 'sim', 'y', ''):
            return
    for p in pastas:
        processar(p)
    logger.info('Concluido. Os arquivos "Previsao AUTO - *.xlsx" estao junto de cada pasta.')


if __name__ == '__main__':
    main()
