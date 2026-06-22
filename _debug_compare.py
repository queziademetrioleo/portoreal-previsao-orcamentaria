#!/usr/bin/env python3
import sys, os, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, '/Users/Usuario/PrevisaoOrcamentaria')
sys.path.insert(0, '/Users/Usuario/PrevisaoOrcamentaria/webapp/backend')
os.chdir('/Users/Usuario/PrevisaoOrcamentaria')
import openpyxl
from gerador_previsao import _norm

destino = '/var/folders/xm/s155brh53vv_31g1_dn4fj7w0000gn/T/tmph0qy_735.xlsx'
manual_path = '/Users/Usuario/Downloads/Quezia - Previsão Orçamentária 3/Sophia I/Sophia I 2024/Previsão 2024.xlsx'

wb_a = openpyxl.load_workbook(destino, data_only=True)
wb_m = openpyxl.load_workbook(manual_path, data_only=True)

ws_a = wb_a[' P R E V I S Ã O ']
ws_m = wb_m[' P R E V I S Ã O ']

def build_row_map(ws, section_start, section_end):
    m = {}
    for r in range(section_start, section_end + 1):
        nome = str(ws.cell(r, 3).value or '').strip()
        if nome:
            d = ws.cell(r, 4).value
            e = ws.cell(r, 5).value
            f = ws.cell(r, 6).value
            m[_norm(nome)] = {'row': r, 'nome': nome, 'D': d, 'E': e, 'F': f}
    return m

sep = '=' * 90

# Receitas
print(sep)
print('RECEITAS'.center(90))
print(sep)
print(f"{'Row':>4} {'Nome':<40} {'Auto D':>12} {'Manual D':>12} {'Auto E':>12} {'Manual E':>12}  {'Status':>6}")
print('-' * 90)

auto_rec = build_row_map(ws_a, 10, 19)
man_rec = build_row_map(ws_m, 10, 19)
for nn in sorted(set(list(auto_rec.keys()) + list(man_rec.keys()))):
    a = auto_rec.get(nn, {})
    m = man_rec.get(nn, {})
    ad = a.get('D', '')
    md = m.get('D', '')
    ae = a.get('E', '')
    me = m.get('E', '')
    status = 'OK'
    if isinstance(ad, (int,float)) and isinstance(md, (int,float)):
        if abs(ad - md) > 0.05:
            status = 'DIVERGE'
    elif ad != md:
        status = 'DIVERGE'
    row = a.get('row', m.get('row', '?'))
    nome = a.get('nome', m.get('nome', nn))
    ad_s = f"{ad:>12.2f}" if isinstance(ad, (int,float)) else f"{str(ad):>12}"
    md_s = f"{md:>12.2f}" if isinstance(md, (int,float)) else f"{str(md):>12}"
    ae_s = f"{ae:>12.2f}" if isinstance(ae, (int,float)) else f"{str(ae):>12}"
    me_s = f"{me:>12.2f}" if isinstance(me, (int,float)) else f"{str(me):>12}"
    print(f"{row:>4} {nome:<40} {ad_s} {md_s} {ae_s} {me_s}  {status:>6}")

# Despesas
print()
print(sep)
print('DESPESAS'.center(90))
print(sep)
print(f"{'Row':>4} {'Nome':<45} {'Auto D':>12} {'Manual D':>12} {'Auto E':>12} {'Manual E':>12}  {'Status':>6}")
print('-' * 90)

auto_desp = build_row_map(ws_a, 22, 47)
man_desp = build_row_map(ws_m, 22, 47)
for nn in sorted(set(list(auto_desp.keys()) + list(man_desp.keys()))):
    a = auto_desp.get(nn, {})
    m = man_desp.get(nn, {})
    ad = a.get('D', '')
    md = m.get('D', '')
    ae = a.get('E', '')
    me = m.get('E', '')
    # Only flag if both have real values that differ, or one has a real value and the other blank
    status = 'OK'
    both_real = isinstance(ad, (int,float)) and isinstance(md, (int,float))
    if both_real:
        if abs(ad - md) > 0.05 and abs(md) > 0.005:
            status = 'DIVERGE'
    elif isinstance(ad, (int,float)) and abs(ad) > 0.005:
        if md is None or md == '' or (isinstance(md, (int,float)) and abs(md) < 0.005):
            status = 'EXTRA'
    elif isinstance(md, (int,float)) and abs(md) > 0.005:
        if ad is None or ad == '' or (isinstance(ad, (int,float)) and abs(ad) < 0.005):
            status = 'MISSING'
    row = a.get('row', m.get('row', '?'))
    nome = a.get('nome', m.get('nome', nn))
    ad_s = f"{ad:>12.2f}" if isinstance(ad, (int,float)) else f"{str(ad if ad != '' else 'blank'):>12}"
    md_s = f"{md:>12.2f}" if isinstance(md, (int,float)) else f"{str(md if md != '' else 'blank'):>12}"
    ae_s = f"{ae:>12.2f}" if isinstance(ae, (int,float)) else f"{str(ae if ae != '' else 'blank'):>12}"
    me_s = f"{me:>12.2f}" if isinstance(me, (int,float)) else f"{str(me if me != '' else 'blank'):>12}"
    print(f"{row:>4} {nome:<45} {ad_s} {md_s} {ae_s} {me_s}  {status:>7}")

# Summary section
print()
print(sep)
print('SUMARIO'.center(90))
print(sep)
pairs = [
    ('SUBTOTAL', 48, 4, 47, 4),
    ('SUBTOTAL(6)', 48, 6, 47, 6),
    ('INFLACAO', 49, 4, 48, 4),
    ('TOTAL', 51, 4, 50, 4),
    ('SALDO', 53, 4, 52, 4),
]
for label, a_row, a_col, m_row, m_col in pairs:
    av = ws_a.cell(a_row, a_col).value
    mv = ws_m.cell(m_row, m_col).value
    a_label = str(ws_a.cell(a_row, 3).value or '').strip()
    m_label = str(ws_m.cell(m_row, 3).value or '').strip()
    status = 'OK'
    if isinstance(av, (int,float)) and isinstance(mv, (int,float)):
        if abs(av - mv) > 0.05:
            status = 'DIVERGE'
    elif av != mv:
        status = 'DIVERGE'
    av_s = f"{av:>12.2f}" if isinstance(av, (int,float)) else f"{str(av):>12}"
    mv_s = f"{mv:>12.2f}" if isinstance(mv, (int,float)) else f"{str(mv):>12}"
    print(f"  {label:<15} Auto col {a_col}={av_s} (row {a_row}, \"{a_label}\")  Manual={mv_s} (row {m_row}, \"{m_label}\")  {status:>7}")

# Inflation rate check
print()
av = ws_a.cell(48, 4).value  # subtotal
infl = ws_a.cell(49, 4).value  # inflation amount
if isinstance(av, (int,float)) and isinstance(infl, (int,float)):
    pct = infl / av
    print(f"  Inflation rate (auto): {pct*100:.2f}% (amount={infl:.2f} / subtotal={av:.2f})")
mv = ws_m.cell(47, 4).value
m_infl = ws_m.cell(48, 4).value
if isinstance(mv, (int,float)) and isinstance(m_infl, (int,float)):
    pct = m_infl / mv
    print(f"  Inflation rate (manual): {pct*100:.2f}% (amount={m_infl:.2f} / subtotal={mv:.2f})")

# Formula check
print()
print(f"Auto has formulas: {any(isinstance(ws_a.cell(r,c).value, str) and ws_a.cell(r,c).value.startswith('=') for r in range(1,60) for c in range(1,10))}")
print(f"Manual has formulas: {any(isinstance(ws_m.cell(r,c).value, str) and ws_m.cell(r,c).value.startswith('=') for r in range(1,60) for c in range(1,10))}")

# Consistency checks
print()
print('Consistency checks:')
av_s = ws_a.cell(48, 4).value  # subtotal
av_t = ws_a.cell(51, 4).value  # total
av_i = ws_a.cell(49, 4).value  # inflation
if isinstance(av_s, (int,float)) and isinstance(av_t, (int,float)) and isinstance(av_i, (int,float)):
    print(f"  Subtotal + Inflacao = Total: {av_s:.2f} + {av_i:.2f} = {av_s+av_i:.2f} vs Total={av_t:.2f} {'OK' if abs(av_s+av_i-av_t) < 0.01 else 'DIVERGE'}")
    print(f"  D/12 = F check (varies by row)")

# Check column E (if num_frac applied)
print(f"  num_frac default = 1 (col E == col D for all)")
for nn, a in auto_desp.items():
    d = a.get('D')
    e = a.get('E')
    if isinstance(d, (int,float)) and isinstance(e, (int,float)):
        if abs(d/12 - e) > 0.05:
            print(f"    {a['nome']}: D/12={d/12:.2f} vs E={e:.2f}   NOTE: E might use num_frac")
        if abs(e - d) < 0.01:
            pass  # num_frac=1, E same as D
